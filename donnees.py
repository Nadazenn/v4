import pandas as pd
import shutil
import os
import math
import joblib
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image
import entrainement_modele as entmod
import database as daba
import sqlite3
from PIL import Image as PILImage




# Fonction pour détecter et standardiser les colonnes
def standardize_columns(df):
    column_map = {
        "Désignation": ["désignation", "description", "nom", "désignation matérielle", "libellé"],
        "Unité": ["uté", "unité", "unit", "unité de mesure", "Unités ENT"],
        "Quantité": ["qté", "qté totale", "quantité", "quantité totale", "nombre", "Qté ENT"],
        "Prix unitaire": ["Prix Unitaire", "PU","Prix en € (ht)", "Prix", "unitaire", "Prix unit"],
    }
    standardized_columns = {}
    for standard_name, variants in column_map.items():
        for variant in variants:
            matches = [
                col for col in df.columns if isinstance(col, str) and variant.lower() in col.lower()
            ]
            if matches:
                standardized_columns[matches[0]] = standard_name
                break
    df = df.rename(columns=standardized_columns)
    required_columns = list(column_map.keys())
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(
            f"Colonnes manquantes : {', '.join(missing_columns)}. Colonnes détectées : {list(df.columns)} "
        )
    return df 

# Lecture du fichier principal avec une détection robuste des en-têtes
def preprocess_excel(file):
    try:
        # Lire le fichier Excel sans définir de ligne d'en-tête
        materiels_df = pd.read_excel(file, engine='openpyxl', header=None)

        # Détection des en-têtes dynamiquement
        column_map = {
            "Désignation": ["désignation", "description", "nom", "désignation matérielle", "libellé"],
            "Unité": ["uté", "unité", "unit", "unité de mesure", "Unités ENT"],
            "Quantité": ["qté", "qté totale", "quantité", "quantité totale", "nombre", "Qté ENT"],
            "Prix unitaire": ["Prix Unitaire", "PU","Prix en € (ht)", "Prix", "unitaire", "Prix unit"],
        }

        header_row = None
        for i, row in materiels_df.iterrows():
            row_lower = row.astype(str).str.lower()
            match_count = 0
            for variants in column_map.values():
                if any(row_lower.str.contains(variant, na=False).any() for variant in variants):
                    match_count += 1
            if match_count >= 2:  # Minimum 2 colonnes correspondant aux mots-clés
                header_row = i
                break

        if header_row is None:
            print("Aucune ligne d'en-tête détectée.")
            print("Contenu des premières lignes :")
            print(materiels_df.head(10))
            raise ValueError(
                "Aucune ligne d'en-tête détectée. Vérifiez les mots-clés ('Désignation', 'Unité', 'Quantité','Prix unitaire') dans vos données ou utilisez le fichier [template_outilIA.xlsx](https://huggingface.co/spaces/Maxfrem/classification/resolve/main/templates/template_outilIA.xlsx) et copier les données de votre bordereau."
            )

        # Définir les en-têtes à partir de la ligne détectée
        materiels_df.columns = materiels_df.iloc[header_row]
        materiels_df = materiels_df.drop(index=range(header_row + 1))
        materiels_df = materiels_df.reset_index(drop=True)

        print("Colonnes détectées après redéfinition :")
        print(materiels_df.columns)

        # Standardiser les colonnes
        materiels_df = standardize_columns(materiels_df)

        # Nettoyer et filtrer les données
        materiels_df = materiels_df.dropna(subset=['Désignation','Quantité','Unité'])
        materiels_df['Quantité'] = pd.to_numeric(materiels_df['Quantité'], errors='coerce').fillna(0)
        materiels_df['Prix unitaire'] = pd.to_numeric(materiels_df['Prix unitaire'], errors='coerce').fillna(0)
        materiels_df = materiels_df[materiels_df['Quantité'] > 0]
        materiels_df['Désignation'] = materiels_df['Désignation'].fillna('')
        if materiels_df.empty:
            raise ValueError("Aucune ligne exploitable après nettoyage.")
        return materiels_df


    except ValueError as ve:
        print(f"Erreur lors du prétraitement du fichier Excel : {ve}")
        raise ve
    except Exception as e:
        raise ValueError(f"Erreur inattendue lors du traitement du fichier Excel : {e}")


# Étape 1 : Générer le fichier temporaire
def process_file(model_choice, file):
    model_name = f"{model_choice}.pkl"
    try:
        # Charger le modèle
        global model
        model_path = os.path.join("models", model_name)
        model, vectorizer, label_encoder = joblib.load(model_path)
        
        # Lecture et prétraitement du fichier principal
        materiels_df = preprocess_excel(file)

        # Transformer les données en vecteurs
        X_new = materiels_df["Désignation"].astype(str)
        X_transformed = vectorizer.transform(X_new)
    
        # Prédictions
        predictions = model.predict(X_transformed)
        probabilities = model.predict_proba(X_transformed)
    
        # Taux de confiance

        max_confidences = np.max(probabilities, axis=1)

        # Ajout des résultats : gérer tous les cas de classes
        try:
            if np.array_equal(getattr(model, "classes_", None), np.arange(len(label_encoder.classes_))):
                cats = label_encoder.inverse_transform(predictions)
            elif np.array_equal(getattr(model, "classes_", None), label_encoder.classes_):
                cats = predictions.astype(str)
            else:
                classes_model = getattr(model, "classes_", None)
                if classes_model is not None and classes_model.dtype.kind in "iu":
                    mapping = {cls: lab for cls, lab in zip(classes_model, label_encoder.classes_)}
                    cats = np.array([mapping.get(p, "Inconnu") for p in predictions])
                else:
                    cats = predictions.astype(str)
        except Exception:
            cats = predictions.astype(str)

        materiels_df["Catégorie Prédite"] = cats
        materiels_df["Taux de Confiance"] = max_confidences

        # Colonnes utiles
        output_df = materiels_df[
            ["Désignation", "Catégorie Prédite", "Taux de Confiance", "Quantité", "Prix unitaire"]
        ].sort_values(by="Taux de Confiance", ascending=True)
        
        # Sauvegarde Excel
        output_file_path2 = os.path.join("sortie", "Bordereau classé et non regroupé.xlsx")
        output_df.to_excel(output_file_path2, index=False)

        # Mise en forme
        wb = load_workbook(output_file_path2)
        ws = wb.active
        
        # Largeur colonnes auto
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
        
        ws.auto_filter.ref = ws.dimensions
        wb.save(output_file_path2)

        # ✅ Retourne message + chemin du fichier généré
        return "✅ Bordereau classé généré avec succès.", output_file_path2

    except Exception as e:
        return f"Erreur pendant le traitement : {e}", None


# Étape 2 : Valider et créer le fichier final
def finalize_file(file, model_choice, nombre_etages, duree_stockage, tarif_stockage, frais_palette, frais_livraison, output_table, output_details_table, choix_camions, choix_conditionnement, entreprise_choice,planning_indice,use_ccc):
    try:


        # Lire les données du fichier Excel source
        materiels_df = pd.read_excel(file, engine='openpyxl')

        # Regrouper par catégorie
        grouped_df = materiels_df.groupby('Catégorie Prédite')['Quantité'].sum().reset_index()

        grouped_df = daba.ajouter_supportage(grouped_df, model_choice)

        # Définir le chemin du fichier original 'outil.xlsm'
        original_file_path  = os.path.join("templates", 'outil.xlsm')

        output_file_path = os.path.join("sortie", 'Outil Logistique - Vinci Energies.xlsm')

        # Copier le fichier
        shutil.copy(original_file_path, output_file_path)

        # Ouvrir le fichier temporaire avec openpyxl pour insérer les données
        wb = load_workbook(output_file_path, keep_vba=True)  

        darkBlue = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        lihtBlue = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        white = Font(color="FFFFFF", bold=True)
        
        # Feuille Données
        
        ws = wb['Données']

        nbzones = output_details_table.shape[0]
        
        # Insérer les données dans les colonnes B et C à partir de la ligne 4
        for i, (category, quantity) in enumerate(zip(grouped_df['Catégorie Prédite'], grouped_df['Quantité']), start=4):
            ws[f'B{i}'] = category
            ws[f'C{i}'] = quantity



        # Aligner le planning sur le format v3 pour l'écriture Excel
        details_v3 = output_details_table.copy()
        numero_col = None
        if "Numero etage (pas de lettres)" in details_v3.columns:
            numero_col = "Numero etage (pas de lettres)"
            details_v3 = details_v3.drop(columns=[numero_col])
        if "Numéro étage (pas de lettres)" in details_v3.columns:
            numero_col = "Numéro étage (pas de lettres)"
            details_v3 = details_v3.drop(columns=[numero_col])

        expected_cols = [
            "Étage",
            "Zone",
            "Nom Zone",
            "Date début phase production",
            "Date début phase terminaux",
            "Délai de livraison avant travaux (jours)",
            "Durée travaux production",
            "Durée travaux terminaux",
        ]
        if all(c in details_v3.columns for c in expected_cols):
            details_v3 = details_v3[expected_cols]
            if numero_col is not None:
                details_v3["Étage"] = output_details_table[numero_col].values

        # En-têtes "Étage" et "Zone" : même logique que v3 (col 0 et 2)
        for i, col_idx in enumerate([0, 2]):
            for j, value in enumerate(details_v3.iloc[:, col_idx]):
                cell = ws.cell(row=1 + i, column=5 + j, value=value)
                if col_idx == 0:
                    cell.fill = darkBlue
                    cell.font = white
                else:
                    cell.fill = lihtBlue
        
        cell = ws.cell(row=1, column=5 + nbzones, value="Exclus")
        cell.fill = darkBlue
        cell.font = white
        cell = ws.cell(row=1, column=6 + nbzones, value="Commentaires")
        cell.fill = darkBlue
        cell.font = white
        

        # Ajuster la largeur des colonnes
        max_length = 0
        for cell in ws['B']:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws.column_dimensions['B'].width = adjusted_width

        # Ajout de la formule dans la colonne D
        for i in range(4, 4 + len(grouped_df)):  # Boucle sur les lignes des données
            last_zone_col_letter = ws.cell(row=1, column=5 + nbzones).column_letter  # Lettre de la dernière colonne des zones
            ws.cell(row=i, column=4, value=f"=C{i}-SUM(E{i}:{last_zone_col_letter}{i})")
        
        # Faire une répartition proportionnelle des matériels dans chaque zone
        for i, (category, quantity) in enumerate(zip(grouped_df['Catégorie Prédite'], grouped_df['Quantité']), start=4):
            # Si la catégorie est "Exclus", on ne fait pas de répartition, mais on met la quantité totale dans la colonne "Exclus"
            if category == "Exclus":
                # Mettre la quantité totale dans la colonne "Exclus" (dernière colonne)
                ws.cell(row=i, column=5 + nbzones, value=quantity)
            else:
                # Calcul de la répartition proportionnelle par zone
                quantity_per_zone = [math.floor(quantity / nbzones)] * nbzones
                
                # Calcul du reste (si la division ne donne pas un entier exact)
                remainder = quantity - sum(quantity_per_zone)
                
                # Distribuer le reste dans la première zone (colonne E)
                if remainder > 0:
                    quantity_per_zone[0] += remainder
                
                # Remplir les colonnes correspondantes pour chaque zone
                for j, qty in enumerate(quantity_per_zone):
                    ws.cell(row=i, column=5 + j, value=qty)
            
        # Feuille paramétrage

        ws = wb['Paramétrage']

        ws["B1"] = model_choice
        ws["B2"] = nombre_etages
        ws["B4"] = duree_stockage
        ws["B5"] = tarif_stockage
        ws["B6"] = frais_palette
        ws["B7"] = frais_livraison
        ws["D1"] = planning_indice

        # Coller les données du premier tableau
        output_table['Numéro étage (pas de lettres)'] = pd.to_numeric(output_table['Numéro étage (pas de lettres)'], errors='coerce').fillna(0).astype(int)
        output_table['Nombre de zones'] = pd.to_numeric(output_table['Nombre de zones'], errors='coerce').fillna(0).astype(int)
        for row_num, row in enumerate(output_table.itertuples(index=False), start=10):
            for col_num, value in enumerate(row, start=1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.fill = lihtBlue

        # Coller les données du deuxième tableau
        details_v3['Délai de livraison avant travaux (jours)'] = pd.to_numeric(details_v3['Délai de livraison avant travaux (jours)'], errors='coerce').fillna(0).astype(int)
        for row_num, row in enumerate(details_v3.itertuples(index=False), start=3):
            for col_num, value in enumerate(row, start=5):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.fill = lihtBlue

                # Vérifier si la colonne correspond aux dates (colonnes 8 et 9 dans Excel format v3)
                if col_num in [8, 9] and isinstance(value, str):  
                    # Convertir le texte en date
                    date_value = pd.to_datetime(value, dayfirst=True).to_pydatetime()  # Convertir en datetime Python
                    cell.value = date_value  # Insère une vraie date Excel
                    cell.number_format = 'DD/MM/YYYY'  # Applique le format de date Excel
        
        ws = wb['Matériel']
        
        # Connexion à la base de données SQLite
        conn = sqlite3.connect('logistique.db')
        cursor = conn.cursor()

        # Exécution de la requête pour récupérer toutes les données de la table matériel
        cursor.execute("SELECT * FROM materiel WHERE lot = ?", (model_choice,))
        rows = cursor.fetchall()

        # Supprimer la première colonne (clé primaire) de chaque ligne
        rows_without_id = [row[1:] for row in rows]  # On supprime l'élément à l'index 0 (clé primaire)        
        
        # Copier les données dans la feuille Matériel à partir de la ligne 2
        for row_index, row in enumerate(rows_without_id, start=2):
            for col_index, value in enumerate(row, start=1):
                 # Vérification des colonnes de conditionnement (6, 8, 10)
                if col_index == 7:  # Colonne 7: Conditionnement 1
                    if value not in choix_conditionnement:
                        continue  # Si conditionnement non choisi, on saute cette colonne et la suivante (8)
                elif col_index == 8:  # Colonne 8: Quantité pour Conditionnement 1
                    if row[6] not in choix_conditionnement:  # Vérification pour la colonne 8
                        continue  # Si conditionnement de la colonne 7 non choisi, on saute cette colonne
                elif col_index == 9:  # Colonne 9: Conditionnement 2
                    if value not in choix_conditionnement:
                        continue  # Si conditionnement non choisi, on saute cette colonne et la suivante (10)
                elif col_index == 10:  # Colonne 10: Quantité pour Conditionnement 2
                    if row[8] not in choix_conditionnement:  # Vérification pour la colonne 9
                        continue  # Si conditionnement de la colonne 9 non choisi, on saute cette colonne
                elif col_index == 11:  # Colonne 11: Conditionnement 3
                    if value not in choix_conditionnement:
                        continue  # Si conditionnement non choisi, on saute cette colonne et la suivante (12)
                elif col_index == 12:  # Colonne 12: Quantité pour Conditionnement 3
                    if row[10] not in choix_conditionnement:  # Vérification pour la colonne 11
                        continue  # Si conditionnement de la colonne 11 non choisi, on saute cette colonne

                # Remplir la cellule si elle n'est pas exclue par le filtre
                ws.cell(row=row_index, column=col_index, value=value)   
       

        ws = wb['Camion']

        # Vérifier si la liste n'est pas vide
        if choix_camions:
            # Création d'une requête SQL dynamique pour gérer plusieurs noms
            placeholders = ", ".join(["?"] * len(choix_camions))  # Génère "?, ?, ?" selon la taille de choix_camions
            query = f"SELECT * FROM camion WHERE nom IN ({placeholders})"

            # Exécution de la requête en passant la liste sous forme de tuple
            cursor.execute(query, tuple(choix_camions))
            rows = cursor.fetchall() 

            # Supprimer la première colonne (clé primaire) de chaque ligne
            rows_without_id = [row[1:] for row in rows]

            # Copier les données dans la feuille Camion à partir de la ligne 2
            for row_index, row in enumerate(rows_without_id, start=2):
                for col_index, value in enumerate(row, start=1):
                    ws.cell(row=row_index, column=col_index, value=value)
        
        ws = wb['Manutention']

        # Vérifier si la liste n'est pas vide
        if choix_conditionnement:
            # Création d'une requête SQL dynamique pour gérer plusieurs noms
            placeholders = ", ".join(["?"] * len(choix_conditionnement))  # Génère "?, ?, ?" selon la taille de choix_conditionnement
            query = f"SELECT * FROM conditionnement WHERE nom IN ({placeholders})"

            # Exécution de la requête en passant la liste sous forme de tuple
            cursor.execute(query, tuple(choix_conditionnement))
            rows = cursor.fetchall() 

            # Supprimer la première colonne (clé primaire) de chaque ligne
            rows_without_id = [row[1:] for row in rows]

            # Copier les données dans la feuille Conditionnement à partir de la ligne 2
            for row_index, row in enumerate(rows_without_id, start=2):
                for col_index, value in enumerate(row, start=1):
                    ws.cell(row=row_index, column=col_index, value=value)

        # Fermer la connexion à la base de données
        conn.close()   

        ws = wb['Livrable']
        
        # Obtenir les dimensions des cellules fusionnées (A1:A4)
        column_width_units = ws.column_dimensions['A'].width  # Largeur en unités Excel
        row_heights_units = [ws.row_dimensions[i].height or 15 for i in range(1, 5)]  # Hauteurs en points

        # Convertir en pixels
        cell_width = int(column_width_units * 7)  # Largeur en pixels (1 unité ≈ 7 pixels)
        cell_height = int(sum(row_heights_units) * 1.25)  # Hauteur totale en pixels (1 point ≈ 1,25 pixel)


        # Construire dynamiquement le chemin de l'image
        image_path = f"images/logos_entreprises/{entreprise_choice}_logo.png"   

        # Charger l'image avec PIL pour la redimensionner
        original_image = PILImage.open(image_path)
        
        # Redimensionner l'image en conservant le ratio
        original_ratio = original_image.width / original_image.height
        target_ratio = cell_width / cell_height

        if original_ratio > target_ratio:
            # L'image est plus large : ajuster la largeur
            new_width = int(cell_width)
            new_height = int(cell_width / original_ratio)
        else:
            # L'image est plus haute : ajuster la hauteur
            new_height = int(cell_height)
            new_width = int(cell_height * original_ratio)

        resized_image = original_image.resize((new_width, new_height), PILImage.Resampling.LANCZOS)


        # Sauvegarder l'image redimensionnée temporairement
        temp_image_path = "sortie/temp_resized_logo.png"
        resized_image.save(temp_image_path)

        img1 = Image(temp_image_path)
        img2 = Image(temp_image_path)
        img3 = Image(temp_image_path)


        # Définir la position de l'image 
        ws.add_image(img1, "A1")
        ws.add_image(img2, "I1")
        ws.add_image(img3, "Q1")

        # Appliquer le formatage pour la première ligne (police Calibri, Gras, taille 18, centré)
        font_ligne1 = Font(name="Calibri", bold=True, size=18)
        alignment1 = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

        # Appliquer le formatage pour la deuxième ligne (police Calibri, Gras, Italique, taille 16, centré)
        font_ligne2 = Font(name="Calibri", bold=True, italic=True, size=14)
        alignment2 = Alignment(horizontal="center", vertical="top", wrap_text=True)

        # Mettre le texte dans deux cellules différentes
        ws['B1'] = f"Etude Logistique - Lot {model_choice}"  # Première ligne
        ws['J1'] = f"Etude Logistique - Lot {model_choice}"
        ws['R1'] = f"Etude Logistique - Lot {model_choice}"
        ws['B3'] = entreprise_choice  # Deuxième ligne
        ws['J3'] = entreprise_choice
        ws['R3'] = entreprise_choice  

        # Appliquer le formatage pour la première ligne (B1)
        ws['B1'].font = font_ligne1
        ws['B1'].alignment = alignment1
        ws['J1'].font = font_ligne1
        ws['J1'].alignment = alignment1
        ws['R1'].font = font_ligne1
        ws['R1'].alignment = alignment1

        # Appliquer le formatage pour la deuxième ligne (B2)
        ws['B3'].font = font_ligne2
        ws['B3'].alignment = alignment2
        ws['J3'].font = font_ligne2
        ws['J3'].alignment = alignment2
        ws['R3'].font = font_ligne2
        ws['R3'].alignment = alignment2

        # Fusionner les cellules d'en-têtes
        ws.merge_cells('B1:G2')
        ws.merge_cells('B3:G4')
        ws.merge_cells('J1:O2')
        ws.merge_cells('J3:O4')
        ws.merge_cells('R1:W2')
        ws.merge_cells('R3:W4')


        # Charger l'image
        image_camions_path = "images/camions_livrable.png"  # Remplacez par votre chemin d'image
        img_camions1 = Image(image_camions_path)

        # Définir la cellule cible
        cellule = "F23"
        cell = ws[cellule]

        # Convertir les dimensions (mm -> pixels)
        largeur_pixels = 20.84 * 3.78
        hauteur_pixels = 50.91 * 3.78

        # Appliquer les dimensions
        img_camions1.width = largeur_pixels
        img_camions1.height = hauteur_pixels

        # Positionner l'image dans la cellule F23
        img_camions1.anchor = cell.coordinate  # Ancrer l'image dans la cellule

        # Ajouter l'image à la feuille
        ws.add_image(img_camions1)

        img_camions2 = Image(image_camions_path)

        # Définir la cellule cible
        cellule = "N23"
        cell = ws[cellule]

        # Convertir les dimensions (mm -> pixels)
        largeur_pixels = 20.84 * 3.78
        hauteur_pixels = 50.91 * 3.78

        # Appliquer les dimensions
        img_camions2.width = largeur_pixels
        img_camions2.height = hauteur_pixels

        # Positionner l'image dans la cellule N23
        img_camions2.anchor = cell.coordinate  # Ancrer l'image dans la cellule

        # Ajouter l'image à la feuille
        ws.add_image(img_camions2)



        # Sauvegarder le fichier modifié
        wb.save(output_file_path)
        wb.close()
        
        # Retourner le chemin du fichier temporaire
        return f"Le fichier final a été créé avec succès.", output_file_path

    except Exception as e:
        return f"Erreur pendant l'insertion des données : {e}", None

# Lien entre le bouton de validation et la sortie
def finalize_wrapper(bordereau_table, model_choice, nombre_etages, duree_stockage, tarif_stockage,
                     frais_palette, frais_livraison, output_table, output_details_table,
                     choix_camions, choix_conditionnement, entreprise_choice,planning_indice,use_ccc):
    # Sauvegarder le tableau édité en Excel temporaire
    temp_path = os.path.join("sortie", "Bordereau_corrigé.xlsx")
    bordereau_table.to_excel(temp_path, index=False)

    # Appeler le pipeline existant pour générer l'outil final
    return finalize_file(temp_path, model_choice, nombre_etages, duree_stockage, tarif_stockage,
                         frais_palette, frais_livraison, output_table, output_details_table,
                         choix_camions, choix_conditionnement, entreprise_choice,planning_indice,use_ccc)

