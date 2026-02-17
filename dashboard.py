import io
import math
import pandas as pd
import plotly.express as px
import streamlit as st
import unicodedata
import database as daba
from openpyxl import load_workbook


def render_dashboard_excel():
    
    st.header("📊 Études logistiques ")
    use_ccc = st.session_state.get("parametrage", {}).get("use_ccc", False)

    #
    # 1) Chargement du fichier Excel final
    #
    file_state = st.session_state.get("pilotage_file", None)
    file_bytes = None
    if isinstance(file_state, bytes):
        file_bytes = file_state
    elif isinstance(file_state, str):
        try:
            with open(file_state, "rb") as f:
                file_bytes = f.read()
        except PermissionError:
            st.warning("Le fichier Excel est ouvert. Ferme-le ou charge une copie.")
    if file_bytes is None:
        uploaded = st.file_uploader(
            "Dépose ici une copie du fichier final (xlsm/xlsx)",
            type=["xlsm", "xlsx"]
        )
        if uploaded is None:
            st.stop()
        file_bytes = uploaded.read()
    excel_io = io.BytesIO(file_bytes)

    #
    # 2) Lecture des feuilles nécessaires
    #
    try:
        xls = pd.ExcelFile(excel_io)
        bg = xls.parse("Bilan Graphique")
        param = xls.parse("Paramétrage")
        materiel = xls.parse("Matériel")
        src = xls.parse("Tableau Source")
    except Exception as e:
        st.error(f"Erreur lecture Excel : {e}")
        st.stop()

    #
    # 3) Pipelines unifiés
    #    - pipeline_sans_ccc : traitement V0
    #    - pipeline_avec_ccc : traitement V1
    #

    def pipeline_sans_ccc(bg_df: pd.DataFrame) -> dict:
        # Palettes par étage / zone
        try:
            palettes_zone = (
                bg_df[["Étage - Zone", "Production", "Terminaux"]]
                .dropna(subset=["Étage - Zone"])
                .copy()
            )
            palettes_zone["Production"] = palettes_zone["Production"].fillna(0)
            palettes_zone["Terminaux"] = palettes_zone["Terminaux"].fillna(0)
            palettes_zone["Palettes"] = (
                palettes_zone["Production"] + palettes_zone["Terminaux"]
            )
        except Exception:
            palettes_zone = pd.DataFrame(columns=["Étage - Zone", "Palettes"])
        total_palettes = float(palettes_zone["Palettes"].sum()) if not palettes_zone.empty else 0.0

        # Flux mensuel palettes
        if {"Mois", "Volume (nombre de palettes équivalentes)"}.issubset(bg_df.columns):
            flux_palettes = (
                bg_df[["Mois", "Volume (nombre de palettes équivalentes)"]]
                .dropna(subset=["Mois"])
                .groupby("Mois", as_index=False)
                .sum()
            )
            if not flux_palettes.empty:
                idx_pic_pal = flux_palettes["Volume (nombre de palettes équivalentes)"].idxmax()
                mois_pic_palettes = flux_palettes.loc[idx_pic_pal, "Mois"]
                pic_palettes = float(
                    flux_palettes.loc[idx_pic_pal, "Volume (nombre de palettes équivalentes)"]
                )
            else:
                mois_pic_palettes = ""
                pic_palettes = 0.0
        else:
            flux_palettes = pd.DataFrame(columns=["Mois", "Volume (nombre de palettes équivalentes)"])
            mois_pic_palettes = ""
            pic_palettes = 0.0

        # Camions par étage / zone (sans CCC)
        if {
            "Étage - Zone",
            "Camions Production sans CCC",
            "Camions Terminaux sans CCC",
        }.issubset(bg_df.columns):
            camions_zone = (
                bg_df[["Étage - Zone", "Camions Production sans CCC", "Camions Terminaux sans CCC"]]
                .dropna(subset=["Étage - Zone"])
                .copy()
            )
            camions_zone["Camions Production sans CCC"] = camions_zone[
                "Camions Production sans CCC"
            ].fillna(0)
            camions_zone["Camions Terminaux sans CCC"] = camions_zone[
                "Camions Terminaux sans CCC"
            ].fillna(0)
            camions_zone["Camions totaux"] = (
                camions_zone["Camions Production sans CCC"]
                + camions_zone["Camions Terminaux sans CCC"]
            )
        else:
            camions_zone = pd.DataFrame(columns=["Étage - Zone", "Camions totaux"])

        # Total camions
        if "Nombre de Camions" in bg_df.columns:
            total_camions = float(bg_df["Nombre de Camions"].fillna(0).sum())
        else:
            total_camions = float(camions_zone["Camions totaux"].sum()) if not camions_zone.empty else 0.0

        # Remplissage par zone + remplissage moyen (sans CCC)
        if "Remplissage camions sans CCC" in bg_df.columns:
            rempl_zone = (
                bg_df[["Étage - Zone", "Remplissage camions sans CCC"]]
                .dropna(subset=["Étage - Zone"])
                .copy()
            )
            rempl_zone["Remplissage (%)"] = (
                rempl_zone["Remplissage camions sans CCC"].fillna(0)
            )
            rempl_brut = bg_df["Remplissage camions sans CCC"].dropna()
            rempl_moyen = float(rempl_brut.mean()) if not rempl_brut.empty else 0.0
        else:
            rempl_zone = pd.DataFrame(columns=["Étage - Zone", "Remplissage (%)"])
            rempl_moyen = 0.0

        # Flux mensuel camions
        if "Nombre de Camions" in bg_df.columns and "Mois" in bg_df.columns:
            flux_camions = (
                bg_df[["Mois", "Nombre de Camions"]]
                .dropna(subset=["Mois"])
                .groupby("Mois", as_index=False)
                .sum()
            )
            if not flux_camions.empty:
                idx_pic_cam = flux_camions["Nombre de Camions"].idxmax()
                mois_pic_camions = flux_camions.loc[idx_pic_cam, "Mois"]
                pic_camions = float(flux_camions.loc[idx_pic_cam, "Nombre de Camions"])
            else:
                mois_pic_camions = ""
                pic_camions = 0.0
        else:
            flux_camions = pd.DataFrame(columns=["Mois", "Nombre de Camions"])
            mois_pic_camions = ""
            pic_camions = 0.0

        # Camions par type
        cols = list(bg_df.columns)
        col_etage_type = None
        for c in cols:
            if c.startswith("Étage") and c != "Étage - Zone":
                col_etage_type = c
                break
        if col_etage_type:
            idx_etage_type = cols.index(col_etage_type)
            col_zone_type = cols[idx_etage_type + 1] if idx_etage_type + 1 < len(cols) else None
            col_type_camion = cols[idx_etage_type + 2] if idx_etage_type + 2 < len(cols) else None
            col_nb_camions_type = cols[idx_etage_type + 3] if idx_etage_type + 3 < len(cols) else None
            if col_type_camion and col_nb_camions_type:
                camions_type = (
                    bg_df[[col_type_camion, col_nb_camions_type]]
                    .dropna(subset=[col_type_camion])
                    .copy()
                )
                camions_type = (
                    camions_type
                    .groupby(col_type_camion, as_index=False)[col_nb_camions_type]
                    .sum()
                )
                camions_type.rename(
                    columns={
                        col_type_camion: "Type de Camion",
                        col_nb_camions_type: "Nombre de Camions",
                    },
                    inplace=True,
                )
            else:
                camions_type = pd.DataFrame(columns=["Type de Camion", "Nombre de Camions"])
        else:
            camions_type = pd.DataFrame(columns=["Type de Camion", "Nombre de Camions"])

        return {
            "palettes_zone": palettes_zone,
            "total_palettes": total_palettes,
            "flux_palettes": flux_palettes,
            "mois_pic_palettes": mois_pic_palettes,
            "pic_palettes": pic_palettes,
            "camions_zone": camions_zone,
            "total_camions": total_camions,
            "flux_camions": flux_camions,
            "mois_pic_camions": mois_pic_camions,
            "pic_camions": pic_camions,
            "rempl_zone": rempl_zone,
            "rempl_moyen": rempl_moyen,
            "camions_type": camions_type,
        }

    def pipeline_avec_ccc(bg_df: pd.DataFrame) -> dict:
        # On réutilise la partie palettes du pipeline sans CCC
        base = pipeline_sans_ccc(bg_df)

        # Flux mensuel camions CCC
        if "Nombre de Camions CCC" in bg_df.columns and "Mois" in bg_df.columns:
            flux_camions_ccc = (
                bg_df[["Mois", "Nombre de Camions CCC"]]
                .dropna(subset=["Mois"])
                .groupby("Mois", as_index=False)
                .sum()
            )
            if not flux_camions_ccc.empty:
                idx_pic_cam_v1 = flux_camions_ccc["Nombre de Camions CCC"].idxmax()
                mois_pic_camions_v1 = flux_camions_ccc.loc[idx_pic_cam_v1, "Mois"]
                pic_camions_v1 = float(
                    flux_camions_ccc.loc[idx_pic_cam_v1, "Nombre de Camions CCC"]
                )
            else:
                mois_pic_camions_v1 = ""
                pic_camions_v1 = 0.0
        else:
            flux_camions_ccc = pd.DataFrame(columns=["Mois", "Nombre de Camions CCC"])
            mois_pic_camions_v1 = ""
            pic_camions_v1 = 0.0

        # Indicateurs globaux CCC
        if "Nombre de Camions CCC" in bg_df.columns:
            total_camions_ccc = float(bg_df["Nombre de Camions CCC"].fillna(0).sum())
        else:
            total_camions_ccc = 0.0

        if "Remplissage camions avec CCC" in bg_df.columns:
            rempl_brut_ccc = bg_df["Remplissage camions avec CCC"].dropna()
            rempl_moyen_ccc = float(rempl_brut_ccc.mean()) if not rempl_brut_ccc.empty else 0.0
            rempl_zone_ccc = (
                bg_df[["Étage - Zone", "Remplissage camions avec CCC"]]
                .dropna(subset=["Étage - Zone"])
                .copy()
            )
            rempl_zone_ccc["Remplissage (%)"] = (
                rempl_zone_ccc["Remplissage camions avec CCC"]
            )
        else:
            rempl_moyen_ccc = 0.0
            rempl_zone_ccc = pd.DataFrame(columns=["Étage - Zone", "Remplissage (%)"])

        # Camions par étage / zone CCC
        if {
            "Étage - Zone",
            "Camions Production avec CCC",
            "Camions Terminaux avec CCC",
        }.issubset(bg_df.columns):
            camions_zone_ccc = (
                bg_df[
                    [
                        "Étage - Zone",
                        "Camions Production avec CCC",
                        "Camions Terminaux avec CCC",
                    ]
                ]
                .dropna(subset=["Étage - Zone"])
                .copy()
            )
            camions_zone_ccc["Total CCC"] = (
                camions_zone_ccc["Camions Production avec CCC"].fillna(0)
                + camions_zone_ccc["Camions Terminaux avec CCC"].fillna(0)
            )
        else:
            camions_zone_ccc = pd.DataFrame(columns=["Étage - Zone", "Total CCC"])

        # On renvoie des noms neutres pour la comparaison
        return {
            "palettes_zone": base["palettes_zone"],
            "total_palettes": base["total_palettes"],
            "flux_palettes": base["flux_palettes"],
            "mois_pic_palettes": base["mois_pic_palettes"],
            "pic_palettes": base["pic_palettes"],
            "camions_zone": camions_zone_ccc.rename(columns={"Total CCC": "Camions"}) if not camions_zone_ccc.empty else pd.DataFrame(columns=["Étage - Zone", "Camions"]),
            "total_camions": total_camions_ccc,
            "flux_camions": flux_camions_ccc.rename(columns={"Nombre de Camions CCC": "Camions"}) if not flux_camions_ccc.empty else pd.DataFrame(columns=["Mois", "Camions"]),
            "mois_pic_camions": mois_pic_camions_v1,
            "pic_camions": pic_camions_v1,
            "rempl_zone": rempl_zone_ccc,
            "rempl_moyen": rempl_moyen_ccc,
            "camions_type": base["camions_type"],  # typologie identique
        }

    def _clean_ccc_familles(df: pd.DataFrame, col: str = "Famille") -> pd.DataFrame:
        exclude = {"stock ccc production", "stock ccc terminaux"}
        if col not in df.columns:
            return df
        mask = (
            df[col]
            .astype(str)
            .str.strip()
            .str.lower()
            .isin(exclude)
        )
        return df.loc[~mask].copy()

    def _df_height(n_rows: int, row_h: int = 32, min_h: int = 180, max_h: int = 700) -> int:
        return max(min_h, min(max_h, (n_rows + 1) * row_h))

    #
    # 3 bis) Préparation des données de base via pipelines
    #

    metrics_v0 = pipeline_sans_ccc(bg)
    palettes_zone = metrics_v0["palettes_zone"]
    total_palettes = metrics_v0["total_palettes"]
    flux_palettes = metrics_v0["flux_palettes"]
    mois_pic_palettes = metrics_v0["mois_pic_palettes"]
    pic_palettes = metrics_v0["pic_palettes"]
    camions_zone = metrics_v0["camions_zone"]
    total_camions = metrics_v0["total_camions"]
    flux_camions = metrics_v0["flux_camions"]
    mois_pic_camions = metrics_v0["mois_pic_camions"]
    pic_camions = metrics_v0["pic_camions"]
    rempl_zone = metrics_v0["rempl_zone"]
    rempl_moyen = metrics_v0["rempl_moyen"]
    camions_type = metrics_v0["camions_type"]

    metrics_v1 = pipeline_avec_ccc(bg)
    total_camions_ccc = metrics_v1["total_camions"]
    flux_camions_ccc = metrics_v1["flux_camions"].rename(columns={"Camions": "Nombre de Camions CCC"}) if not metrics_v1["flux_camions"].empty else pd.DataFrame(columns=["Mois", "Nombre de Camions CCC"])
    mois_pic_camions_v1 = metrics_v1["mois_pic_camions"]
    pic_camions_v1 = metrics_v1["pic_camions"]
    rempl_zone_ccc = metrics_v1["rempl_zone"]
    rempl_moyen_ccc = metrics_v1["rempl_moyen"]
    camions_zone_ccc = metrics_v1["camions_zone"].rename(columns={"Camions": "Total CCC"}) if not metrics_v1["camions_zone"].empty else pd.DataFrame(columns=["Étage - Zone", "Total CCC"])

    # Hypothèses générales
    lot_col = param.columns[1]  # même logique que ta macro
    try:
        nb_etages = param.loc[param["Lot"] == "Nombre étage :", lot_col].iloc[0]
    except Exception:
        nb_etages = ""
    familles_identifiees = (
        materiel["Nom"].dropna().astype(str).sort_values().unique().tolist()
        if "Nom" in materiel.columns
        else []
    )

    # DPGF + indice planning depuis la session
    dpgf_date = st.session_state.get("dpgf_date", "")
    dpgf_date = st.text_input("Date du DPGF :", value=dpgf_date)
    st.session_state["dpgf_date"] = dpgf_date
    planning_indice = st.session_state.get("parametrage", {}).get("planning_indice", "")

    #
    # 4) Gestion des variantes (initialisation)
    #
    if "variants" not in st.session_state:
        st.session_state["variants"] = {}      # {"V2": {"with_ccc": bool, "bytes": ...}}
    if "variant_counter" not in st.session_state:
        st.session_state["variant_counter"] = 2

    # 4) Onglets V0 / V1 / Variantes / Comparatif
    tabs = []

    if use_ccc:
        tabs.append("V1")
    else:
        tabs.append("V0")

    tabs.extend(["Variantes", "Comparatif"])

    tab_objects = st.tabs(tabs)
    main_tab = tab_objects[0]
    tab_var = tab_objects[1]
    tab_comp = tab_objects[2]

    def _norm(s: str) -> str:
        v = "" if s is None else str(s)
        v = unicodedata.normalize("NFKD", v)
        v = "".join(c for c in v if not unicodedata.combining(c))
        return v.lower().strip()

    def _find_col(columns, target: str):
        t = _norm(target)
        for col in columns:
            if _norm(col) == t:
                return col
        return None

    def _coerce_quantite_cols(df: pd.DataFrame) -> pd.DataFrame:
        if df is None or df.empty:
            return df
        for col in df.columns:
            if _norm(col).startswith("quantit"):
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        return df

    dpgf_date_str = dpgf_date
    dpgf_indice = planning_indice if "planning_indice" in locals() else ""



    with main_tab:
        if use_ccc:


            try:
                src_v1 = src.copy()
            except Exception:
                src_v1 = pd.DataFrame()

            # 3 onglets internes : Hypothèses / Palettes / Camions
            ong_hyp_v1, ong_pal_v1, ong_cam_v1 = st.tabs(
                ["📘 Hypothèses", "📦 Palettes", "🚚 Camions"]
            )

            #
            # 📘 ONGLET HYPOTHÈSES V1
            #
            with ong_hyp_v1:
                st.markdown("### 📘 Hypothèses")

                h1, h2 = st.columns(2)
                with h1:
                    # DPGF + PIC
                    st.markdown("#### 📄 Document de source")

                    if dpgf_date and planning_indice:
                        default_dpgf_v1 = f"DPGF du {dpgf_date} – Indice {planning_indice}"
                    elif dpgf_date:
                        default_dpgf_v1 = f"DPGF du {dpgf_date}"
                    elif planning_indice:
                        default_dpgf_v1 = f"Indice {planning_indice}"
                    else:
                        default_dpgf_v1 = ""

                    dpgf_txt_v1 = st.text_area(
                        "DPGF + Indice :",
                        value=default_dpgf_v1,
                        key="dpgf_v1",
                        placeholder="DPGF du … – Indice …",
                    )

                    pic_file_v1 = st.file_uploader(
                        "Veuillez joindre le fichier PIC", key="pic_v1"
                    )

                with h2:
                    # Hypothèse planning
                    st.markdown("#### 🕒 Hypothèse planning")
                    st.markdown(f"- Planning indice : **{planning_indice or '…'}**")

                    planning_hyp_v1 = st.text_area(
                        "Hypothèses prises pour le planning :",
                        key="planning_hyp_v1",
                        placeholder="Décrire ici les hypothèses utilisées pour le planning…",
                    )

                h3, h4 = st.columns(2)
                with h3:
                    # Hypothèses de l’étude
                    st.markdown("#### 📄 Hypothèses de l’étude")
                    st.markdown("- regroupement du matériel en grandes catégories")
                    st.markdown("- conversion des conditionnements en équivalent palette")
                    st.markdown("- 2 phases de travaux par étage")

                with h4:
                    # Paramètres CCC
                    st.markdown("#### ⚙️ Paramètres CCC")

                    def _val_param_ccc(libel):
                        try:
                            return param.loc[param["Lot"] == libel, lot_col].iloc[0]
                        except:
                            return ""

                    duree_ccc = _val_param_ccc("Durée de stockage CCC (en mois)")
                    tarif_mois = _val_param_ccc("Tarif mois de stockage (en €)")
                    frais_sup = _val_param_ccc("Frais supplémentaires/palette (en €)")
                    frais_liv = _val_param_ccc("Frais de livraison par camion")

                    st.markdown(f"- Durée stockage : **{duree_ccc} mois**")
                    st.markdown(f"- Tarif de stockage : **{tarif_mois} €/mois**")
                    st.markdown(f"- Frais supplémentaires/palette : **{frais_sup} €**")
                    st.markdown(f"- Frais de livraison : **{frais_liv} €**")

                with st.container():
                    # Familles CCC (Tableau Source + BG)
                    st.markdown("#### Hypothèse de base déportée par famille")

                    if (
                        {"Nom de l'élément", "Utilisation d'une CCC"}.issubset(src_v1.columns)
                        and {"Matériel CCC", "Nombre de matériels CCC"}.issubset(bg.columns)
                    ):
                        df_src = src_v1[["Nom de l'élément", "Utilisation d'une CCC"]].dropna()
                        df_src["use_ccc"] = df_src["Utilisation d'une CCC"].astype(str).str.lower().isin(
                            ["oui", "yes", "y", "1"]
                        )

                        df_yes = (
                            df_src.groupby("Nom de l'élément", as_index=False)["use_ccc"]
                            .any()
                            .rename(columns={"Nom de l'élément": "Famille"})
                        )

                        df_qty = (
                            bg[["Matériel CCC", "Nombre de matériels CCC"]]
                            .dropna(subset=["Matériel CCC"])
                            .groupby("Matériel CCC", as_index=False)["Nombre de matériels CCC"]
                            .sum()
                            .rename(columns={"Matériel CCC": "Famille", "Nombre de matériels CCC": "Quantité"})
                        )

                        df_merge = pd.merge(df_yes, df_qty, on="Famille", how="left")
                        df_merge["Quantité"] = df_merge["Quantité"].fillna(0)
                        df_merge["Stocké en CCC ?"] = df_merge["use_ccc"].apply(lambda x: "✔️" if x else "❌")

                        df_merge = _clean_ccc_familles(df_merge)
                        display_df = df_merge[["Famille", "Stocké en CCC ?", "Quantité"]]
                        st.dataframe(
                            display_df,
                            use_container_width=True,
                            height=_df_height(len(display_df)),
                        )

                    else:
                        st.info("Colonnes nécessaires introuvables dans Tableau Source / BG")
    
                #
            # 📦 ONGLET PALETTES (V1)
            #
            with ong_pal_v1:

                st.markdown("### 📦 Palettes")

                colA, colB = st.columns(2)
                with colA:
                    st.metric(
                        "Palettes équivalentes totales (identiques V0/V1)",
                        f"{total_palettes:,.0f}".replace(",", " "),
                    )
                with colB:
                    st.metric(
                        "Surface totale (m²)",
                        f"{(total_palettes * 0.96):,.0f}".replace(",", " "),
                    )

                c1, c2 = st.columns(2)

                # Palettes par famille (Tableau Source)
                with c1:
                    st.markdown("#### Palettes par famille")
                    col_fam = None
                    for name in ["Nom de l'élément", "Nom de l'element", "Nom de l'élement"]:
                        if name in src_v1.columns:
                            col_fam = name
                            break
                    col_pal_eq = None
                    for name in ["Nombre palettes equivalent total", "Nombre palettes équivalent total"]:
                        if name in src_v1.columns:
                            col_pal_eq = name
                            break
                    if col_fam and col_pal_eq:
                        df_fam_pal = src_v1[[col_fam, col_pal_eq]].copy()
                        df_fam_pal = df_fam_pal.dropna(subset=[col_fam])
                        df_fam_pal = df_fam_pal[
                            ~df_fam_pal[col_fam].astype(str).str.lower().str.startswith("stock ccc")
                        ]
                        df_fam_pal[col_pal_eq] = pd.to_numeric(
                            df_fam_pal[col_pal_eq], errors="coerce"
                        ).fillna(0)
                        df_fam_pal = (
                            df_fam_pal.groupby(col_fam, as_index=False)[col_pal_eq]
                            .sum()
                            .sort_values(col_pal_eq, ascending=False)
                        )
                        fig_fam_pal = px.bar(
                            df_fam_pal,
                            x=col_pal_eq,
                            y=col_fam,
                            orientation="h",
                            color=col_fam,
                            color_discrete_sequence=[
                                "#F4A261",
                                "#2A9D8F",
                                "#E76F51",
                                "#264653",
                                "#8AB17D",
                                "#F1C453",
                                "#6D597A",
                            ],
                        )
                        fig_fam_pal.update_layout(
                            showlegend=False,
                            yaxis={"categoryorder": "total ascending"},
                            margin=dict(l=10, r=10, t=20, b=10),
                        )
                        st.plotly_chart(fig_fam_pal, key="palettes_famille_v1", use_container_width=True)
                    else:
                        st.info("Colonnes famille/palettes introuvables dans Tableau Source.")

                # Flux palettes
                with c2:
                    st.markdown("#### Flux mensuel de palettes")

                    flux_palettes_plot = flux_palettes.copy()
                    if not flux_palettes_plot.empty:
                        flux_palettes_plot["Mois"] = flux_palettes_plot["Mois"].astype(str)

                    fig_flux_pal_v1 = px.area(
                        flux_palettes_plot,
                        x="Mois",
                        y="Volume (nombre de palettes équivalentes)",
                    )

                    # PIC palettes V1
                    if pic_palettes > 0:
                        mois_pic_palettes_str = str(mois_pic_palettes)

                        fig_flux_pal_v1.add_vline(
                            x=mois_pic_palettes_str,
                            line_dash="dot",
                            line_color="red",
                        )
                        fig_flux_pal_v1.add_hline(
                            y=pic_palettes,
                            line_dash="dot",
                            line_color="red",
                        )
                        fig_flux_pal_v1.add_scatter(
                            x=[mois_pic_palettes_str],
                            y=[pic_palettes],
                            mode="markers",
                            marker=dict(color="red", size=10),
                            name="Pic",
                        )
                        fig_flux_pal_v1.add_annotation(
                            x=mois_pic_palettes_str,
                            y=pic_palettes,
                            text=f"Pic : {pic_palettes:.0f} palettes ({mois_pic_palettes_str})",
                            showarrow=True,
                            arrowhead=2,
                            ax=0,
                            ay=-40,
                            font=dict(color="red"),
                        )

                    fig_flux_pal_v1.update_layout(margin=dict(l=10, r=10, t=30, b=40))

                    st.plotly_chart(
                        fig_flux_pal_v1,
                        use_container_width=True,
                        key="flux_palettes_v1",
                    )


                # Matériaux stockés en CCC + Répartition palettes
                c3, c4 = st.columns(2)
                with c3:
                    st.markdown("#### Matériaux stockés en CCC")

                    mat_col = None
                    qty_col = None
                    for c in bg.columns:
                        cname = str(c).strip().lower()
                        if cname.startswith("matériel ccc"):
                            mat_col = c
                        if cname.startswith("nombre de matériels ccc"):
                            qty_col = c

                    if mat_col and qty_col:
                        df_v1_mat = (
                            bg[[mat_col, qty_col]]
                            .dropna(subset=[mat_col])
                            .groupby(mat_col, as_index=False)[qty_col]
                            .sum()
                        )

                        st.markdown("#### Répartition des matériaux stockés en CCC")
                        fig_mat_v1 = px.bar(
                            df_v1_mat,
                            x=qty_col,
                            y=mat_col,
                            orientation="h",
                        )
                        st.plotly_chart(fig_mat_v1, key="bar_mat_v1", use_container_width=True)
                    else:
                        st.info("Colonnes Matériel CCC absentes")

                with c4:
                    st.markdown("#### Répartition des palettes par étage / zone")
                    fig_pal_v1 = px.bar(
                        palettes_zone,
                        x="Étage - Zone",
                        y="Palettes",
                        color="Palettes",
                    )
                    st.plotly_chart(fig_pal_v1, key="palettes_zone_v1", use_container_width=True)

            #
            # 🚚 ONGLET CAMIONS (V1)
            #
            with ong_cam_v1:

                st.markdown("### 🚚 Camions")

                colA, colB = st.columns(2)
                colA.metric("Nombre total de camions (CCC)", f"{total_camions_ccc:,.0f}")
                colB.metric("Remplissage moyen (CCC)", f"{rempl_moyen_ccc:.1f} %")

                c1, c2 = st.columns(2)

                # Camions par zone
                with c1:
                    st.markdown("#### Camions par étage")
                    if not camions_zone_ccc.empty:
                        fig_zone_ccc = px.bar(
                            camions_zone_ccc,
                            x="Étage - Zone",
                            y="Total CCC",
                            color="Total CCC",
                        )
                        st.plotly_chart(fig_zone_ccc, key="camions_zone_v1", use_container_width=True)
                    else:
                        st.info("Colonnes camions CCC manquantes")

                # Flux camions CCC
                with c2:
                    st.markdown("#### Flux mensuel de camions")

                    if not flux_camions_ccc.empty:

                        flux_camions_plot_v1 = flux_camions_ccc.copy()
                        flux_camions_plot_v1["Nombre de Camions CCC"] = pd.to_numeric(
                            flux_camions_plot_v1["Nombre de Camions CCC"], errors="coerce"
                        ).fillna(0)

                        fig_flux_ccc = px.area(
                            flux_camions_plot_v1,
                            x="Mois",
                            y="Nombre de Camions CCC",
                        )

                        # PIC camions V1
                        if mois_pic_camions_v1:
                            fig_flux_ccc.add_vline(
                                x=mois_pic_camions_v1,
                                line_dash="dot",
                                line_color="red",
                            )
                            fig_flux_ccc.add_hline(
                                y=pic_camions_v1,
                                line_dash="dot",
                                line_color="red",
                            )
                            fig_flux_ccc.add_scatter(
                                x=[mois_pic_camions_v1],
                                y=[pic_camions_v1],
                                mode="markers",
                                marker=dict(color="red", size=10),
                                name="Pic",
                            )
                            fig_flux_ccc.add_annotation(
                                x=mois_pic_camions_v1,
                                y=pic_camions_v1,
                                text=f"Pic : {pic_camions_v1:.0f} camions ({mois_pic_camions_v1})",
                                showarrow=True,
                                arrowhead=2,
                                ax=0,
                                ay=-40,
                                font=dict(color="red"),
                            )

                        fig_flux_ccc.update_layout(margin=dict(l=10, r=10, t=30, b=40))

                        st.plotly_chart(
                            fig_flux_ccc,
                            use_container_width=True,
                            key="flux_camions_ccc_v1",
                        )

                    else:
                        st.info("Aucun flux de camions CCC")

                
                
                
                
                
                
                
                

                c_rempl_ccc, c_typo_ccc = st.columns(2)
                with c_rempl_ccc:
                    # Remplissage CCC
                    st.markdown("#### Remplissage par étage")
                    if not rempl_zone_ccc.empty:
                        x_col = None
                        for c in rempl_zone_ccc.columns:
                            cname = str(c).lower()
                            if "tage" in cname and "zone" in cname:
                                x_col = c
                                break
                        if x_col is None:
                            x_col = rempl_zone_ccc.columns[0]
                        fig_r_ccc = px.bar(
                            rempl_zone_ccc,
                            x=x_col,
                            y="Remplissage (%)",
                            color="Remplissage (%)",
                        )
                        st.plotly_chart(fig_r_ccc, key="remplissage_ccc_v1", use_container_width=True)
                    else:
                        st.info("Aucune donnÃ©e de remplissage CCC disponible")
                with c_typo_ccc:
                    # Typologie des camions (V1 â€“ CCC, colonnes X et Y)
                    st.markdown("## 🚚 Typologie des camions")
                    try:
                        # Colonnes strictes par position
                        col_type = bg.columns[23]   # colonne X
                        col_nb   = bg.columns[24]   # colonne Y

                        df_camions_ccc = (
                            bg[[col_type, col_nb]]
                            .dropna(subset=[col_type])
                            .groupby(col_type, as_index=False)[col_nb]
                            .sum()
                            .rename(columns={
                                col_type: "Type de Camion",
                                col_nb: "Nombre de Camions"
                            })
                        )

                        if df_camions_ccc.empty:
                            st.info("Aucun camion (CCC) trouvÃ© dans les colonnes X et Y.")
                        else:
                            for _, row in df_camions_ccc.iterrows():
                                nom_camion = str(row["Type de Camion"]).strip()
                                quantite = int(row["Nombre de Camions"])

                                # Filtrer uniquement les camions connus
                                if nom_camion not in daba.liste_camions:
                                    continue

                                img_path = f"images/image_camions/{nom_camion}.png"
                                c_img, c_nom, c_nb = st.columns([1, 3, 1])

                                with c_img:
                                    try:
                                        st.image(img_path, width=70)
                                    except:
                                        st.write("ðŸšš")

                                with c_nom:
                                    st.write(f"**{nom_camion}**")

                                with c_nb:
                                    st.write(f"**{quantite}**")

                    except Exception as e:
                        st.error(f"Erreur lecture typologie V1 (colonnes X et Y): {e}")





        else: 


            # Sous-onglets internes
            ong_hyp, ong_pal, ong_cam = st.tabs(
                ["📘 Hypothèses", "📦 Palettes", "🚚 Camions"]
            )

            #
            # 📘 ONGLET HYPOTHÈSES (V0)
            #
            with ong_hyp:
                st.markdown("### 📘 Hypothèses")
                h1, h2 = st.columns(2)
                with h1:
                    # Document de source
                    st.markdown("### 📄 Document de source")
                    # Pre-remplissage DPGF + Indice a partir de dpgf_date et planning_indice
                    if dpgf_date and planning_indice:
                        default_dpgf = f"DPGF du {dpgf_date} - Indice {planning_indice}"
                    elif dpgf_date:
                        default_dpgf = f"DPGF du {dpgf_date}"
                    elif planning_indice:
                        default_dpgf = f"Indice {planning_indice}"
                    else:
                        default_dpgf = ""
                    dpgf_txt = st.text_area(
                        "DPGF + Indice :",
                        value=default_dpgf,
                        key="dpgf_v0",
                        placeholder="DPGF du JJ/MM/AAAA - Indice X"
                    )
                    st.file_uploader("Veuillez joindre le fichier PIC", key="pic_v0")
                with h2:
                    # Hypothèse planning
                    st.markdown("### 🕒 Hypothèse planning")
                    st.markdown(f"- Planning indice : **{planning_indice or 'N/A'}**")
                    planning_hyp = st.text_area(
                        "Hypotheses prises pour planning :",
                        key="planning_hyp_v0",
                        placeholder="Decrire ici les hypotheses utilisees pour le planning."
                    )
                    if planning_hyp.strip() == "":
                        st.markdown(
                            "<p style='color:red;'>Hypotheses planning non completees</p>",
                            unsafe_allow_html=True,
                        )
                    else:
                        st.markdown(
                            "<p style='color:green;'>Hypotheses planning completees</p>",
                            unsafe_allow_html=True,
                        )
                h3, h4 = st.columns(2)
                with h3:
                    # Hypothèses de l’étude
                    st.markdown("### 📄 Hypothèses de l’étude")
                    st.markdown("- regroupement du materiel en grandes categories")
                    st.markdown("- conversion des conditionnements en equivalent palette (palette europeenne 1,2 x 0,8)")
                    st.markdown("- 2 phases de travaux par etage : Production et Terminaux")
                with h4:
                    # Hypothese de base par famille
                    st.markdown("###  Hypothèse de base déportée par famille")
                    if "Nom de l'élément" in src.columns:
                        familles_src = (src["Nom de l'élément"].dropna().astype(str).sort_values().unique())
                    elif "Nom" in materiel.columns:
                        familles_src = (materiel["Nom"].dropna().astype(str).sort_values().unique())
                    else:
                        familles_src = []
                    if len(familles_src) == 0:
                        st.info("Aucune famille trouvee.")
                    else:
                        df_fam = pd.DataFrame(
                            {
                                "Famille": familles_src,
                                "Stocké en CCC ?": ["❌"] * len(familles_src),
                                "Quantité": [0] * len(familles_src),
                            }
                        )
                        st.dataframe(df_fam, use_container_width=True)
            # 📦 ONGLET PALETTES (V0)
            #
            with ong_pal:
                st.markdown("### 📦 Palettes")

                # Métriques : palettes + surface
                colA, colB = st.columns(2)

                with colA:
                    st.metric(
                        "Palettes équivalentes totales (V0)",
                        f"{total_palettes:,.0f}".replace(",", " "),
                    )

                surface_totale_v0 = total_palettes * 0.96
                with colB:
                    st.metric(
                        "Surface totale occupée (m²)",
                        f"{surface_totale_v0:,.0f}".replace(",", " "),
                    )

                # Deux graphiques côte à côte
                c1, c2 = st.columns(2)

                # Palettes par famille (Tableau Source)
                with c1:
                    st.markdown("#### Palettes par famille")
                    col_fam = None
                    for name in ["Nom de l'élément", "Nom de l'element", "Nom de l'élement"]:
                        if name in src.columns:
                            col_fam = name
                            break
                    col_pal_eq = None
                    for name in ["Nombre palettes equivalent total", "Nombre palettes équivalent total"]:
                        if name in src.columns:
                            col_pal_eq = name
                            break
                    if col_fam and col_pal_eq:
                        df_fam_pal = src[[col_fam, col_pal_eq]].copy()
                        df_fam_pal = df_fam_pal.dropna(subset=[col_fam])
                        df_fam_pal = df_fam_pal[
                            ~df_fam_pal[col_fam].astype(str).str.lower().str.startswith("stock ccc")
                        ]
                        df_fam_pal[col_pal_eq] = pd.to_numeric(
                            df_fam_pal[col_pal_eq], errors="coerce"
                        ).fillna(0)
                        df_fam_pal = (
                            df_fam_pal.groupby(col_fam, as_index=False)[col_pal_eq]
                            .sum()
                            .sort_values(col_pal_eq, ascending=False)
                        )
                        fig_fam_pal = px.bar(
                            df_fam_pal,
                            x=col_pal_eq,
                            y=col_fam,
                            orientation="h",
                            color=col_fam,
                            color_discrete_sequence=[
                                "#F4A261",
                                "#2A9D8F",
                                "#E76F51",
                                "#264653",
                                "#8AB17D",
                                "#F1C453",
                                "#6D597A",
                            ],
                        )
                        fig_fam_pal.update_layout(
                            showlegend=False,
                            yaxis={"categoryorder": "total ascending"},
                            margin=dict(l=10, r=10, t=20, b=10),
                        )
                        st.plotly_chart(fig_fam_pal, key="palettes_famille_v0", use_container_width=True)
                    else:
                        st.info("Colonnes famille/palettes introuvables dans Tableau Source.")

                # Flux mensuel de palettes + PIC (V0)
                with c2:
                    st.markdown("#### Flux mensuel de palettes")

                    flux_palettes_plot_v0 = flux_palettes.copy()
                    if not flux_palettes_plot_v0.empty:
                        flux_palettes_plot_v0["Mois"] = flux_palettes_plot_v0["Mois"].astype(str)
                    mois_pic_palettes_str_v0 = str(mois_pic_palettes)

                    fig_flux_pal_v0 = px.area(
                        flux_palettes_plot_v0,
                        x="Mois",
                        y="Volume (nombre de palettes équivalentes)",
                    )

                    if pic_palettes > 0:
                        fig_flux_pal_v0.add_vline(
                            x=mois_pic_palettes_str_v0, line_dash="dot", line_color="red"
                        )
                        fig_flux_pal_v0.add_hline(
                            y=pic_palettes, line_dash="dot", line_color="red"
                        )
                        fig_flux_pal_v0.add_scatter(
                            x=[mois_pic_palettes_str_v0],
                            y=[pic_palettes],
                            mode="markers",
                            marker=dict(color="red", size=10),
                            name="Pic de livraison",
                        )
                        fig_flux_pal_v0.add_annotation(
                            x=mois_pic_palettes_str_v0,
                            y=pic_palettes,
                            text=(
                                f"Pic : {pic_palettes:.0f} palettes "
                                f"({mois_pic_palettes_str_v0})"
                            ),
                            showarrow=True,
                            arrowhead=2,
                            ax=0,
                            ay=-40,
                            font=dict(color="red"),
                        )

                    fig_flux_pal_v0.update_layout(
                        margin=dict(l=10, r=10, t=30, b=40)
                    )
                    st.plotly_chart(
                        fig_flux_pal_v0,
                        use_container_width=True,
                        height=320,
                        key="flux_palettes_v0",
                    )

                # Répartition par étage / zone
                c3, c4 = st.columns(2)
                with c3:
                    st.markdown("#### Répartition des palettes par étage / zone")
                    fig_pal_zone_v0 = px.bar(
                        palettes_zone,
                        x="Étage - Zone",
                        y="Palettes",
                        color="Palettes",
                    )
                    fig_pal_zone_v0.update_layout(margin=dict(l=10, r=10, t=30, b=40))
                    st.plotly_chart(
                        fig_pal_zone_v0,
                        use_container_width=True,
                        height=320,
                        key="palettes_zone_v0",
                    )
                with c4:
                    st.empty()

            #
            # 🚚 ONGLET CAMIONS (V0)
            #
            with ong_cam:
                st.markdown("### 🚚 Camions ")

                colA, colB = st.columns(2)
                with colA:
                    st.metric(
                        "Nombre total de camions ",
                        f"{total_camions:,.0f}".replace(",", " "),
                    )
                with colB:
                    st.metric(
                        "Remplissage moyen camions ",
                        f"{rempl_moyen:.1f} %",
                    )

                c2_, c3_ = st.columns(2)

                # CAMIONS PAR ÉTAGE / ZONE
                with c2_:
                    st.markdown("#### Camions par étage ")
                    fig_cam_zone = px.bar(
                        camions_zone,
                        x="Étage - Zone",
                        y="Camions totaux",
                        color="Camions totaux",
                    )
                    fig_cam_zone.update_layout(
                        margin=dict(l=10, r=10, t=30, b=40)
                    )
                    st.plotly_chart(
                        fig_cam_zone,
                        use_container_width=True,
                        key="cam_v0_zone",
                    )

                # FLUX MENSUEL CAMIONS (CORRIGÉ AVEC LIGNE HORIZONTALE)
                with c3_:
                    st.markdown("#### Flux mensuel de camions ")

                    # Sécurisation des données (important)
                    flux_camions_plot = flux_camions.copy()
                    flux_camions_plot["Nombre de Camions"] = pd.to_numeric(
                        flux_camions_plot["Nombre de Camions"], errors="coerce"
                    ).fillna(0)

                    fig_flux_cam = px.area(
                        flux_camions_plot,
                        x="Mois",
                        y="Nombre de Camions",
                    )

                    # PIC – Affichage même si pic_camions 0
                    if not flux_camions_plot.empty and mois_pic_camions:

                        # Ligne verticale
                        fig_flux_cam.add_vline(
                            x=mois_pic_camions,
                            line_dash="dot",
                            line_color="red"
                        )

                        # Ligne horizontale (toujours affichée)
                        fig_flux_cam.add_hline(
                            y=pic_camions,
                            line_dash="dot",
                            line_color="red"
                        )

                        # Point rouge
                        fig_flux_cam.add_scatter(
                            x=[mois_pic_camions],
                            y=[pic_camions],
                            mode="markers",
                            marker=dict(color="red", size=10),
                            name="Pic de livraison",
                        )

                        # Annotation
                        fig_flux_cam.add_annotation(
                            x=mois_pic_camions,
                            y=pic_camions,
                            text=f"Pic : {pic_camions:.0f} camions ({mois_pic_camions})",
                            showarrow=True,
                            arrowhead=2,
                            ax=0,
                            ay=-40,
                            font=dict(color="red"),
                        )

                    fig_flux_cam.update_layout(
                        margin=dict(l=10, r=10, t=30, b=40)
                    )

                    st.plotly_chart(
                        fig_flux_cam,
                        use_container_width=True,
                        key="cam_v0_flux_pic",
                    )

                # REMPLISSAGE PAR ÉTAGE / ZONE
                c4, c5 = st.columns(2)
                with c4:
                    st.markdown("#### Remplissage des camions par étage ")
                    if not rempl_zone.empty:
                        fig_rempl = px.bar(
                            rempl_zone,
                            x="Étage - Zone",
                            y="Remplissage (%)",
                            color="Remplissage (%)",
                            color_continuous_scale="Purples",
                        )
                        fig_rempl.update_layout(
                            margin=dict(l=10, r=10, t=30, b=40)
                        )
                        st.plotly_chart(
                            fig_rempl,
                            use_container_width=True,
                            key="cam_v0_rempl",
                        )
                    else:
                        st.info("Aucune donnée de remplissage disponible.")
                with c5:
                    st.empty()


                # Typologie des camions (V0 – colonnes T et U)
                c6, c7 = st.columns(2)
                with c6:
                    st.markdown("## 🚚 Typologie des camions")
                    try:
                        # Identification stricte des colonnes T et U
                        col_type = bg.columns[19]   # colonne T
                        col_nb   = bg.columns[20]   # colonne U

                        df_camions_v0 = (
                            bg[[col_type, col_nb]]
                            .dropna(subset=[col_type])
                            .groupby(col_type, as_index=False)[col_nb]
                            .sum()
                            .rename(columns={
                                col_type: "Type de Camion",
                                col_nb:   "Nombre de Camions"
                            })
                        )

                        if df_camions_v0.empty:
                            st.info("Aucun camion trouve dans les colonnes T et U.")
                        else:
                            for _, row in df_camions_v0.iterrows():
                                nom_camion = str(row["Type de Camion"]).strip()
                                quantite = int(row["Nombre de Camions"])

                                # Filtrer uniquement les camions connus
                                if nom_camion not in daba.liste_camions:
                                    continue

                                img_path = f"images/image_camions/{nom_camion}.png"
                                c_img, c_nom, c_nb = st.columns([1, 3, 1])

                                with c_img:
                                    try:
                                        st.image(img_path, width=70)
                                    except:
                                        st.write("camion")

                                with c_nom:
                                    st.write(f"**{nom_camion}**")

                                with c_nb:
                                    st.write(f"**{quantite}**")

                    except Exception as e:
                        st.error(f"Erreur lecture typologie V0 (colonnes T et U): {e}")
                with c7:
                    st.empty()





    # VARIANTES
    with tab_var:
        st.subheader("Variantes personnalisées")

        # Choix type de la nouvelle variante (même logique V0/V1)
        choix_type = st.radio(
            "Type de la nouvelle variante :",
            ["Sans CCC", "Avec CCC"],
            horizontal=True,
            key="type_variante_crea"
        )

        # CRÉATION D'UNE NOUVELLE VARIANTE
        if st.button("Créer une variante"):
            vid = f"V{st.session_state['variant_counter']}"

            # Détection du format XLSX / XLSM à partir du fichier d'origine
            import zipfile
            try:
                zip_test = zipfile.ZipFile(io.BytesIO(file_bytes))
                if any("vbaProject.bin" in f.filename for f in zip_test.filelist):
                    ext = ".xlsm"
                else:
                    ext = ".xlsx"
            except:
                st.error("Le fichier Excel d'origine est invalide.")
                st.stop()

            file_name = f"{vid}{ext}"

            # Écriture du fichier sur disque
            try:
                with open(file_name, "wb") as f:
                    f.write(file_bytes)
            except Exception as e:
                st.error(f"Erreur lors de la création de la variante : {e}")
                st.stop()

            # Ajout dans la session
            with open(file_name, "rb") as f:
                st.session_state["variants"][vid] = {
                    "with_ccc": (choix_type == "Avec CCC"),
                    "bytes": f.read(),
                    "ext": ext,
                }

            # Bouton de téléchargement immédiat
            mime = (
                "application/vnd.ms-excel" if ext == ".xlsm"
                else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            with open(file_name, "rb") as f:
                st.download_button(
                    f"Télécharger {file_name}",
                    data=f.read(),
                    file_name=file_name,
                    mime=mime,
                    key=f"download_{vid}_creation",
                )

            st.session_state["variant_counter"] += 1
            st.success(f"Variante {vid} créée.")

        # Liste des variantes existantes
        variants = st.session_state.get("variants", {})

        if not variants:
            st.info("Aucune variante pour le moment.")
        else:
            for vid, meta in variants.items():
                mois_pic_palettes_var = None

                st.markdown(
                    f"### {vid} – {'avec CCC' if meta['with_ccc'] else 'sans CCC'}"
                )

                # Bouton téléchargement du fichier de la variante
                if meta.get("bytes"):
                    st.download_button(
                        f"Télécharger {vid}{meta.get('ext', '.xlsx')}",
                        data=meta["bytes"],
                        file_name=f"{vid}{meta.get('ext', '.xlsx')}",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_{vid}_existing",
                    )

                # Upload d'un fichier Excel modifié pour cette variante
                uploaded_var = st.file_uploader(
                    f"Uploader le fichier modifié pour {vid}",
                    type=["xlsx", "xlsm"],
                    key=f"upload_{vid}",
                )

                if uploaded_var is not None:
                    meta["bytes"] = uploaded_var.read()
                    st.success(f"Fichier de {vid} mis à jour.")

                if meta.get("bytes") is None:
                    st.info("Aucun fichier chargé pour cette variante.")
                    continue

                # Lecture des feuilles du fichier VARIANTE
                try:
                    excel_io_var = io.BytesIO(meta["bytes"])
                    xls_var = pd.ExcelFile(excel_io_var)
                    bg_var = xls_var.parse("Bilan Graphique")
                    param_var = xls_var.parse("Paramétrage")
                    materiel_var = xls_var.parse("Matériel")
                    src_var = xls_var.parse("Tableau Source")
                except Exception as e:
                    st.error(f"Erreur lecture Excel pour {vid}: {e}")
                    continue

                # Application du pipeline adapté
                if meta["with_ccc"]:
                    metrics_var = pipeline_avec_ccc(bg_var)
                else:
                    metrics_var = pipeline_sans_ccc(bg_var)

                palettes_zone_var = metrics_var["palettes_zone"]
                total_palettes_var = metrics_var["total_palettes"]
                flux_palettes_var = metrics_var["flux_palettes"]
                mois_pic_palettes_var = metrics_var["mois_pic_palettes"]
                pic_palettes_var = metrics_var["pic_palettes"]
                camions_zone_var = metrics_var["camions_zone"]
                total_camions_var = metrics_var["total_camions"]
                flux_camions_var = metrics_var["flux_camions"]
                mois_pic_camions_var = metrics_var["mois_pic_camions"]
                pic_camions_var = metrics_var["pic_camions"]
                rempl_zone_var = metrics_var["rempl_zone"]
                rempl_moyen_var = metrics_var["rempl_moyen"]
                camions_type_var = metrics_var["camions_type"]

                with st.expander(f"Afficher le dashboard de {vid}", expanded=False):

                    # EXACTEMENT la même structure : Hypothèses / Palettes / Camions
                    ong_hyp_v, ong_pal_v, ong_cam_v = st.tabs(
                        ["📘 Hypothèses", "📦 Palettes", "🚚 Camions"]
                    )

                    #  VARIANTE SANS CCC  → miroir du DASHBOARD V0
                    if not meta["with_ccc"]:

                        # HYPO V0
                        with ong_hyp_v:
                            st.markdown("### 📘 Hypothèses")
                            h1, h2 = st.columns(2)
                            with h1:
                                st.markdown("#### 📄 Document de source")
                                if dpgf_date_str and dpgf_indice:
                                    default_dpgf = f"DPGF indice {dpgf_indice} du {dpgf_date_str}"
                                elif dpgf_date_str:
                                    default_dpgf = f"DPGF du {dpgf_date_str}"
                                elif dpgf_indice:
                                    default_dpgf = f"DPGF indice {dpgf_indice}"
                                else:
                                    default_dpgf = ""

                                st.text_area(
                                    "DPGF + Indice :",
                                    value=default_dpgf,
                                    key=f"dpgf_var_{vid}",
                                    placeholder="DPGF indice ? du ?",
                                )

                                st.file_uploader(
                                    "Veuillez joindre le fichier PIC",
                                    key=f"pic_var_{vid}",
                                )

                            with h2:
                                st.markdown("#### 🕒 Hypothèse planning")
                                st.markdown(f"- Planning indice : **{planning_indice or '…'}**")

                                planning_hyp_var = st.text_area(
                                    "Hypothèses prises pour planning :",
                                    key=f"planning_hyp_var_{vid}",
                                    placeholder="Décrire ici les hypothèses utilisées pour le planning…",
                                )

                                if planning_hyp_var.strip() == "":
                                    st.markdown(
                                        "<p style='color:red;'>Hypothèses planning non complétées</p>",
                                        unsafe_allow_html=True,
                                    )
                                else:
                                    st.markdown(
                                        "<p style='color:green;'>Hypothèses planning complétées</p>",
                                        unsafe_allow_html=True,
                                    )

                            h3, h4 = st.columns(2)
                            with h3:
                                # Hypothèses de l’étude (idem V0)
                                st.markdown("#### 📄 Hypothèses de l’étude")
                                st.markdown("- regroupement du matériel en grandes catégories")
                                st.markdown(
                                    "- conversion des conditionnements en équivalent palette "
                                    "(palette européenne 1,2 × 0,8)"
                                )
                                st.markdown(
                                    "- 2 phases de travaux par étage : Production et Terminaux"
                                )

                            with h4:
                                st.markdown("#### Hypothèse de base déportée par famille")

                                if "Nom de l'élément" in src_var.columns:
                                    familles_src_v = (
                                        src_var["Nom de l'élément"]
                                        .dropna()
                                        .astype(str)
                                        .sort_values()
                                        .unique()
                                    )
                                elif "Nom" in materiel_var.columns:
                                    familles_src_v = (
                                        materiel_var["Nom"]
                                        .dropna()
                                        .astype(str)
                                        .sort_values()
                                        .unique()
                                    )
                                else:
                                    familles_src_v = []
    
                                if len(familles_src_v) == 0:
                                    st.info(
                                        "Aucune famille trouvée dans Tableau Source / Matériel de la variante."
                                    )
                                else:
                                    df_fam_v = pd.DataFrame(
                                        {
                                            "Famille": familles_src_v,
                                            "Stocké en CCC ?": ["❌" for _ in familles_src_v],
                                        }
                                    )
                                    st.dataframe(df_fam_v, use_container_width=True)
    
                            # PALETTES V0 (variante)
                        with ong_pal_v:
                            st.markdown("### 📦 Palettes")

                            cA, cB = st.columns(2)
                            with cA:
                                st.metric(
                                    "Palettes équivalentes totales (Variante)",
                                    f"{total_palettes_var:,.0f}".replace(",", " "),
                                )
                            with cB:
                                st.metric(
                                    "Surface totale occupée (m²)",
                                    f"{(total_palettes_var * 0.96):,.0f}".replace(",", " "),
                                )

                            c1, c2 = st.columns(2)

                            # Répartition matériaux (on refait comme en V0 mais sur bg_var)
                            with c1:
                                st.markdown("#### Palettes par famille")
                                col_fam_v = (
                                    _find_col(src_var.columns, "Nom de l'element")
                                    or _find_col(src_var.columns, "Nom de l'élément")
                                    or _find_col(src_var.columns, "Nom de l'élement")
                                )
                                col_pal_eq_v = _find_col(src_var.columns, "Nombre palettes equivalent total")
                                if col_fam_v and col_pal_eq_v:
                                    df_fam_pal_v = src_var[[col_fam_v, col_pal_eq_v]].copy()
                                    df_fam_pal_v = df_fam_pal_v.dropna(subset=[col_fam_v])
                                    df_fam_pal_v = df_fam_pal_v[
                                        ~df_fam_pal_v[col_fam_v].astype(str).str.lower().str.startswith("stock ccc")
                                    ]
                                    df_fam_pal_v[col_pal_eq_v] = pd.to_numeric(
                                        df_fam_pal_v[col_pal_eq_v], errors="coerce"
                                    ).fillna(0)
                                    df_fam_pal_v = (
                                        df_fam_pal_v.groupby(col_fam_v, as_index=False)[col_pal_eq_v]
                                        .sum()
                                        .sort_values(col_pal_eq_v, ascending=False)
                                    )
                                    fig_fam_pal_v = px.bar(
                                        df_fam_pal_v,
                                        x=col_pal_eq_v,
                                        y=col_fam_v,
                                        orientation="h",
                                        color=col_fam_v,
                                        color_discrete_sequence=[
                                            "#F4A261",
                                            "#2A9D8F",
                                            "#E76F51",
                                            "#264653",
                                            "#8AB17D",
                                            "#F1C453",
                                            "#6D597A",
                                        ],
                                    )
                                    fig_fam_pal_v.update_layout(
                                        showlegend=False,
                                        yaxis={"categoryorder": "total ascending"},
                                        margin=dict(l=10, r=10, t=20, b=10),
                                    )
                                    st.plotly_chart(
                                        fig_fam_pal_v,
                                        key=f"pal_fam_v0_{vid}",
                                        use_container_width=True,
                                    )
                                else:
                                    st.info("Colonnes famille/palettes introuvables dans Tableau Source.")

                            # Flux mensuel palettes (variante)
                            with c2:
                                st.markdown("#### Flux mensuel de palettes ")

                                flux_palettes_plot_v = flux_palettes_var.copy()
                                if not flux_palettes_plot_v.empty:
                                    flux_palettes_plot_v["Mois"] = flux_palettes_plot_v["Mois"].astype(str)

# 2e colonne valeur
                                if not flux_palettes_plot_v.empty:
                                    y_col_pal = [
                                        c for c in flux_palettes_plot_v.columns if c != "Mois"
                                    ][0]
                                else:
                                    y_col_pal = "Volume (nombre de palettes équivalentes)"

                                fig_flux_pal_v = px.area(
                                    flux_palettes_plot_v,
                                    x="Mois",
                                    y=y_col_pal,
                                )

                                if pic_palettes_var > 0 and mois_pic_palettes_var:
                                    mois_pic_palettes_str_v = str(mois_pic_palettes_var)
                                    fig_flux_pal_v.add_vline(
                                        x=mois_pic_palettes_str_v,
                                        line_dash="dot",
                                        line_color="red",
                                    )
                                    fig_flux_pal_v.add_hline(
                                        y=pic_palettes_var,
                                        line_dash="dot",
                                        line_color="red",
                                    )
                                    fig_flux_pal_v.add_scatter(
                                        x=[mois_pic_palettes_str_v],
                                        y=[pic_palettes_var],
                                        mode="markers",
                                        marker=dict(color="red", size=10),
                                        name="Pic de livraison",
                                    )

                                fig_flux_pal_v.update_layout(
                                    margin=dict(l=10, r=10, t=30, b=40)
                                )
                                st.plotly_chart(
                                    fig_flux_pal_v,
                                    use_container_width=True,
                                )

                            # Palettes par étage / zone
                            c3, c4 = st.columns(2)
                            with c3:
                                st.markdown("#### Répartition des palettes par étage / zone ")
                                fig_pal_zone_v = px.bar(
                                    palettes_zone_var,
                                    x="Étage - Zone",
                                    y="Palettes",
                                    color="Palettes",
                                )
                                fig_pal_zone_v.update_layout(
                                    margin=dict(l=10, r=10, t=30, b=40)
                                )
                                st.plotly_chart(
                                    fig_pal_zone_v,
                                    use_container_width=True,
                                )
                            with c4:
                                st.empty()

                        # CAMIONS V0 (variante)
                        with ong_cam_v:
                            st.markdown("### 🚚 Camions")

                            colA, colB = st.columns(2)
                            with colA:
                                st.metric(
                                    "Nombre total de camions",
                                    f"{total_camions_var:,.0f}".replace(",", " "),
                                )
                            with colB:
                                st.metric(
                                    "Remplissage moyen camions ",
                                    f"{rempl_moyen_var:.1f} %",
                                )

                            c2_, c3_ = st.columns(2)

                            # Camions par étage / zone
                            with c2_:
                                st.markdown("#### Camions par étage ")
                                fig_cam_zone_v = px.bar(
                                    camions_zone_var,
                                    x="Étage - Zone",
                                    y=camions_zone_var.columns[-1],
                                    color=camions_zone_var.columns[-1],
                                )
                                fig_cam_zone_v.update_layout(
                                    margin=dict(l=10, r=10, t=30, b=40)
                                )
                                st.plotly_chart(
                                    fig_cam_zone_v,
                                    use_container_width=True,
                                )

                            # Flux mensuel camions (variante)
                            with c3_:
                                st.markdown("#### Flux mensuel de camions ")

                                fc_v = flux_camions_var.copy()
                                if not fc_v.empty:
                                    fc_v["Mois"] = fc_v["Mois"].astype(str)
                                    # Nom de la colonne Y (peut être 'Nombre de Camions' ou 'Camions')
                                    y_candidates = [c for c in fc_v.columns if c != "Mois"]
                                    y_col_cam = y_candidates[0] if y_candidates else None
                                    if y_col_cam:
                                        fc_v[y_col_cam] = pd.to_numeric(
                                            fc_v[y_col_cam],
                                            errors="coerce",
                                        ).fillna(0)
                                else:
                                    y_col_cam = None

                                if y_col_cam:
                                    fig_flux_cam_v = px.area(
                                        fc_v,
                                        x="Mois",
                                        y=y_col_cam,
                                    )
                                else:
                                    fig_flux_cam_v = None

                                if mois_pic_camions_var:
                                    fig_flux_cam_v.add_vline(
                                        x=mois_pic_camions_var,
                                        line_dash="dot",
                                        line_color="red",
                                    )
                                    fig_flux_cam_v.add_hline(
                                        y=pic_camions_var,
                                        line_dash="dot",
                                        line_color="red",
                                    )

                                if fig_flux_cam_v is not None:
                                    fig_flux_cam_v.update_layout(
                                        margin=dict(l=10, r=10, t=30, b=40)
                                    )
                                    st.plotly_chart(
                                        fig_flux_cam_v,
                                        use_container_width=True,
                                    )
                                else:
                                    st.info("Aucun flux de camions disponible pour cette variante.")

                            # Remplissage par étage / zone
                            c4, c5 = st.columns(2)
                            with c4:
                                st.markdown("#### Remplissage des camions par étage")
                                if not rempl_zone_var.empty:
                                    fig_rempl_v = px.bar(
                                        rempl_zone_var,
                                        x="Étage - Zone",
                                        y="Remplissage (%)",
                                        color="Remplissage (%)",
                                    )
                                    fig_rempl_v.update_layout(
                                        margin=dict(l=10, r=10, t=30, b=40)
                                    )
                                    st.plotly_chart(
                                        fig_rempl_v,
                                        use_container_width=True,
                                    )
                                else:
                                    st.info("Aucune donnée de remplissage disponible pour cette variante.")
                            with c5:
                                # Typologie camions (variante) – même info que V0
                                # 🚚 Typologie des camions – Variante Sans CCC (structure V0)

                                st.markdown("## 🚚 Typologie des camions ")

                                try:
                                    # Identification stricte des colonnes T et U (comme V0)
                                    col_type = bg_var.columns[19]   # colonne T
                                    col_nb   = bg_var.columns[20]   # colonne U

                                    df_camions_var = (
                                        bg_var[[col_type, col_nb]]
                                        .dropna(subset=[col_type])
                                        .groupby(col_type, as_index=False)[col_nb]
                                        .sum()
                                        .rename(columns={
                                            col_type: "Type de Camion",
                                            col_nb:   "Nombre de Camions"
                                        })
                                    )

                                    if df_camions_var.empty:
                                        st.info("Aucun camion trouvé dans les colonnes T et U pour cette variante.")
                                    else:
                                        for _, row in df_camions_var.iterrows():
                                            nom_camion = str(row["Type de Camion"]).strip()
                                            quantite = int(row["Nombre de Camions"])

                                            # Filtrer uniquement les camions connus
                                            if nom_camion not in daba.liste_camions:
                                                continue

                                            img_path = f"images/image_camions/{nom_camion}.png"
                                            c_img, c_nom, c_nb = st.columns([1, 3, 1])

                                            with c_img:
                                                try:
                                                    st.image(img_path, width=70)
                                                except:
                                                    st.write("🚚")

                                            with c_nom:
                                                st.write(f"**{nom_camion}**")

                                            with c_nb:
                                                st.write(f"**{quantite}**")

                                except Exception as e:
                                    st.error(f"Erreur lecture typologie variante Sans CCC : {e}")

                    #  VARIANTE AVEC CCC  → miroir du DASHBOARD V1
                    else:
                        # HYPO V1 (variante)
                        with ong_hyp_v:
                            st.markdown("### 📘 Hypothèses")
                            h1, h2 = st.columns(2)
                            with h1:
                                st.markdown("#### 📄 Document de source")
                                if dpgf_date_str and dpgf_indice:
                                    default_dpgf = f"DPGF indice {dpgf_indice} du {dpgf_date_str}"
                                elif dpgf_date_str:
                                    default_dpgf = f"DPGF du {dpgf_date_str}"
                                elif dpgf_indice:
                                    default_dpgf = f"DPGF indice {dpgf_indice}"
                                else:
                                    default_dpgf = ""

                                st.text_area(
                                    "DPGF + Indice :",
                                    value=default_dpgf,
                                    key=f"dpgf_var_{vid}",
                                    placeholder="DPGF indice ? du ?",
                                )

                                st.file_uploader(
                                    "Veuillez joindre le fichier PIC",
                                    key=f"pic_var_{vid}",
                                )

                            with h2:
                                st.markdown("#### 🕒 Hypothèse planning")
                                st.markdown(f"- Planning indice : **{planning_indice or '…'}**")

                                planning_hyp_var = st.text_area(
                                    "Hypothèses prises pour planning :",
                                    key=f"planning_hyp_var_{vid}",
                                    placeholder="Décrire ici les hypothèses utilisées pour le planning…",
                                )

                                if planning_hyp_var.strip() == "":
                                    st.markdown(
                                        "<p style='color:red;'>Hypothèses planning non complétées</p>",
                                        unsafe_allow_html=True,
                                    )
                                else:
                                    st.markdown(
                                        "<p style='color:green;'>Hypothèses planning complétées</p>",
                                        unsafe_allow_html=True,
                                    )

                            h3, h4 = st.columns(2)
                            with h3:
                                st.markdown("#### 📄 Hypothèses de l’étude")
                                st.markdown("- regroupement du matériel en grandes catégories")
                                st.markdown("- conversion des conditionnements en équivalent palette")
                                st.markdown("- 2 phases de travaux par étage")

                            with h4:
                                # Paramètres CCC depuis la feuille Paramétrage de la variante
                                st.markdown("#### ⚙️ Paramètres CCC")
                                def _val_param_ccc_var(libel):
                                    try:
                                        lot_col_var = param_var.columns[1]
                                        return param_var.loc[
                                            param_var["Lot"] == libel,
                                            lot_col_var,
                                        ].iloc[0]
                                    except Exception:
                                        return ""

                                duree_ccc_v = _val_param_ccc_var("Durée de stockage CCC (en mois)")
                                tarif_mois_v = _val_param_ccc_var("Tarif mois de stockage (en €)")
                                frais_sup_v = _val_param_ccc_var("Frais supplémentaires/palette (en €)")
                                frais_liv_v = _val_param_ccc_var("Frais de livraison par camion")

                                st.markdown(f"- Durée stockage : **{duree_ccc_v} mois**")
                                st.markdown(f"- Tarif de stockage : **{tarif_mois_v} €/mois**")
                                st.markdown(f"- Frais supplémentaires/palette : **{frais_sup_v} €**")
                                st.markdown(f"- Frais de livraison : **{frais_liv_v} €**")

                            st.markdown("#### Hypothèse de base déportée par famille")

                            if (
                                {"Nom de l'élément", "Utilisation d'une CCC"}.issubset(src_var.columns)
                                and {"Matériel CCC", "Nombre de matériels CCC"}.issubset(bg_var.columns)
                            ):
                                df_src_loc = src_var[
                                    ["Nom de l'élément", "Utilisation d'une CCC"]
                                ].dropna()
                                df_src_loc["use_ccc"] = (
                                    df_src_loc["Utilisation d'une CCC"]
                                    .astype(str)
                                    .str.lower()
                                    .isin(["oui", "yes", "y", "1"])
                                )
    
                                df_yes_v = (
                                    df_src_loc.groupby("Nom de l'élément", as_index=False)["use_ccc"]
                                    .any()
                                    .rename(columns={"Nom de l'élément": "Famille"})
                                )
    
                                df_qty_v = (
                                    bg_var[["Matériel CCC", "Nombre de matériels CCC"]]
                                    .dropna(subset=["Matériel CCC"])
                                    .groupby("Matériel CCC", as_index=False)["Nombre de matériels CCC"]
                                    .sum()
                                    .rename(
                                        columns={
                                            "Matériel CCC": "Famille",
                                            "Nombre de matériels CCC": "Quantité",
                                        }
                                    )
                                )
    
                                df_merge_v = pd.merge(df_yes_v, df_qty_v, on="Famille", how="left")
                                df_merge_v["Quantité"] = pd.to_numeric(
                                    df_merge_v["Quantité"], errors="coerce"
                                ).fillna(0)
                                df_merge_v = _coerce_quantite_cols(df_merge_v)
                                df_merge_v["Stocké en CCC ?"] = df_merge_v["use_ccc"].apply(
                                    lambda x: "✔️" if x else "❌"
                                )
    
                                df_merge_v = _clean_ccc_familles(df_merge_v)
                                display_df_v = df_merge_v[["Famille", "Stocké en CCC ?", "Quantité"]]
                                st.dataframe(
                                    display_df_v,
                                    use_container_width=True,
                                    height=_df_height(len(display_df_v)),
                                )
                            else:
                                st.info(
                                    "Colonnes nécessaires introuvables dans Tableau Source / Bilan Graphique de la variante."
                                )
    
                            # PALETTES V1 (variante)
                        with ong_pal_v:
                            st.markdown("### 📦 Palettes ")

                            colA, colB = st.columns(2)
                            with colA:
                                st.metric(
                                    "Palettes équivalentes totales (Variante)",
                                    f"{total_palettes_var:,.0f}".replace(",", " "),
                                )
                            with colB:
                                st.metric(
                                    "Surface totale (m²)",
                                    f"{(total_palettes_var * 0.96):,.0f}".replace(",", " "),
                                )

                            c1, c2 = st.columns(2)

                            # Palettes par famille
                            with c1:
                                st.markdown("#### Palettes par famille")
                                col_fam_v = (
                                    _find_col(src_var.columns, "Nom de l'element")
                                    or _find_col(src_var.columns, "Nom de l'élément")
                                    or _find_col(src_var.columns, "Nom de l'élement")
                                )
                                col_pal_eq_v = _find_col(src_var.columns, "Nombre palettes equivalent total")
                                if col_fam_v and col_pal_eq_v:
                                    df_fam_pal_v = src_var[[col_fam_v, col_pal_eq_v]].copy()
                                    df_fam_pal_v = df_fam_pal_v.dropna(subset=[col_fam_v])
                                    df_fam_pal_v = df_fam_pal_v[
                                        ~df_fam_pal_v[col_fam_v].astype(str).str.lower().str.startswith("stock ccc")
                                    ]
                                    df_fam_pal_v[col_pal_eq_v] = pd.to_numeric(
                                        df_fam_pal_v[col_pal_eq_v], errors="coerce"
                                    ).fillna(0)
                                    df_fam_pal_v = (
                                        df_fam_pal_v.groupby(col_fam_v, as_index=False)[col_pal_eq_v]
                                        .sum()
                                        .sort_values(col_pal_eq_v, ascending=False)
                                    )
                                    fig_fam_pal_v = px.bar(
                                        df_fam_pal_v,
                                        x=col_pal_eq_v,
                                        y=col_fam_v,
                                        orientation="h",
                                        color=col_fam_v,
                                        color_discrete_sequence=[
                                            "#F4A261",
                                            "#2A9D8F",
                                            "#E76F51",
                                            "#264653",
                                            "#8AB17D",
                                            "#F1C453",
                                            "#6D597A",
                                        ],
                                    )
                                    fig_fam_pal_v.update_layout(
                                        showlegend=False,
                                        yaxis={"categoryorder": "total ascending"},
                                        margin=dict(l=10, r=10, t=20, b=10),
                                    )
                                    st.plotly_chart(
                                        fig_fam_pal_v,
                                        key=f"pal_fam_v1_{vid}",
                                        use_container_width=True,
                                    )
                                else:
                                    st.info("Colonnes famille/palettes introuvables dans Tableau Source.")

                            # Flux palettes (identique logique V1)
                            with c2:
                                st.markdown("#### Flux mensuel de palettes")

                                flux_palettes_plot_v1 = flux_palettes_var.copy()
                                if not flux_palettes_plot_v1.empty:
                                    flux_palettes_plot_v1["Mois"] = flux_palettes_plot_v1[
                                        "Mois"
                                    ].astype(str)
                                    y_col_pal_v1 = [
                                        c for c in flux_palettes_plot_v1.columns if c != "Mois"
                                    ][0]
                                else:
                                    y_col_pal_v1 = "Volume (nombre de palettes équivalentes)"

                                fig_flux_pal_v1_var = px.area(
                                    flux_palettes_plot_v1,
                                    x="Mois",
                                    y=y_col_pal_v1,
                                )

                                if pic_palettes_var > 0 and mois_pic_palettes_var:
                                    mois_pic_palettes_str_v1 = str(mois_pic_palettes_var)
                                    fig_flux_pal_v1_var.add_vline(
                                        x=mois_pic_palettes_str_v1,
                                        line_dash="dot",
                                        line_color="red",
                                    )
                                    fig_flux_pal_v1_var.add_hline(
                                        y=pic_palettes_var,
                                        line_dash="dot",
                                        line_color="red",
                                    )

                                fig_flux_pal_v1_var.update_layout(
                                    margin=dict(l=10, r=10, t=30, b=40)
                                )
                                st.plotly_chart(
                                    fig_flux_pal_v1_var,
                                    use_container_width=True,
                                )

                            # Répartition palettes par étage / zone
                            c3, c4 = st.columns(2)
                            with c3:
                                st.markdown("#### Matériaux stockés en CCC")

                                mat_col_v = None
                                qty_col_v = None
                                for c in bg_var.columns:
                                    cname = str(c).lower()
                                    if cname.startswith("matériel ccc"):
                                        mat_col_v = c
                                    if cname.startswith("nombre de matériels ccc"):
                                        qty_col_v = c

                                if mat_col_v and qty_col_v:
                                    df_v1_mat_var = (
                                        bg_var[[mat_col_v, qty_col_v]]
                                        .dropna(subset=[mat_col_v])
                                        .groupby(mat_col_v, as_index=False)[qty_col_v]
                                        .sum()
                                    )

                                    st.markdown("#### Répartition des matériaux stockés en CCC")
                                    fig_mat_v1_var = px.bar(
                                        df_v1_mat_var,
                                        x=qty_col_v,
                                        y=mat_col_v,
                                        orientation="h",
                                    )
                                    st.plotly_chart(
                                        fig_mat_v1_var,
                                        use_container_width=True,
                                    )
                                else:
                                    st.info("Colonnes Matériel CCC absentes dans le BG de la variante.")
                            with c4:
                                st.markdown("#### Répartition des palettes par étage / zone")
                                fig_pal_v1_var = px.bar(
                                    palettes_zone_var,
                                    x="Étage - Zone",
                                    y="Palettes",
                                    color="Palettes",
                                )
                                st.plotly_chart(
                                    fig_pal_v1_var,
                                    use_container_width=True,
                                )

                        # CAMIONS V1 (variante)
                        with ong_cam_v:
                            st.markdown("### 🚚 Camions")

                            colA, colB = st.columns(2)
                            with colA:
                                st.metric(
                                    "Nombre total de camions (CCC – Variante)",
                                    f"{total_camions_var:,.0f}".replace(",", " "),
                                )
                            with colB:
                                st.metric(
                                    "Remplissage moyen (CCC – Variante)",
                                    f"{rempl_moyen_var:.1f} %",
                                )

                            c1, c2 = st.columns(2)

                            # Camions par étage / zone CCC
                            with c1:
                                st.markdown("#### Camions par étage")
                                if not camions_zone_var.empty:
# colonne Y 'Camions' (pipeline_avec_ccc)
                                    y_col_zone = [
                                        c for c in camions_zone_var.columns if c != "Étage - Zone"
                                    ][0]
                                    fig_zone_ccc_var = px.bar(
                                        camions_zone_var,
                                        x="Étage - Zone",
                                        y=y_col_zone,
                                        color=y_col_zone,
                                    )
                                    st.plotly_chart(
                                        fig_zone_ccc_var,
                                        use_container_width=True,
                                    )
                                else:
                                    st.info("Aucune donnée de camions CCC par zone pour cette variante.")

                            # Flux camions CCC
                            with c2:
                                st.markdown("#### Flux mensuel de camions")

                                if not flux_camions_var.empty:
                                    fc_v1 = flux_camions_var.copy()
                                    fc_v1["Mois"] = fc_v1["Mois"].astype(str)
                                    y_col_cam_v1 = [
                                        c for c in fc_v1.columns if c != "Mois"
                                    ][0]
                                    fc_v1[y_col_cam_v1] = pd.to_numeric(
                                        fc_v1[y_col_cam_v1],
                                        errors="coerce",
                                    ).fillna(0)

                                    fig_flux_ccc_v = px.area(
                                        fc_v1,
                                        x="Mois",
                                        y=y_col_cam_v1,
                                    )

                                    if mois_pic_camions_var:
                                        fig_flux_ccc_v.add_vline(
                                            x=mois_pic_camions_var,
                                            line_dash="dot",
                                            line_color="red",
                                        )
                                        fig_flux_ccc_v.add_hline(
                                            y=pic_camions_var,
                                            line_dash="dot",
                                            line_color="red",
                                        )

                                    fig_flux_ccc_v.update_layout(
                                        margin=dict(l=10, r=10, t=30, b=40)
                                    )
                                    st.plotly_chart(
                                        fig_flux_ccc_v,
                                        use_container_width=True,
                                    )
                                else:
                                    st.info("Aucun flux de camions CCC pour cette variante.")

                            c3, c4 = st.columns(2)
                            with c3:
                                # Remplissage CCC
                                st.markdown("#### Remplissage par étage")
                                if not rempl_zone_var.empty:
                                    fig_r_ccc_v = px.bar(
                                        rempl_zone_var,
                                        x="Étage - Zone",
                                        y="Remplissage (%)",
                                        color="Remplissage (%)",
                                    )
                                    st.plotly_chart(
                                        fig_r_ccc_v,
                                        use_container_width=True,
                                    )
                                else:
                                    st.info("Aucune donnée de remplissage CCC pour cette variante.")

                            with c4:
                                # Typologie camions CCC (variante)
                                # 🚚 Typologie des camions – Variante Avec CCC (structure V1)

                                st.markdown("## 🚚 Typologie des camions")

                                try:
                                    # Colonnes X et Y index 23 et 24 (comme V1)
                                    col_type = bg_var.columns[23]   # colonne X
                                    col_nb   = bg_var.columns[24]   # colonne Y

                                    df_camions_ccc_var = (
                                        bg_var[[col_type, col_nb]]
                                        .dropna(subset=[col_type])
                                        .groupby(col_type, as_index=False)[col_nb]
                                        .sum()
                                        .rename(columns={
                                            col_type: "Type de Camion",
                                            col_nb:   "Nombre de Camions"
                                        })
                                    )

                                    if df_camions_ccc_var.empty:
                                        st.info("Aucun camion (CCC) trouvé dans les colonnes X et Y pour cette variante.")
                                    else:
                                        for _, row in df_camions_ccc_var.iterrows():
                                            nom_camion = str(row["Type de Camion"]).strip()
                                            quantite = int(row["Nombre de Camions"])

                                            # Filtrer uniquement les camions connus
                                            if nom_camion not in daba.liste_camions:
                                                continue

                                            img_path = f"images/image_camions/{nom_camion}.png"
                                            c_img, c_nom, c_nb = st.columns([1, 3, 1])

                                            with c_img:
                                                try:
                                                    st.image(img_path, width=70)
                                                except:
                                                    st.write("🚚")

                                            with c_nom:
                                                st.write(f"**{nom_camion}**")

                                            with c_nb:
                                                st.write(f"**{quantite}**")

                                except Exception as e:
                                    st.error(f"Erreur lecture typologie variante Avec CCC : {e}")



        with tab_comp:
            st.subheader("Comparatif multi-versions")

            # 1) Construire la liste de toutes les versions
            all_versions = {
                "V0": {"with_ccc": False, "source": "base"},
                "V1": {"with_ccc": True,  "source": "base"},
            }

            for vid, meta in st.session_state.get("variants", {}).items():
                all_versions[vid] = {
                    "with_ccc": meta["with_ccc"],
                    "source": "variant",
                    "bytes": meta["bytes"],
                }

            version_names = list(all_versions.keys())

            if len(version_names) < 1:
                st.info("Aucune version disponible.")
                st.stop()

            # 2) Choix des versions à comparer (multi-sélection)
            selected_versions = st.multiselect(
                "Choisir les versions à comparer",
                version_names,
                default=[v for v in ["V0", "V1"] if v in version_names],
                key="comp_versions_multiselect",
            )

            if len(selected_versions) == 0:
                st.info("Sélectionner au moins une version.")
                st.stop()



            def compute_metrics_for_version(vname, info):
                # Charger les bonnes feuilles
                if info["source"] == "base":
                    bg_loc = bg
                    param_loc = param
                    src_loc = src
                else:
                    try:
                        excel_io_loc = io.BytesIO(info["bytes"])
                        xls_loc = pd.ExcelFile(excel_io_loc)
                        bg_loc = xls_loc.parse("Bilan Graphique")
                        param_loc = xls_loc.parse("Paramétrage")
                        src_loc = xls_loc.parse("Tableau Source")
                    except Exception:
                        return {"ok": False}

                # Palettes
                try:
                    pz = (
                        bg_loc[["Étage - Zone", "Production", "Terminaux"]]
                        .dropna(subset=["Étage - Zone"])
                        .copy()
                    )
                    pz["Production"] = pz["Production"].fillna(0)
                    pz["Terminaux"] = pz["Terminaux"].fillna(0)
                    pz["Palettes"] = pz["Production"] + pz["Terminaux"]
                    total_pal = float(pz["Palettes"].sum())
                except Exception:
                    pz = pd.DataFrame()
                    total_pal = 0.0

                # Flux palettes
                if {"Mois", "Volume (nombre de palettes équivalentes)"}.issubset(bg_loc.columns):
                    flux_pal = (
                        bg_loc[["Mois", "Volume (nombre de palettes équivalentes)"]]
                        .dropna(subset=["Mois"])
                        .groupby("Mois", as_index=False)
                        .sum()
                    )
                else:
                    flux_pal = pd.DataFrame(columns=["Mois", "Volume (nombre de palettes équivalentes)"])

                # Camions
                if info["with_ccc"]:
                    col_cam_total = "Nombre de Camions CCC"
                    col_cam_prod = "Camions Production avec CCC"
                    col_cam_term = "Camions Terminaux avec CCC"
                    col_rempl = "Remplissage camions avec CCC"
                else:
                    col_cam_total = "Nombre de Camions"
                    col_cam_prod = "Camions Production sans CCC"
                    col_cam_term = "Camions Terminaux sans CCC"
                    col_rempl = "Remplissage camions sans CCC"

                # Total camions
                if col_cam_total in bg_loc.columns:
                    total_cam = float(bg_loc[col_cam_total].fillna(0).sum())
                    flux_cam = (
                        bg_loc[["Mois", col_cam_total]]
                        .dropna(subset=["Mois"])
                        .groupby("Mois", as_index=False)
                        .sum()
                        .rename(columns={col_cam_total: "Camions"})
                    )
                else:
                    total_cam = 0.0
                    flux_cam = pd.DataFrame(columns=["Mois", "Camions"])

                # Camions zone
                if {"Étage - Zone", col_cam_prod, col_cam_term}.issubset(bg_loc.columns):
                    cz = (
                        bg_loc[["Étage - Zone", col_cam_prod, col_cam_term]]
                        .dropna(subset=["Étage - Zone"])
                        .copy()
                    )
                    cz["Camions"] = cz[col_cam_prod].fillna(0) + cz[col_cam_term].fillna(0)
                else:
                    cz = pd.DataFrame(columns=["Étage - Zone", "Camions"])

                # Remplissage
                if col_rempl in bg_loc.columns:
                    rz = (
                        bg_loc[["Étage - Zone", col_rempl]]
                        .dropna(subset=["Étage - Zone"])
                        .copy()
                    )
                    rz["Remplissage (%)"] = rz[col_rempl].fillna(0)
                    rbrut = bg_loc[col_rempl].dropna()
                    rmoy = float(rbrut.mean()) if not rbrut.empty else 0.0
                else:
                    rz = pd.DataFrame(columns=["Étage - Zone", "Remplissage (%)"])
                    rmoy = 0.0

                # Coût total logistique
                if "Coût total logistique" in bg_loc.columns:
                    cout_total = float(bg_loc["Coût total logistique"].fillna(0).sum())
                else:
                    cout_total = None

                return {
                    "ok": True,
                    "with_ccc": info["with_ccc"],
                    "bg": bg_loc,
                    "param": param_loc,
                    "src": src_loc,
                    "palettes_zone": pz,
                    "flux_palettes": flux_pal,
                    "total_palettes": total_pal,
                    "camions_zone": cz,
                    "total_camions": total_cam,
                    "flux_camions": flux_cam,
                    "rempl_zone": rz,
                    "rempl_moyen": rmoy,
                    "cout_total": cout_total,
                }

            # 4) Calcul des données pour toutes les versions sélectionnées

            data_versions = {}
            for v in selected_versions:
                data_versions[v] = compute_metrics_for_version(v, all_versions[v])

            if any(not dv["ok"] for dv in data_versions.values()):
                st.error("Impossible de lire les données pour au moins une version.")
                st.stop()

            st.markdown("### 🧩 Performance & Coûts CCC")  # titre plus petit


            def read_ccc_from_bytes(excel_bytes: bytes) -> dict:
                wb = load_workbook(filename=io.BytesIO(excel_bytes), data_only=True, keep_vba=True)
                ws = wb["Bilan Graphique"]

                labels = [ws.cell(row=1, column=c).value for c in range(31, 37)]
                values = [ws.cell(row=2, column=c).value for c in range(31, 37)]

                out = {}
                for k, v in zip(labels, values):
                    if k is None:
                        continue
                    out[str(k).strip()] = v
                return out


            def fmt_percent(x):
                try:
                    return f"{x * 100:.0f} %"
                except Exception:
                    return "—"

            def fmt_signed_percent(x):
                try:
                    sign = "+" if x >= 0 else "-"
                    return f"{sign}{abs(x) * 100:.0f} %"
                except Exception:
                    return "—"

            def fmt_arrow_percent(x, positive_is_good: bool) -> str:
                try:
                    val = float(x) * 100
                except Exception:
                    return "<span style='color:#9AA0A6'>—</span>"
                is_pos = val >= 0
                arrow = "↑" if is_pos else "↓"
                sign = "+" if is_pos else "-"
                good = is_pos if positive_is_good else not is_pos
                color = "#0F9D58" if good else "#DB4437"
                return f"<span style='color:{color}; font-weight:600;'>{arrow} {sign}{abs(val):.0f} %</span>"

            def fmt_euro(x):
                try:
                    return f"{int(round(x)):,}".replace(",", " ") + " €"
                except Exception:
                    return "—"



            for v in selected_versions:
                info = all_versions[v]

                if not info["with_ccc"]:
                    continue

                excel_bytes_v = file_bytes if info["source"] == "base" else info["bytes"]
                ccc = read_ccc_from_bytes(excel_bytes_v)

                st.markdown(f"#### {v}")

                col1, col2, col3 = st.columns(3)

                with col1:
                    with st.container(border=True):
                        st.markdown("<span style='color:gray'>% Stock CCC</span>", unsafe_allow_html=True)
                        st.markdown(f"<h3>{fmt_percent(ccc.get('% Stock CCC'))}</h3>", unsafe_allow_html=True)

                with col2:
                    with st.container(border=True):
                        st.markdown("**KPI camions**")
                        st.markdown("<span style='color:gray'>Réduction camions</span>", unsafe_allow_html=True)
                        red_val = ccc.get("% réduction Camions")
                        if "V0" in data_versions and v != "V0":
                            try:
                                base = data_versions["V0"]["total_camions"]
                                curr = data_versions[v]["total_camions"]
                                red_val = (curr - base) / base if base else None
                            except Exception:
                                pass
                        st.markdown(
                            fmt_arrow_percent(red_val, positive_is_good=False),
                            unsafe_allow_html=True,
                        )
                        st.markdown("<span style='color:gray'>Remplissage moyen des camions</span>", unsafe_allow_html=True)
                        st.markdown(
                            fmt_arrow_percent(ccc.get("% remplissage moyen des camions"), positive_is_good=True),
                            unsafe_allow_html=True,
                        )

                with col3:
                    with st.container(border=True):
                        st.markdown("**KPI Coûts**")
                        c_cost1, c_cost2 = st.columns(2)
                        with c_cost1:
                            st.markdown("<span style='color:gray'>Stockage</span>", unsafe_allow_html=True)
                            st.markdown(
                                f"<h3>{fmt_euro(ccc.get('Coût CCC stockage'))}</h3>",
                                unsafe_allow_html=True
                            )
                        with c_cost2:
                            st.markdown("<span style='color:gray'>Livraison</span>", unsafe_allow_html=True)
                            st.markdown(
                                f"<h3>{fmt_euro(ccc.get('Coût CCC livraison'))}</h3>",
                                unsafe_allow_html=True
                            )
                        st.markdown("<span style='color:gray'>Total</span>", unsafe_allow_html=True)
                        st.markdown(
                            f"<h2 style='color:#0F9D58'>{fmt_euro(ccc.get('Coût CCC Total'))}</h2>",
                            unsafe_allow_html=True
                        )

                st.markdown("---")

            # ensuite viennent les onglets
            ong_hyp_comp, ong_pal_comp, ong_cam_comp = st.tabs(["📘 Hypothèses", "📦 Palettes", "🚚 Camions"])


            # ONGLET HYPOTHÈSES
            with ong_hyp_comp:
                st.markdown("### 📘 Hypothèses de l’étude")
                st.markdown("- regroupement du matériel en grandes catégories")
                st.markdown("- conversion en équivalent palette (1,2 × 0,8 m)")
                st.markdown("- 2 phases par étage : Production & Terminaux")

                # Paramètres CCC par version
                st.markdown("### ⚙️ Paramètres CCC par version")
                for v, dv in data_versions.items():
                    if not dv["with_ccc"]:
                        continue

                    st.markdown(f"#### {v} – Avec CCC")

                    def _get_param_ccc(dv_local, libel):
                        try:
                            col = dv_local["param"].columns[1]
                            return dv_local["param"].loc[dv_local["param"]["Lot"] == libel, col].iloc[0]
                        except Exception:
                            return ""

                    st.markdown(
                        f"- Durée de stockage CCC : **{_get_param_ccc(dv, 'Durée de stockage CCC (en mois)')} mois**"
                    )
                    st.markdown(
                        f"- Tarif mois de stockage : **{_get_param_ccc(dv, 'Tarif mois de stockage (en €)')} €**"
                    )
                    st.markdown(
                        f"- Frais supplémentaires/palette : **{_get_param_ccc(dv, 'Frais supplémentaires/palette (en €)')} €**"
                    )
                    st.markdown(
                        f"- Frais de livraison par camion : **{_get_param_ccc(dv, 'Frais de livraison par camion')} €**"
                    )

                st.markdown("---")


                st.markdown("### 📦 Hypothèse de base déportée par famille ")

                # df_final existe déjà avec la colonne "Famille"
                # On le reconstruit proprement à partir de ce qui est déjà calculé

                familles = sorted(
                    set(
                        bg["Matériel CCC"].dropna().astype(str)
                        if "Matériel CCC" in bg.columns
                        else []
                    )
                    | set(
                        src["Nom de l'élément"].dropna().astype(str)
                        if "Nom de l'élément" in src.columns
                        else []
                    )
                )

                df_final = pd.DataFrame({"Famille": familles})
                df_final = _clean_ccc_familles(df_final)

                for v, dv in data_versions.items():

                    qty_col = f"Quantité_{v}"
                    flag_col = f"Stocké en CCC ? ({v})"

                    if v == "V0" or not dv["with_ccc"]:
                        if {"Désignation", "Production", "Terminaux"}.issubset(dv["bg"].columns):
                            df_qty = (
                                dv["bg"][["Désignation", "Production", "Terminaux"]]
                                .dropna(subset=["Désignation"])
                                .assign(
                                    Quantité=lambda x: x["Production"].fillna(0)
                                    + x["Terminaux"].fillna(0)
                                )
                                .groupby("Désignation", as_index=False)["Quantité"]
                                .sum()
                                .rename(columns={"Désignation": "Famille"})
                            )
                        else:
                            df_qty = pd.DataFrame(columns=["Famille", "Quantité"])
                    else:
                        if {"Matériel CCC", "Nombre de matériels CCC"}.issubset(dv["bg"].columns):
                            df_qty = (
                                dv["bg"][["Matériel CCC", "Nombre de matériels CCC"]]
                                .dropna(subset=["Matériel CCC"])
                                .groupby("Matériel CCC", as_index=False)["Nombre de matériels CCC"]
                                .sum()
                                .rename(columns={
                                    "Matériel CCC": "Famille",
                                    "Nombre de matériels CCC": "Quantité"
                                })
                            )
                        else:
                            df_qty = pd.DataFrame(columns=["Famille", "Quantité"])

                    df_final = df_final.merge(df_qty, on="Famille", how="left")
                    df_final.rename(columns={"Quantité": qty_col}, inplace=True)
                    df_final[qty_col] = df_final[qty_col].fillna(0).astype(int)
                    df_final[flag_col] = df_final[qty_col].apply(lambda x: "✔️" if x > 0 else "❌")

                # 🔥 CETTE LIGNE EST OBLIGATOIRE
                st.dataframe(
                    df_final,
                    use_container_width=True,
                    height=_df_height(len(df_final)),
                )


                                            



            # ONGLET PALETTES
            with ong_pal_comp:
                st.markdown("### 📦 Comparaison des palettes")

                # Totaux par version
                cols_tot = st.columns(len(selected_versions))
                for i, v in enumerate(selected_versions):
                    dv = data_versions[v]
                    with cols_tot[i]:
                        st.metric(
                            f"Palettes totales – {v}",
                            f"{dv['total_palettes']:,.0f}".replace(",", " "),
                        )

                pal_left, pal_right = st.columns(2)

                # Palettes par étage / zone
                with pal_left:
                    df_pal = pd.concat(
                        [
                            dv["palettes_zone"].assign(Version=v)
                            for v, dv in data_versions.items()
                            if not dv["palettes_zone"].empty
                        ],
                        ignore_index=True,
                    )

                    if not df_pal.empty:
                        fig_pal = px.bar(
                            df_pal,
                            x="Étage - Zone",
                            y="Palettes",
                            color="Version",
                            barmode="group",
                            title="Palettes par étage / zone",
                        )
                        st.plotly_chart(fig_pal, use_container_width=True)
                    else:
                        st.info("Aucune donnée palettes pour ces versions.")

                # Flux palettes
                with pal_right:
                    df_flux_pal = pd.concat(
                        [
                            dv["flux_palettes"].assign(Version=v)
                            for v, dv in data_versions.items()
                            if not dv["flux_palettes"].empty
                        ],
                        ignore_index=True,
                    )

                    if not df_flux_pal.empty:
                        fig_flux_pal = px.line(
                            df_flux_pal,
                            x="Mois",
                            y="Volume (nombre de palettes équivalentes)",
                            color="Version",
                            title="Flux mensuel de palettes",
                        )
                        st.plotly_chart(fig_flux_pal, use_container_width=True)
                    else:
                        st.info("Aucun flux palettes pour ces versions.")

            # ONGLET CAMIONS
            with ong_cam_comp:
                st.markdown("### 🚚 Comparaison des camions")

                # Totaux + remplissage
                cols_cam = st.columns(len(selected_versions))
                for i, v in enumerate(selected_versions):
                    dv = data_versions[v]
                    with cols_cam[i]:
                        st.metric(
                            f"Camions totaux – {v}",
                            f"{dv['total_camions']:,.0f}".replace(",", " "),
                        )
                        st.metric(
                            f"Remplissage moyen – {v}",
                            f"{dv['rempl_moyen']:.1f} %",
                        )

                st.markdown("---")
                cam_left, cam_right = st.columns(2)

                with cam_left:
                    st.markdown("### 🚚 Camions par étage / zone")

                    df_cam = pd.concat(
                        [
                            dv["camions_zone"].assign(Version=v)
                            for v, dv in data_versions.items()
                            if not dv["camions_zone"].empty
                        ],
                        ignore_index=True,
                    )

                    if not df_cam.empty:
                        fig_cam = px.bar(
                            df_cam,
                            x="Étage - Zone",
                            y="Camions",
                            color="Version",
                            barmode="group",
                            title="Camions par étage / zone",
                        )
                        st.plotly_chart(fig_cam, use_container_width=True)
                    else:
                        st.info("Aucune donnée camions par étage pour ces versions.")

                with cam_right:
                    st.markdown("### 📈 Flux mensuel de camions")

                    df_flux_cam = pd.concat(
                        [
                            dv["flux_camions"].assign(Version=v)
                            for v, dv in data_versions.items()
                            if not dv["flux_camions"].empty
                        ],
                        ignore_index=True,
                    )

                    if not df_flux_cam.empty:
                        fig_flux_cam = px.line(
                            df_flux_cam,
                            x="Mois",
                            y="Camions",
                            color="Version",
                            title="Flux mensuel de camions",
                        )
                        st.plotly_chart(fig_flux_cam, use_container_width=True)
                    else:
                        st.info("Aucun flux camions pour ces versions.")

                st.markdown("---")
                rem_left, rem_right = st.columns(2)
                with rem_left:
                    st.markdown("### 📦 Remplissage des camions par étage / zone")

                    df_rempl = pd.concat(
                        [
                            dv["rempl_zone"].assign(Version=v)
                            for v, dv in data_versions.items()
                            if not dv["rempl_zone"].empty
                        ],
                        ignore_index=True,
                    )

                    if not df_rempl.empty:
                        fig_rempl = px.bar(
                            df_rempl,
                            x="Étage - Zone",
                            y="Remplissage (%)",
                            color="Version",
                            barmode="group",
                            title="Remplissage des camions par étage / zone",
                        )
                        st.plotly_chart(fig_rempl, use_container_width=True)
                    else:
                        st.info("Aucune donnée de remplissage disponible pour ces versions.")
                with rem_right:
                    st.empty()


























