import streamlit as st
import pandas as pd
from datetime import date, datetime
import os
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode

import tempfile
import unicodedata, re
import math
# Imports backend
import donnees as do
import parametrage as pa
import database as daba
import entrainement_modele as entmod
import pilotage_excel as pex
import database as daba
from openpyxl import load_workbook
import io

import pandas as pd
import plotly.express as px

import shutil
import tempfile




st.set_page_config(layout="wide")

st.markdown(
    "<h1 style='text-align: center;'> Logicast - Outil de chiffrage </h1>",
    unsafe_allow_html=True
)

st.markdown("""
⚠️ **Information importante :**  
Cet outil est une **première version de test**.  
Les fonctionnalités sont en cours de développement et peuvent évoluer.
""")


def _to_internal_lot(value: str) -> str:
    v = "" if value is None else str(value).strip()
    return "GLOBAL" if v.upper() in {"GLOBAL", "TCE"} else v


def _to_display_lot(value: str) -> str:
    v = "" if value is None else str(value).strip()
    return "TCE" if v.upper() == "GLOBAL" else v


def _enrich_lot_from_db(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    if "Catégorie Prédite" not in df.columns:
        return df
    try:
        df_lots = pd.read_sql_query("SELECT nom, lot FROM materiel", daba.conn)
        lot_map = {str(r["nom"]).strip().lower(): r["lot"] for _, r in df_lots.iterrows()}
        out = df.copy()
        mapped = out["Catégorie Prédite"].astype(str).str.strip().str.lower().map(lot_map)
        if "Lot" in out.columns:
            out["Lot"] = out["Lot"].where(out["Lot"].astype(str).str.strip() != "", mapped)
        else:
            out["Lot"] = mapped
        return out
    except Exception:
        return df


def _format_date_fr(value):
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, str):
        return value.strip()
    return ""


# Menu principal
menu = st.sidebar.radio(
    "Navigation",
    ["Paramétrage", "Données", "Pilotage Excel", "Dashboard", "Entraînement modèles", "Base de données"]
)

if menu == "Paramétrage":
    st.header("⚙️ Paramétrage")

    # Persistance de la page entière
    if "parametrage_page" not in st.session_state:
        st.session_state["parametrage_page"] = {}
    p = st.session_state["parametrage_page"]

    def _param_number_input(label, key, default, **kwargs):
        state_key = f"param_{key}"
        if state_key not in st.session_state:
            st.session_state[state_key] = p.get(key, default)
        if kwargs.get("format") == "%d":
            st.session_state[state_key] = int(round(float(st.session_state[state_key])))
        val = st.number_input(label, key=state_key, **kwargs)
        p[key] = val
        return val

    # Explication générale

    # Sélections principales
    entreprises = [f.replace("_logo.png", "") for f in os.listdir("images/logos_entreprises") if f.endswith(".png")]
    if "Choix entreprise" not in entreprises:
        entreprises.insert(0, "Choix entreprise")
    # Modèles disponibles
    models = [m.replace(".pkl", "") for m in os.listdir("models") if m.endswith(".pkl")]

    # Afficher TCE à la place de GLOBAL dans l'UI
    models = ["TCE" if m == "GLOBAL" else m for m in models]
    if "TCE" not in models:
        models.insert(0, "TCE")
    if "Choix lot" not in models:
        models.insert(0, "Choix lot")

    p["entreprise_choice"] = st.selectbox(
        "Sélectionnez votre entreprise",
        entreprises,
        index=entreprises.index(p.get("entreprise_choice")) if "entreprise_choice" in p and p.get("entreprise_choice") in entreprises else 0,
    )

    # Modèles disponibles
    models = [m.replace(".pkl", "") for m in os.listdir("models") if m.endswith(".pkl")]

    # Afficher TCE à la place de GLOBAL dans l'UI
    models = ["TCE" if m == "GLOBAL" else m for m in models]
    if "TCE" not in models:
        models.insert(0, "TCE")
    if "Choix lot" not in models:
        models.insert(0, "Choix lot")

    current_model_display = _to_display_lot(p.get("model_choice"))
    p["model_choice"] = st.selectbox(
        "Sélectionnez le modèle (TCE ou spécifique)",
        models,
        index=models.index(current_model_display) if current_model_display in models else 0,
    )
    p["model_choice"] = _to_internal_lot(p["model_choice"])



    # Caractéristiques du bâtiment
    st.subheader("Caractéristiques du bâtiment")
    col1, col2, col3 = st.columns(3)
    with col1:
        _param_number_input("Nombre détages", "nombre_etages", 1, min_value=1, step=1, format="%d")
    with col2:
        _param_number_input("Zones par étage par défaut", "zones_par_etage_defaut", 1, min_value=1, step=1, format="%d")
    with col3:
        _param_number_input("Numéro étage inférieur", "numero_etage_inf", 0, step=1, format="%d")

    # Étages / Zones
    st.subheader("Étages / Zones")
    if st.button("Valider Étages / Zones"):
        df1 = pa.generate_table(p["nombre_etages"], p["zones_par_etage_defaut"], p["numero_etage_inf"])
        p["output_table"] = df1
        st.success("✅ Étages / Zones générés")

    if "output_table" in p:
        st.dataframe(p["output_table"], use_container_width=True)

    # Planning
    st.subheader("Planning")
    if st.button("➕ Insérer un indice de planning"):
        p["show_popup_planning"] = True

    if p.get("show_popup_planning", False):
        p["planning_indice"] = st.text_input(
            "👉 Indiquez l'indice du planning :",
            value=p.get("planning_indice", "")
        )
        if p["planning_indice"]:
            st.success(f"Indice enregistré : {p['planning_indice']}")
            p["show_popup_planning"] = False

    st.caption("Production = début travaux techniques, Terminaux = pose terminaux plus tard.")

    col1, col2, col3 = st.columns(3)
    with col1:
        p["date_debut_prod"] = st.text_input("Début Production (JJ/MM/AAAA)", p.get("date_debut_prod", "01/01/2025"))
    with col2:
        p["date_debut_term"] = st.text_input("Début Terminaux (JJ/MM/AAAA)", p.get("date_debut_term", "01/05/2025"))
    with col3:
        _param_number_input("Intervalle entre étages (jours)", "intervalle_par_etage", 14, min_value=0, step=1, format="%d")

    col1, col2, col3 = st.columns(3)
    with col1:
        _param_number_input("Délai livraison avant travaux (jours)", "delai_livraison", 0, min_value=0, step=1, format="%d")
    with col2:
        _param_number_input("Durée moyenne Production (jours)", "duree_prodmoyen_paretage", 30, min_value=0, step=1, format="%d")
    with col3:
        _param_number_input("Durée moyenne Terminaux (jours)", "duree_termmoyen_paretage", 30, min_value=0, step=1, format="%d")

    # Planning détaillé
    st.markdown("**Mode planning détaillé**")
    p["planning_mode"] = st.radio(
        "Choisissez votre mode de travail :",
        ["Application", "Excel"],
        index=0 if p.get("planning_mode", "Application") == "Application" else 1,
        horizontal=True,
    )

    if "output_table" in p:
        if st.button("Créer le planning détaillé"):
            etages_zones = p["output_table"]["Numéro étage (pas de lettres)"].tolist()
            zones_per_etage = p["output_table"]["Nombre de zones"].tolist()
            df2 = pa.generate_details_table(
                etages_zones, zones_per_etage,
                p["delai_livraison"], p["date_debut_prod"], p["date_debut_term"],
                p["intervalle_par_etage"],
                p["duree_prodmoyen_paretage"], p["duree_termmoyen_paretage"]
            )
            p["output_details_table"] = df2
            st.success("✅ Planning généré")

    # Tableau Détails (modifiable)
    if "output_details_table" in p:
        df_base = p["output_details_table"]

        if p.get("planning_mode", "Application") == "Application":
            st.markdown("**Tableau Détails (application)**")
            gb = GridOptionsBuilder.from_dataframe(df_base)
            gb.configure_pagination(enabled=True)
            gb.configure_default_column(editable=True, wrapText=True, autoHeight=True)
            grid_options = gb.build()

            grid_response = AgGrid(
                df_base,
                gridOptions=grid_options,
                data_return_mode='AS_INPUT',
                update_mode=GridUpdateMode.NO_UPDATE,
                fit_columns_on_grid_load=True,
                allow_unsafe_jscode=True,
                key="planning_grid"
            )

            if st.button("💾 Enregistrer le planning"):
                p["output_details_table"] = pd.DataFrame(grid_response["data"])
                st.success("✅ Planning enregistré avec succès")
        else:
            st.markdown("**Tableau Détails (Excel)**")
            st.caption("Téléchargez, modifiez dans Excel, puis ré-uploadez.")

            excel_buffer = io.BytesIO()
            df_base.to_excel(excel_buffer, index=False, sheet_name="Planning_Details")
            excel_buffer.seek(0)
            st.download_button(
                "📥 Télécharger le planning détaillé (Excel)",
                data=excel_buffer.getvalue(),
                file_name="planning_detaille_modifiable.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            uploaded_planning = st.file_uploader(
                "📤 Uploader le planning détaillé modifié (Excel)",
                type=["xlsx"],
                key="planning_details_upload",
            )
            if uploaded_planning is not None:
                try:
                    df_uploaded = pd.read_excel(uploaded_planning, engine="openpyxl")
                    required_cols = list(df_base.columns)
                    missing = [c for c in required_cols if c not in df_uploaded.columns]
                    if missing:
                        st.error(
                            "Le fichier uploadé ne contient pas les colonnes attendues : "
                            + ", ".join(missing)
                        )
                    else:
                        df_uploaded = df_uploaded[required_cols].copy()
                        for date_col in ["Date début phase production", "Date début phase terminaux"]:
                            if date_col in df_uploaded.columns:
                                df_uploaded[date_col] = df_uploaded[date_col].map(_format_date_fr)
                        p["output_details_table"] = df_uploaded
                        df_base = p["output_details_table"]
                        st.success("✅ Planning importé depuis Excel.")
                except Exception as e:
                    st.error(f"Erreur lors de l'import Excel : {e}")

            st.markdown("**Aperçu planning rechargé**")
            st.dataframe(df_base, use_container_width=True)

    # CCC
    # Activation CCC
    st.subheader("Utilisation de la CCC")
    use_ccc = st.radio(
        "Souhaitez-vous utiliser une CCC ?",
        ["Oui", "Non"],
        index=1 if p.get("use_ccc") is False else 0
    )

    p["use_ccc"] = (use_ccc == "Oui")

    if p.get("use_ccc", False):

        st.subheader("Caractéristiques du CCC")
        col1, col2, col3, col4 = st.columns(4)

        with col1:
            _param_number_input("Durée stockage (mois)", "duree_stockage", 2, min_value=0, step=1, format="%d")
        with col2:
            _param_number_input("Tarif stockage (€/mois)", "tarif_stockage", 19, min_value=0)
        with col3:
            _param_number_input("Frais palette (€)", "frais_palette", 10, min_value=0)
        with col4:
            _param_number_input("Frais livraison (€)", "frais_livraison", 175, min_value=0)

    st.subheader("Logistique du chantier")
    st.image("images/camions.png", caption="Camions disponibles")
    p["choix_camions"] = st.multiselect("Sélectionner les camions possibles", daba.liste_camions, default=p.get("choix_camions", daba.liste_camions))
    st.image("images/conditionnements.png", caption="Conditionnements disponibles")
    p["choix_conditionnement"] = st.multiselect("Sélectionner les conditionnements possibles", daba.liste_conditionnement, default=p.get("choix_conditionnement", daba.liste_conditionnement))

    # Validation finale
    if st.button("✅ Valider le paramétrage"):

        df_final = p["output_details_table"]
        _, msg = pa.validate_parametrage()

        st.session_state["parametrage"] = {
            "entreprise": p["entreprise_choice"],
            "lot": p["model_choice"],
            "nombre_etages": p["nombre_etages"],
            "duree_stockage": p["duree_stockage"],
            "tarif_stockage": p["tarif_stockage"],
            "frais_palette": p["frais_palette"],
            "frais_livraison": p["frais_livraison"],
            "df1": p.get("output_table"),
            "param_details": df_final,
            "camions": p["choix_camions"],
            "conditionnements": p["choix_conditionnement"],
            "date_debut_prod": p["date_debut_prod"],
            "date_debut_term": p["date_debut_term"],
            "intervalle_par_etage": p["intervalle_par_etage"],
            "delai_livraison": p["delai_livraison"],
            "duree_prodmoyen_paretage": p["duree_prodmoyen_paretage"],
            "duree_termmoyen_paretage": p["duree_termmoyen_paretage"],
            "planning_indice": p.get("planning_indice", ""),
            "use_ccc": p.get("use_ccc", False)

        }

        st.success(msg)

# Onglet 2 : Données
elif menu == "Données":
    st.header("📊 Classification des bordereaux")

    params = st.session_state.get("parametrage", None)
    if not params:
        st.warning("⚠️ Veuillez d'abord compléter et valider l'onglet Paramétrage avant de continuer.")
        st.stop()

    
    #  CHOIX MODE A / MODE B
    
    mode = st.radio(
        "Méthode de travail :",
        ["Travailler directement dans l'application", "Télécharger et ré-uploader après modification"]
    )

    
    

    if "donnees_page" not in st.session_state:
        st.session_state["donnees_page"] = {}
    d = st.session_state["donnees_page"]

    st.subheader("DPGF ")
    if st.button("Insérer un indice DPGF"):
        d["show_popup_dpgf"] = True

    if d.get("show_popup_dpgf", False):
        d["dpgf_indice"] = st.text_input(
            "Indiquez l'indice du DPGF :",
            value=d.get("dpgf_indice", "")
        )
        if d["dpgf_indice"]:
            st.success(f"Indice DPGF enregistré : {d['dpgf_indice']}")
            d["show_popup_dpgf"] = False

    d["dpgf_date"] = d.get("dpgf_date")
    st.session_state["dpgf_indice"] = d.get("dpgf_indice", "")
    st.session_state["dpgf_date"] = d.get("dpgf_date")
#  MODE B — Télécharger → Modifier dans Excel → Ré-uploader 
    
    if mode == "Télécharger et ré-uploader après modification":

        st.subheader("📥 Mode Édition Externe")

        st.info(
            "1️⃣ Déposez votre bordereau initial\n"
            "2️⃣ Générez un bordereau classé\n"
            "3️⃣ Téléchargez-le et modifiez-le dans Excel\n"
            "4️⃣ Ré-uploadez-le → il remplacera le tableau modifiable"
        )

        uploaded_init = st.file_uploader("📂 Bordereau Excel initial", type=["xlsx"], key="init_upload_B")

        if st.button("Créer le Bordereau classé (Mode B)") and uploaded_init:
            message, temp_path = do.process_file(params["lot"], uploaded_init)
            st.text_area("Message", value=message, height=80)

            if temp_path:
                with open(temp_path, "rb") as f:
                    st.download_button(
                        "📥 Télécharger pour modification dans Excel",
                        f,
                        file_name="bordereau_classé.xlsx"
                    )

        uploaded_modified = st.file_uploader(
            "📤 Ré-uploadez votre fichier modifié",
            type=["xlsx"],
            key="upload_modified_B"
        )

        if uploaded_modified:
            df_uploaded = pd.read_excel(uploaded_modified)
            st.session_state["bordereau_modifie"] = df_uploaded
            st.success("📌 Bordereau modifié chargé et prêt pour la génération finale.")

    
    # mode  A : Direct app 
    
    if mode == "Travailler directement dans l'application":

        # 📂 Upload du fichier Excel source
        uploaded_file = st.file_uploader("📂 Déposez le Bordereau Excel", type=["xlsx"])

        # 🚀 Création du Bordereau classé
        if st.button("Créer le Bordereau classé") and uploaded_file is not None:
            message, temp_path = do.process_file(params["lot"], uploaded_file)

            st.session_state["process_message"] = message
            st.text_area("Message", value=message, height=80)

            if temp_path:
                df_classed = pd.read_excel(temp_path)
                if "Taux de Confiance" in df_classed.columns:
                    df_classed["Taux de Confiance"] = df_classed["Taux de Confiance"].apply(lambda x: f"{x*100:.1f}%")

                st.session_state["bordereau_table"] = df_classed
                st.session_state["bordereau_modifie"] = df_classed.copy()

                with open(temp_path, "rb") as f:
                    st.download_button(
                        "📥 Télécharger le Bordereau classé (non regroupé)",
                        f,
                        file_name="bordereau_classé_non_regroupé.xlsx"
                    )
            else:
                st.error("Le bordereau n'a pas pu être généré. Corrige le fichier source puis réessaie.")

        # 📝 Tableau modifiable (AgGrid)
        if "bordereau_modifie" in st.session_state:
            st.subheader("📑 Bordereau classé (modifiable)")

            df_base = st.session_state["bordereau_modifie"]

            gb = GridOptionsBuilder.from_dataframe(df_base)
            gb.configure_pagination(enabled=True)
            gb.configure_default_column(editable=True, wrapText=True, autoHeight=True)

            for col_name in ["Quantité", "Quantite", "Prix unitaire", "Prix Unitaire"]:
                if col_name in df_base.columns:
                    gb.configure_column(col_name, hide=True)

            if "Catégorie Prédite" in df_base.columns:
                lot_actuel = params["lot"]
                try:
                    if _to_internal_lot(lot_actuel).upper() == "GLOBAL":
                        query = "SELECT DISTINCT nom FROM materiel"
                        df_cat = pd.read_sql_query(query, daba.conn)
                    else:
                        query = "SELECT DISTINCT nom FROM materiel WHERE lot = ?"
                        df_cat = pd.read_sql_query(query, daba.conn, params=(lot_actuel,))
                    categories_completes = (
                        df_cat["nom"].dropna().astype(str).unique().tolist()
                    )
                except:
                    categories_completes = (
                        df_base["Catégorie Prédite"].dropna().astype(str).unique().tolist()
                    )

                gb.configure_column(
                    "Catégorie Prédite",
                    editable=True,
                    cellEditor="agSelectCellEditor",
                    cellEditorParams={"values": sorted(categories_completes)}
                )

            grid_options = gb.build()
            grid_response = AgGrid(
                df_base,
                gridOptions=grid_options,
                data_return_mode='AS_INPUT',
                update_mode=GridUpdateMode.NO_UPDATE,
                fit_columns_on_grid_load=True,
                allow_unsafe_jscode=True,
                key="aggrid_bordereau"
            )

            if st.button("💾 Enregistrer le bordereau"):
                df_saved = pd.DataFrame(grid_response["data"])
                if _to_internal_lot(params.get("lot", "")).upper() == "GLOBAL" and "Catégorie Prédite" in df_saved.columns:
                    try:
                        df_lots = pd.read_sql_query("SELECT nom, lot FROM materiel", daba.conn)
                        lot_map = {
                            str(r["nom"]).strip().lower(): r["lot"]
                            for _, r in df_lots.iterrows()
                        }
                        df_saved["Lot"] = df_saved["Catégorie Prédite"].astype(str).str.strip().str.lower().map(lot_map)
                    except Exception:
                        pass
                st.session_state["bordereau_modifie"] = df_saved
                st.success("Bordereau enregistré. Vous pouvez maintenant créer le fichier final.")


    if "bordereau_modifie" in st.session_state:
        st.subheader("Vous pouvez télécharger le fichier si vous préférez travailler sur Excel")

        if st.button("Valider et créer le fichier final"):
            msg, path = do.finalize_wrapper(
                st.session_state["bordereau_modifie"],
                params["lot"],
                params["nombre_etages"],
                params["duree_stockage"],
                params["tarif_stockage"],
                params["frais_palette"],
                params["frais_livraison"],
                params["df1"],
                params["param_details"],
                params["camions"],
                params["conditionnements"],
                params["entreprise"],
                params["planning_indice"],
                params["use_ccc"]
            )

            st.text_area("Message final", value=msg, height=80)

            if path:
                st.session_state["pilotage_file"] = path
                with open(path, "rb") as f:
                    st.download_button(
                        "📥 Télécharger le fichier final",
                        f,
                        file_name="Outil_Logistique.xlsm"
                    )

# onglet 2 : pilotag execel 

#onglet 2 pilotage excel : 
elif menu == "Pilotage Excel":
    st.header("Pilotage Excel")

    st.subheader("Feuille Données")

    params = st.session_state.get("parametrage", None)
    if not params:
        st.warning("Veuillez d'abord completer l'onglet Parametrage.")
        st.stop()

    if "bordereau_modifie" not in st.session_state:
        st.warning("Veuillez d'abord completer l'onglet Donnees.")
        st.stop()

    if _to_internal_lot(params.get("lot", "")).upper() == "GLOBAL":
        st.session_state["bordereau_modifie"] = _enrich_lot_from_db(st.session_state["bordereau_modifie"])

    planning = params.get("param_details")
    if planning is None or planning.empty:
        st.error("Planning detaille manquant ou vide.")
        st.stop()

    if "pilotage" not in st.session_state:
        st.session_state["pilotage"] = {}
    if "donnees_grid" not in st.session_state["pilotage"]:
        st.session_state["pilotage"]["donnees_grid"] = pex.build_donnees_grid(
            st.session_state["bordereau_modifie"],
            planning,
            params.get("lot", "")
        )

    df_donnees = st.session_state["pilotage"]["donnees_grid"]
    df_display = df_donnees.astype(str)
    edited_donnees = st.data_editor(
        df_display,
        use_container_width=True,
        num_rows="dynamic",
        disabled=False,
        key="donnees_editor"
    )

    if st.button("Enregistrer Donnees", key="save_donnees"):
        st.session_state["pilotage"]["donnees_grid"] = edited_donnees.copy()
        st.success("Feuille Donnees mise a jour en session.")

    fichier_excel = st.session_state.get("pilotage_file")

    st.subheader("Feuille Tableau Source")

    # Empêcher l'affichage automatique du tableau
    if "df_source" not in st.session_state:
        st.session_state["df_source"] = None
    if "df_source_modif" not in st.session_state:
        st.session_state["df_source_modif"] = None

    # Bouton creer tableau source
    if st.button("Creer Tableau Source", key="create_source"):
        try:
            st.session_state["df_source"] = pex.build_tableau_source(
                st.session_state.get("bordereau_modifie", st.session_state["pilotage"]["donnees_grid"]),
                st.session_state["pilotage"]["donnees_grid"],
                params.get("lot", "")
            )
            st.session_state["df_source_modif"] = None
        except Exception as e:
            st.error(f"Erreur creation Tableau Source : {e}")

    # Bouton afficher tableau uniquement si df_source existe
    if st.session_state["df_source"] is not None or st.session_state["df_source_modif"] is not None:
        df_source = (
            st.session_state["df_source_modif"]
            if st.session_state["df_source_modif"] is not None
            else st.session_state["df_source"]
        )

        gb = GridOptionsBuilder.from_dataframe(df_source)
        gb.configure_default_column(editable=True)

        if "Type de camion requis" in df_source.columns:
            values = sorted(set(df_source["Type de camion requis"].dropna().unique()) | set(daba.liste_camions))
            gb.configure_column(
                "Type de camion requis",
                editable=True,
                cellEditor="agSelectCellEditor",
                cellEditorParams={"values": values}
            )

        if "Conditionnement" in df_source.columns:
            values = sorted(set(df_source["Conditionnement"].dropna().unique()) | set(daba.liste_conditionnement))
            gb.configure_column(
                "Conditionnement",
                editable=True,
                cellEditor="agSelectCellEditor",
                cellEditorParams={"values": values}
            )

        if "Utilisation d'un CCC" in df_source.columns:
            values = sorted(set(df_source["Utilisation d'un CCC"].dropna().unique()) | {"Oui", "Non"})
            gb.configure_column(
                "Utilisation d'un CCC",
                editable=True,
                cellEditor="agSelectCellEditor",
                cellEditorParams={"values": values}
            )

        grid_options = gb.build()
        grid_response = AgGrid(
            df_source,
            gridOptions=grid_options,
            update_on=["cellValueChanged"],
            fit_columns_on_grid_load=True,
            allow_unsafe_jscode=True,
            key="tableau_source"
        )

        st.session_state["df_source_modif"] = pd.DataFrame(grid_response["data"])

        if st.button("Enregistrer Tableau Source", key="save_source"):
            st.session_state["df_source"] = st.session_state["df_source_modif"].copy()
            st.success("Feuille Tableau Source mise a jour en session.")


# Onglet 3 : Dashboard :
elif menu == "Dashboard":
    import io
    st.header("📊 Études logistiques ")
    params = st.session_state.get("parametrage", None)
    use_ccc = st.session_state.get("parametrage", {}).get("use_ccc", False)
    default_mode = 1 if not params else 0
    data_mode = st.radio(
        "Source des donnees du dashboard",
        ["Dashboard actuel", "Visualiser un anciens Dahsboard"],
        index=default_mode,
        horizontal=True,
    )

    if data_mode == "Visualiser un anciens Dahsboard":
        choix_ccc = st.radio(
            "As-tu utilisé une CCC ?",
            ["Oui", "Non"],
            horizontal=True,
            key="ancien_dashboard_ccc",
        )
        # Règle demandée : Oui => V1 (avec CCC), Non => V0 (sans CCC)
        use_ccc_override = True if choix_ccc == "Oui" else False
        st.session_state.setdefault("parametrage", {})["use_ccc"] = use_ccc_override
        # Forcer l'import d'un Excel, même si un fichier existe déjà
        st.session_state["pilotage_file"] = None

        from dashboard import render_dashboard_excel
        render_dashboard_excel()
        st.stop()
    
    # 1) Prerequis : donnees de session (sans Excel)
    if not params:
        st.warning("Veuillez d'abord completer l'onglet Parametrage.")
        st.stop()
    if "bordereau_modifie" not in st.session_state:
        st.warning("Veuillez d'abord completer l'onglet Donnees.")
        st.stop()

    planning = params.get("param_details")
    if planning is None or planning.empty:
        st.warning("Planning detaille manquant. Validez le Parametrage.")
        st.stop()

    # 2) Helpers : normalisation + construction des donnees
    def _norm(s: str) -> str:
        v = "" if s is None else str(s)
        v = unicodedata.normalize("NFKD", v)
        v = "".join(c for c in v if not unicodedata.combining(c))
        return v.lower().strip()

    def _find_col(columns, target: str):
        t = _norm(target)
        for col in columns:
            if _norm(col) == t:
                return col
        return None

    def _col(df, target: str):
        return _find_col(df.columns, target)

    def _find_col_contains(columns, *tokens: str):
        tokens_norm = [_norm(t) for t in tokens]
        for col in columns:
            col_norm = _norm(col)
            if all(t in col_norm for t in tokens_norm):
                return col
        return None

    def _fix_mojibake(value: str) -> str:
        if value is None:
            return value
        text = str(value)
        for _ in range(2):
            try:
                fixed = text.encode("latin1").decode("utf-8")
            except Exception:
                break
            if fixed == text:
                break
            text = fixed
        return text

    def _fix_df_columns(df: pd.DataFrame) -> pd.DataFrame:
        if df is None or df.empty:
            return df
        col_map = {}
        for col in df.columns:
            col_map[col] = _fix_mojibake(col)
        return df.rename(columns=col_map)

    def _coerce_quantite_cols(df: pd.DataFrame) -> pd.DataFrame:
        if df is None or df.empty:
            return df
        for col in df.columns:
            if _norm(col).startswith("quantit"):
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        return df

    def _as_date(value):
        if value is None or value == "":
            return None
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime()
        if hasattr(value, "year"):
            return value
        dt = pd.to_datetime(value, dayfirst=True, errors="coerce")
        return dt.to_pydatetime() if not pd.isna(dt) else None

    def _build_param_df(p):
        rows = [
            ("Nombre etage :", p.get("nombre_etages", "")),
            ("Duree de stockage CCC (en mois)", p.get("duree_stockage", "")),
            ("Tarif mois de stockage (en EUR)", p.get("tarif_stockage", "")),
            ("Frais supplementaires/palette (en EUR)", p.get("frais_palette", "")),
            ("Frais de livraison par camion", p.get("frais_livraison", "")),
        ]
        return pd.DataFrame(rows, columns=["Lot", "Valeur"])

    def _load_materiel_df(lot_value: str):
        import sqlite3
        conn = sqlite3.connect("logistique.db")
        try:
            if _to_internal_lot(lot_value).upper() == "GLOBAL":
                df = pd.read_sql("SELECT nom, lot FROM materiel", conn)
            else:
                df = pd.read_sql("SELECT nom, lot FROM materiel WHERE lot = ?", conn, params=(lot_value,))
        finally:
            conn.close()
        if "nom" in df.columns:
            df = df.rename(columns={"nom": "Nom"})
        if "lot" in df.columns:
            df = df.rename(columns={"lot": "Lot"})
        return df

    def _ensure_donnees_grid(bordereau_df, planning_df):
        if "pilotage" in st.session_state and "donnees_grid" in st.session_state["pilotage"]:
            return st.session_state["pilotage"]["donnees_grid"]
        return pex.build_donnees_grid(bordereau_df, planning_df, params.get("lot", ""))

    def _ensure_tableau_source(bordereau_df, donnees_grid_df, lot_value: str):
        if st.session_state.get("df_source_modif") is not None:
            return st.session_state["df_source_modif"]
        if st.session_state.get("df_source") is not None:
            return st.session_state["df_source"]
        return pex.build_tableau_source(bordereau_df, donnees_grid_df, lot_value)

    def _load_refs():
        import sqlite3
        conn = sqlite3.connect("logistique.db")
        cur = conn.cursor()
        cur.execute("SELECT * FROM conditionnement")
        cond_rows = cur.fetchall()
        cur.execute("SELECT * FROM camion")
        cam_rows = cur.fetchall()
        conn.close()

        cond_dict = {}
        for r in cond_rows:
            try:
                nb = float(r[3]) if r[3] not in (None, "") else 1.0
            except Exception:
                nb = 1.0
            cond_dict[_norm(r[1])] = {"nom": r[1], "type_camion": r[2], "nb_pal_eq": nb}

        cam_by_type = {}
        for r in cam_rows:
            typ = r[2]
            cap = r[3] if r[3] is not None else 0
            cam_by_type.setdefault(typ, []).append({"nom": r[1], "capacite": cap})

        return cond_dict, cam_by_type

    def _pick_camion(nb_palettes, type_cam, cam_by_type):
        candidates = cam_by_type.get(type_cam, [])
        if not candidates:
            return None
        best = None
        meilleur_nb = 999999
        meilleur_taux = -1
        for c in candidates:
            cap = c["capacite"] if c["capacite"] else 1
            nb_camions = max(1, math.ceil(nb_palettes / cap))
            taux = (nb_palettes / (nb_camions * cap)) * 100 if nb_camions * cap > 0 else 0
            palettes_dernier = nb_palettes % cap
            taux_dernier = 100 if palettes_dernier == 0 else (palettes_dernier / cap) * 100
            if nb_palettes == 1:
                if taux > meilleur_taux:
                    best = c
                    meilleur_taux = taux
            else:
                if (nb_camions < meilleur_nb) or (nb_camions == meilleur_nb and taux > meilleur_taux and taux_dernier > 50):
                    best = c
                    meilleur_nb = nb_camions
                    meilleur_taux = taux
        return best

    def _fill_stock_camions(df_source, cond_dict, cam_by_type):
        df = df_source.copy()
        col_nom = _find_col(df.columns, "Nom de l'element")
        col_pal = _find_col(df.columns, "Nombre palettes equivalent total")
        col_cam = _find_col(df.columns, "Nombre de camions necessaires")
        col_full = _find_col(df.columns, "Dont camions pleins")
        col_fill = _find_col(df.columns, "Remplissage camion non plein")
        col_type = _find_col(df.columns, "Type de camion requis")
        if not col_nom or not col_pal:
            return df
        if not col_cam:
            df["Nombre de camions necessaires"] = None
            col_cam = "Nombre de camions necessaires"
        if not col_full:
            df["Dont camions pleins"] = None
            col_full = "Dont camions pleins"
        if not col_fill:
            df["Remplissage camion non plein"] = None
            col_fill = "Remplissage camion non plein"
        if not col_type:
            df["Type de camion requis"] = None
            col_type = "Type de camion requis"

        is_stock = df[col_nom].astype(str).str.lower().str.startswith("stock ccc")
        for idx in df[is_stock].index:
            nb_pal = df.at[idx, col_pal]
            try:
                nb_pal = float(nb_pal)
            except Exception:
                nb_pal = 0
            if nb_pal <= 0:
                continue
            cond = cond_dict.get(_norm("Palette"))
            if not cond:
                continue
            type_cam = cond["type_camion"]
            chosen = _pick_camion(nb_pal, type_cam, cam_by_type)
            if not chosen:
                continue
            cap = chosen["capacite"] if chosen["capacite"] else 1
            nb_cam = int(math.ceil(nb_pal / cap))
            full_trucks = int(nb_pal // cap)
            fill_last = round((nb_pal / cap) - full_trucks, 2) if cap else None
            df.at[idx, col_cam] = nb_cam
            df.at[idx, col_full] = full_trucks
            df.at[idx, col_fill] = fill_last
            df.at[idx, col_type] = chosen["nom"]
        return df

    def _compute_camions_type(df_source):
        col_type = _find_col(df_source.columns, "Type de camion requis")
        col_cam = _find_col(df_source.columns, "Nombre de camions necessaires")
        col_nom = _find_col(df_source.columns, "Nom de l'element")
        if not col_type or not col_cam:
            return pd.DataFrame(columns=["Type de Camion", "Nombre de Camions"])
        df = df_source.copy()
        if col_nom:
            is_stock = df[col_nom].astype(str).str.lower().str.startswith("stock ccc")
            df = df[~is_stock]
        df[col_cam] = pd.to_numeric(df[col_cam], errors="coerce").fillna(0)
        df_type = (
            df[[col_type, col_cam]]
            .dropna(subset=[col_type])
            .groupby(col_type, as_index=False)[col_cam]
            .sum()
            .rename(columns={col_type: "Type de Camion", col_cam: "Nombre de Camions"})
        )
        return df_type

    def _compute_flux(planning_df, zone_palettes, zone_camions, zone_camions_ccc):
        etage_col = _find_col(planning_df.columns, "Numero etage (pas de lettres)")
        zone_col = _find_col(planning_df.columns, "Nom Zone")
        date_prod_col = _find_col(planning_df.columns, "Date debut phase production")
        date_term_col = _find_col(planning_df.columns, "Date debut phase terminaux")
        delai_col = _find_col(planning_df.columns, "Delai de livraison avant travaux (jours)")
        duree_prod_col = _find_col(planning_df.columns, "Duree travaux production")
        duree_term_col = _find_col(planning_df.columns, "Duree travaux terminaux")
        if not etage_col or not zone_col:
            return pd.DataFrame(columns=["Mois", "Volume (nombre de palettes équivalentes)", "Nombre de Camions", "Nombre de Camions CCC"])

        agg = {}

        def _add_to_months(date_debut, duree_jours, volume, camions, camions_ccc):
            if date_debut is None:
                return
            try:
                duree_jours = float(duree_jours)
            except Exception:
                duree_jours = 0
            nombre_mois = int(math.ceil(duree_jours / 30)) if duree_jours > 0 else 1
            if nombre_mois <= 1:
                parts = [1.0]
            else:
                parts = [0.5] + [0.5 / (nombre_mois - 1)] * (nombre_mois - 1)
            for i, part in enumerate(parts):
                mois = (pd.Timestamp(date_debut) + pd.DateOffset(months=i)).strftime("%Y-%m")
                if mois not in agg:
                    agg[mois] = [0.0, 0.0, 0.0]
                agg[mois][0] += (volume or 0) * part
                agg[mois][1] += (camions or 0) * part
                agg[mois][2] += (camions_ccc or 0) * part

        for _, row in planning_df.iterrows():
            etage = row.get(etage_col)
            zone = row.get(zone_col)
            key = f"{etage} - {zone}"

            date_prod = _as_date(row.get(date_prod_col))
            date_term = _as_date(row.get(date_term_col))
            delai = row.get(delai_col, 0)

            try:
                delai_days = int(float(delai))
            except Exception:
                delai_days = 0

            prod_start = date_prod - pd.Timedelta(days=delai_days) if date_prod else None
            term_start = date_term - pd.Timedelta(days=delai_days) if date_term else None

            vol_prod = zone_palettes.get((key, "Production"), 0)
            vol_term = zone_palettes.get((key, "Terminaux"), 0)
            cam_prod = zone_camions.get((key, "Production"), 0)
            cam_term = zone_camions.get((key, "Terminaux"), 0)
            cam_prod_ccc = zone_camions_ccc.get((key, "Production"), 0)
            cam_term_ccc = zone_camions_ccc.get((key, "Terminaux"), 0)

            _add_to_months(prod_start, row.get(duree_prod_col, 0), vol_prod, cam_prod, cam_prod_ccc)
            _add_to_months(term_start, row.get(duree_term_col, 0), vol_term, cam_term, cam_term_ccc)

        rows = [
            {
                "Mois": k,
                "Volume (nombre de palettes équivalentes)": v[0],
                "Nombre de Camions": v[1],
                "Nombre de Camions CCC": v[2],
            }
            for k, v in sorted(agg.items())
        ]
        return pd.DataFrame(rows)

    def _compute_bilan_graphique(df_source, planning_df, donnees_grid=None):
        col_etage = _find_col(df_source.columns, "Etage")
        col_zone = _find_col(df_source.columns, "Zone")
        col_phase = _find_col(df_source.columns, "Phase de traveaux")
        col_pal = _find_col(df_source.columns, "Nombre palettes equivalent total")
        col_cam = _find_col(df_source.columns, "Nombre de camions necessaires")
        col_full = _find_col(df_source.columns, "Dont camions pleins")
        col_fill = _find_col(df_source.columns, "Remplissage camion non plein")
        col_ccc = (
            _find_col(df_source.columns, "Utilisation d'une CCC")
            or _find_col(df_source.columns, "Utilisation d'un CCC")
        )
        col_nom = _find_col(df_source.columns, "Nom de l'element")
        col_qty = _find_col(df_source.columns, "Quantite")
        col_type = _find_col(df_source.columns, "Type de camion requis")

        cond_dict, cam_by_type = _load_refs()
        df = _fill_stock_camions(df_source, cond_dict, cam_by_type)

        if not col_etage or not col_zone or not col_phase or not col_pal:
            return pd.DataFrame(), pd.DataFrame(columns=["Type de Camion", "Nombre de Camions"])

        df[col_pal] = pd.to_numeric(df[col_pal], errors="coerce").fillna(0)
        if col_cam:
            df[col_cam] = pd.to_numeric(df[col_cam], errors="coerce").fillna(0)
        if col_full:
            df[col_full] = pd.to_numeric(df[col_full], errors="coerce").fillna(0)
        if col_fill:
            df[col_fill] = pd.to_numeric(df[col_fill], errors="coerce").fillna(0)

        df["zone_key"] = df[col_etage].astype(str).str.strip() + " - " + df[col_zone].astype(str).str.strip()
        is_stock = df[col_nom].astype(str).str.lower().str.startswith("stock ccc") if col_nom else pd.Series(False, index=df.index)
        ccc_val = df[col_ccc].fillna("").astype(str).str.strip().str.lower() if col_ccc else pd.Series("", index=df.index)
        has_ccc_val = ccc_val != ""

        # Exclure les lignes Stock CCC des palettes (comme demandé)
        df_pal_src = df[~is_stock].copy()
        pal_group = df_pal_src.groupby(["zone_key", col_phase], as_index=False)[col_pal].sum()
        pal_prod = pal_group[pal_group[col_phase] == "Production"].set_index("zone_key")[col_pal].to_dict()
        pal_term = pal_group[pal_group[col_phase] == "Terminaux"].set_index("zone_key")[col_pal].to_dict()
        zones = sorted(set(df["zone_key"].dropna().unique().tolist()))

        df_pal = pd.DataFrame({
           "Étage - Zone": zones,
            "Production": [pal_prod.get(z, 0) for z in zones],
            "Terminaux": [pal_term.get(z, 0) for z in zones],
        })

        zone_camions = {}
        zone_camions_ccc = {}
        zone_rempl = {}
        zone_rempl_ccc = {}

        if col_cam:
            for z in zones:
                for phase in ["Production", "Terminaux"]:
                    mask_base = (df["zone_key"] == z) & (df[col_phase] == phase)
                    mask_sans = mask_base & has_ccc_val
                    mask_avec = mask_base & (ccc_val != "oui")
                    zone_camions[(z, phase)] = float(df.loc[mask_sans, col_cam].sum())
                    zone_camions_ccc[(z, phase)] = float(df.loc[mask_avec, col_cam].sum())

                mask_sans_tot = (df["zone_key"] == z) & has_ccc_val
                mask_avec_tot = (df["zone_key"] == z) & (ccc_val != "oui")

                def _rempl(mask):
                    total_cam = df.loc[mask, col_cam].sum()
                    if total_cam <= 0:
                        return 0.0
                    total_full = df.loc[mask, col_full].sum() if col_full else 0
                    total_part = df.loc[mask, col_fill].sum() if col_fill else 0
                    ratio = float((total_full + total_part) / total_cam)
                    factor = 10 ** 2
                    return math.ceil(ratio * factor) / factor

                zone_rempl[z] = _rempl(mask_sans_tot)
                zone_rempl_ccc[z] = _rempl(mask_avec_tot)

        df_cam = pd.DataFrame({
           "Étage - Zone": zones,
            "Camions Production sans CCC": [zone_camions.get((z, "Production"), 0) for z in zones],
            "Camions Terminaux sans CCC": [zone_camions.get((z, "Terminaux"), 0) for z in zones],
            "Camions Production avec CCC": [zone_camions_ccc.get((z, "Production"), 0) for z in zones],
            "Camions Terminaux avec CCC": [zone_camions_ccc.get((z, "Terminaux"), 0) for z in zones],
            "Remplissage camions sans CCC": [zone_rempl.get(z, 0) * 100 for z in zones],
            "Remplissage camions avec CCC": [zone_rempl_ccc.get(z, 0) * 100 for z in zones],
        })

        if col_nom and col_qty:
            df_qty = df.loc[(ccc_val == "oui"), [col_nom, col_qty]].copy()
            df_qty[col_qty] = pd.to_numeric(df_qty[col_qty], errors="coerce").fillna(0)
            df_mat = (
                df_qty.groupby(col_nom, as_index=False)[col_qty]
                .sum()
                .rename(columns={col_nom: "Matériel CCC", col_qty: "Nombre de matériels CCC"})
            )
        else:
            df_mat = pd.DataFrame(columns=["Matériel CCC", "Nombre de matériels CCC"])

        # Materiel complet (depuis Donnees)
        if donnees_grid is not None and {"2", "3"}.issubset(donnees_grid.columns):
            qty_col = "3"
            try:
                # Si la colonne 3 est "Lot", la quantité se trouve en colonne 4
                header_c3 = str(donnees_grid.iloc[1, 2]).strip().lower()
                if header_c3 == "lot" and "4" in donnees_grid.columns:
                    qty_col = "4"
            except Exception:
                qty_col = "3"
            df_mat_full = (
                donnees_grid.loc[3:, ["2", qty_col]]
                .rename(columns={"2": "Matériel complet", qty_col: "Nombre total de matériels"})
            )
            df_mat_full["Matériel complet"] = df_mat_full["Matériel complet"].astype(str).str.strip()
            df_mat_full["Nombre total de matériels"] = pd.to_numeric(
                df_mat_full["Nombre total de matériels"], errors="coerce"
            ).fillna(0)
            df_mat_full = df_mat_full[df_mat_full["Matériel complet"] != ""]
        else:
            df_mat_full = pd.DataFrame(columns=["Matériel complet", "Nombre total de matériels"])

        zone_palettes = {(z, "Production"): pal_prod.get(z, 0) for z in zones}
        zone_palettes.update({(z, "Terminaux"): pal_term.get(z, 0) for z in zones})
        df_flux = _compute_flux(planning_df, zone_palettes, zone_camions, zone_camions_ccc)

        # Typologie camions (sans CCC / avec CCC)
        if col_cam and col_type:
            df_type = (
                df.loc[has_ccc_val, [col_type, col_cam, col_etage, col_zone]]
                .copy()
            )
            df_type[col_cam] = pd.to_numeric(df_type[col_cam], errors="coerce").fillna(0)
            df_type["zone_key_raw"] = (
                df_type[col_etage].astype(str).str.strip()
                + df_type[col_zone].astype(str).str.strip()
            )
            df_type = (
                df_type.groupby([ "zone_key_raw", col_type ], as_index=False)[col_cam]
                .sum()
                .rename(columns={
                    "zone_key_raw": "Zone",
                    col_type: "Type de Camion",
                    col_cam: "Nombre de Camions.1",
                })
            )
            df_type.insert(0, "Étage", "")
        else:
            df_type = pd.DataFrame(columns=["Étage", "Zone", "Type de Camion", "Nombre de Camions.1"])

        if col_cam and col_type:
            df_type_ccc = (
                df.loc[ccc_val != "oui", [col_type, col_cam, col_etage, col_zone]]
                .copy()
            )
            df_type_ccc[col_cam] = pd.to_numeric(df_type_ccc[col_cam], errors="coerce").fillna(0)
            df_type_ccc["zone_key_raw"] = (
                df_type_ccc[col_etage].astype(str).str.strip()
                + df_type_ccc[col_zone].astype(str).str.strip()
            )
            df_type_ccc = (
                df_type_ccc.groupby([ "zone_key_raw", col_type ], as_index=False)[col_cam]
                .sum()
                .rename(columns={
                    "zone_key_raw": "Étage.1",
                    col_type: "Type de Camion.1",
                    col_cam: "Nombre de Camions avec CCC",
                })
            )
        else:
            df_type_ccc = pd.DataFrame(columns=["Étage.1", "Type de Camion.1", "Nombre de Camions avec CCC"])

        # KPI CCC (ligne 2 dans Excel)
        total_palettes = float(df.loc[has_ccc_val, col_pal].sum()) if col_pal else 0.0
        stock_ccc = float(
            df.loc[
                df[col_nom].astype(str).str.strip().isin(["Stock CCC Production", "Stock CCC Terminaux"]),
                col_pal,
            ].sum()
        ) if col_nom and col_pal else 0.0
        total_camions = float(df.loc[has_ccc_val, col_cam].sum()) if col_cam else 0.0
        total_camions_ccc = float(df.loc[ccc_val != "oui", col_cam].sum()) if col_cam else 0.0
        if col_cam and col_full and col_fill:
            total_full = df.loc[has_ccc_val, col_full].sum()
            total_part = df.loc[has_ccc_val, col_fill].sum()
            rempl_moyen = float((total_full + total_part) / total_camions) if total_camions else 0.0

            total_full_ccc = df.loc[ccc_val != "oui", col_full].sum()
            total_part_ccc = df.loc[ccc_val != "oui", col_fill].sum()
            rempl_moyen_ccc = float((total_full_ccc + total_part_ccc) / total_camions_ccc) if total_camions_ccc else 0.0
        else:
            rempl_moyen = 0.0
            rempl_moyen_ccc = 0.0

        camion_amelioration = ((total_camions_ccc - total_camions) / total_camions) if total_camions else 0.0
        rempl_amelioration = ((rempl_moyen_ccc - rempl_moyen) / rempl_moyen) if rempl_moyen else 0.0

        duree_ccc = float(params.get("duree_stockage", 0) or 0)
        tarif_mois = float(params.get("tarif_stockage", 0) or 0)
        frais_palette = float(params.get("frais_palette", 0) or 0)
        frais_livraison = float(params.get("frais_livraison", 0) or 0)
        cout_stock = (tarif_mois * duree_ccc + frais_palette) * stock_ccc
        cout_liv = frais_livraison * stock_ccc / 9 if stock_ccc else 0.0
        cout_total = cout_stock + cout_liv

        # Construire le Bilan Graphique avec separateurs
        sections = [
            df_pal,
            df_cam,
            df_flux,
            df_type,
            df_type_ccc,
            df_mat,
            df_mat_full,
        ]
        max_len = max([len(s) for s in sections] + [2])
        bg = pd.DataFrame(index=range(max_len))

        def _put(col, series):
            if series is None:
                bg[col] = None
            else:
                bg[col] = pd.Series(series).reset_index(drop=True)

        _put("Unnamed: 0", None)
        _put("Étage - Zone", df_pal.get("Étage - Zone"))
        _put("Production", df_pal.get("Production"))
        _put("Terminaux", df_pal.get("Terminaux"))
        _put("Unnamed: 4", None)
        _put("Étage - Zone.1", df_cam.get("Étage - Zone"))
        _put("Camions Production sans CCC", df_cam.get("Camions Production sans CCC"))
        _put("Camions Terminaux sans CCC", df_cam.get("Camions Terminaux sans CCC"))
        _put("Camions Production avec CCC", df_cam.get("Camions Production avec CCC"))
        _put("Camions Terminaux avec CCC", df_cam.get("Camions Terminaux avec CCC"))
        _put("Remplissage camions sans CCC", df_cam.get("Remplissage camions sans CCC"))
        _put("Remplissage camions avec CCC", df_cam.get("Remplissage camions avec CCC"))
        _put("Mois", df_flux.get("Mois"))
        _put("Volume (nombre de palettes équivalentes)", df_flux.get("Volume (nombre de palettes équivalentes)"))
        _put("Nombre de Camions", df_flux.get("Nombre de Camions"))
        _put("Nombre de Camions CCC", df_flux.get("Nombre de Camions CCC"))
        _put("Unnamed: 16", None)
        _put("Étage", df_type.get("Étage"))
        _put("Zone", df_type.get("Zone"))
        _put("Type de Camion", df_type.get("Type de Camion"))
        _put("Nombre de Camions.1", df_type.get("Nombre de Camions.1"))
        _put("Unnamed: 21", None)
        _put("Étage.1", df_type_ccc.get("Étage.1"))
        _put("Type de Camion.1", df_type_ccc.get("Type de Camion.1"))
        _put("Nombre de Camions avec CCC", df_type_ccc.get("Nombre de Camions avec CCC"))
        _put("Unnamed: 25", None)
        _put("Matériel CCC", df_mat.get("Matériel CCC"))
        _put("Nombre de matériels CCC", df_mat.get("Nombre de matériels CCC"))
        _put("Matériel complet", df_mat_full.get("Matériel complet"))
        _put("Nombre total de matériels", df_mat_full.get("Nombre total de matériels"))

        # KPI CCC 
        bg.loc[0, "% Stock CCC"] = (stock_ccc / total_palettes) if total_palettes else 0.0
        bg.loc[0, "% réduction Camions"] = camion_amelioration
        bg.loc[0, "% remplissage moyen des camions"] = rempl_amelioration
        bg.loc[0, "Coût CCC stockage"] = cout_stock
        bg.loc[0, "Coût CCC livraison"] = cout_liv
        bg.loc[0, "Coût CCC Total"] = cout_total

        cam_type = df_type[["Type de Camion", "Nombre de Camions.1"]].rename(columns={"Nombre de Camions.1": "Nombre de Camions"})
        return bg, cam_type

    def _build_lot_share_df(df_source: pd.DataFrame, ccc_only: bool = False) -> pd.DataFrame:
        col_lot = _find_col(df_source.columns, "Lot")
        col_pal = _find_col(df_source.columns, "Nombre palettes equivalent total")
        col_use = (
            _find_col(df_source.columns, "Utilisation d'une CCC")
            or _find_col(df_source.columns, "Utilisation d'un CCC")
        )
        col_nom = (
            _find_col(df_source.columns, "Nom de l'element")
            or _find_col(df_source.columns, "Nom de l'élément")
            or _find_col(df_source.columns, "Nom de l'élement")
        )
        if not col_lot or not col_pal:
            return pd.DataFrame(columns=["Lot", "Palettes", "Pourcentage"])

        keep_cols = [col_lot, col_pal]
        if col_use:
            keep_cols.append(col_use)
        dfp = df_source[keep_cols].copy()
        if col_nom and col_nom in df_source.columns:
            dfp[col_nom] = df_source[col_nom]
            dfp = dfp[~dfp[col_nom].astype(str).str.lower().str.startswith("stock ccc")]

        if ccc_only:
            if not col_use:
                return pd.DataFrame(columns=["Lot", "Palettes", "Pourcentage"])
            dfp[col_use] = dfp[col_use].astype(str).str.strip().str.lower()
            dfp = dfp[dfp[col_use].isin(["oui", "yes", "y", "1"])]

        dfp[col_lot] = dfp[col_lot].astype(str).str.strip()
        dfp[col_pal] = pd.to_numeric(dfp[col_pal], errors="coerce").fillna(0)
        dfp = dfp[dfp[col_lot] != ""]
        dfp = dfp.groupby(col_lot, as_index=False)[col_pal].sum()
        dfp = dfp[dfp[col_pal] > 0]
        if dfp.empty:
            return pd.DataFrame(columns=["Lot", "Palettes", "Pourcentage"])

        dfp = dfp.rename(columns={col_lot: "Lot", col_pal: "Palettes"})
        total = float(dfp["Palettes"].sum())
        dfp["Pourcentage"] = (dfp["Palettes"] / total * 100.0) if total > 0 else 0.0
        return dfp.sort_values("Palettes", ascending=False).reset_index(drop=True)

    def _plot_palettes_par_lot_pie(df_source: pd.DataFrame, key: str):
        dfp = _build_lot_share_df(df_source, ccc_only=False)

        if dfp.empty:
            st.info("Aucune palette par lot à afficher.")
            return dfp

        fig = px.pie(
            dfp,
            names="Lot",
            values="Palettes",
            hole=0.45,
            title="Pourcentage de palettes par lot",
        )
        fig.update_traces(textposition="inside", textinfo="percent+label")
        fig.update_layout(margin=dict(l=10, r=10, t=45, b=10))
        st.plotly_chart(fig, key=key, use_container_width=True)
        return dfp

    def _plot_ccc_par_lot_pie(df_source: pd.DataFrame, key: str):
        dfc = _build_lot_share_df(df_source, ccc_only=True)

        if dfc.empty:
            st.info("Aucune palette CCC par lot à afficher.")
            return dfc

        fig = px.pie(
            dfc,
            names="Lot",
            values="Palettes",
            hole=0.45,
            title="Pourcentage CCC par lot",
        )
        fig.update_traces(textposition="inside", textinfo="percent+label")
        fig.update_layout(margin=dict(l=10, r=10, t=45, b=10))
        st.plotly_chart(fig, key=key, use_container_width=True)
        return dfc

    if _to_internal_lot(params.get("lot", "")).upper() == "GLOBAL":
        st.session_state["bordereau_modifie"] = _enrich_lot_from_db(st.session_state["bordereau_modifie"])

    donnees_grid = _ensure_donnees_grid(st.session_state["bordereau_modifie"], planning)
    donnees_grid = _fix_df_columns(donnees_grid)
    src = _ensure_tableau_source(st.session_state["bordereau_modifie"], donnees_grid, params.get("lot", ""))
    src = _fix_df_columns(src)
    if "Utilisation d'un CCC" in src.columns and "Utilisation d'une CCC" not in src.columns:
        src["Utilisation d'une CCC"] = src["Utilisation d'un CCC"]

    lots_selectionnes = []
    donnees_grid_dashboard = donnees_grid
    if _to_internal_lot(params.get("lot", "")).upper() == "GLOBAL" and "Lot" in src.columns:
        lots_disponibles = sorted(src["Lot"].dropna().astype(str).str.strip().unique().tolist())
        lots_selectionnes = st.multiselect(
            "Lots à visualiser",
            options=lots_disponibles,
            default=lots_disponibles,
            key="dashboard_lots_filter",
        )
        if lots_selectionnes:
            src = src[src["Lot"].astype(str).isin(lots_selectionnes)].copy()
            try:
                if "3" in donnees_grid.columns and str(donnees_grid.iloc[1, 2]).strip().lower() == "lot":
                    entete = donnees_grid.iloc[:3, :].copy()
                    data = donnees_grid.iloc[3:, :].copy()
                    data = data[data["3"].astype(str).isin(lots_selectionnes)]
                    donnees_grid_dashboard = pd.concat([entete, data], ignore_index=True)
            except Exception:
                donnees_grid_dashboard = donnees_grid

    param = _build_param_df(params)
    param = _fix_df_columns(param)
    materiel = _load_materiel_df(params.get("lot", ""))
    if lots_selectionnes and "Lot" in materiel.columns:
        materiel = materiel[materiel["Lot"].astype(str).isin(lots_selectionnes)].copy()
    bg, camions_type_base = _compute_bilan_graphique(src, planning, donnees_grid_dashboard)
    bg = _fix_df_columns(bg)

    if bg is None or bg.empty:
        st.error("Impossible de construire le dashboard sans Excel. Verifiez les donnees.")
        st.stop()

    file_bytes = None
    file_state = st.session_state.get("pilotage_file", None)
    if isinstance(file_state, bytes):
        file_bytes = file_state
    elif isinstance(file_state, str):
        try:
            with open(file_state, "rb") as f:
                file_bytes = f.read()
        except Exception:
            file_bytes = None
    # 3) Pipelines unifiés
    #    - pipeline_sans_ccc : traitement V0
    #    - pipeline_avec_ccc : traitement V1

    def pipeline_sans_ccc(bg_df: pd.DataFrame) -> dict:
        col_zone = _col(bg_df, "Étage - Zone")
        col_prod = _col(bg_df, "Production")
        col_term = _col(bg_df, "Terminaux")

        try:
            if col_zone and col_prod and col_term:
                palettes_zone = (
                    bg_df[[col_zone, col_prod, col_term]]
                    .dropna(subset=[col_zone])
                    .copy()
                    .rename(columns={col_zone: "Étage - Zone", col_prod: "Production", col_term: "Terminaux"})
                )
                palettes_zone["Production"] = pd.to_numeric(
                    palettes_zone["Production"], errors="coerce"
                ).fillna(0)
                palettes_zone["Terminaux"] = pd.to_numeric(
                    palettes_zone["Terminaux"], errors="coerce"
                ).fillna(0)
                palettes_zone["Palettes"] = palettes_zone["Production"] + palettes_zone["Terminaux"]
            else:
                palettes_zone = pd.DataFrame(columns=["Étage - Zone", "Palettes"])
        except Exception:
            palettes_zone = pd.DataFrame(columns=["Étage - Zone", "Palettes"])
        total_palettes = float(palettes_zone["Palettes"].sum()) if not palettes_zone.empty else 0.0

        col_mois = _col(bg_df, "Mois")
        col_vol = _col(bg_df, "Volume (nombre de palettes équivalentes)")
        if col_mois and col_vol:
            flux_palettes = (
                bg_df[[col_mois, col_vol]]
                .dropna(subset=[col_mois])
                .groupby(col_mois, as_index=False)
                .sum()
                .rename(columns={col_mois: "Mois", col_vol: "Volume (nombre de palettes équivalentes)"})
            )
            if not flux_palettes.empty:
                idx_pic_pal = flux_palettes["Volume (nombre de palettes équivalentes)"].idxmax()
                mois_pic_palettes = flux_palettes.loc[idx_pic_pal, "Mois"]
                pic_palettes = float(flux_palettes.loc[idx_pic_pal, "Volume (nombre de palettes équivalentes)"])
            else:
                mois_pic_palettes = ""
                pic_palettes = 0.0
        else:
            flux_palettes = pd.DataFrame(columns=["Mois", "Volume (nombre de palettes équivalentes)"])
            mois_pic_palettes = ""
            pic_palettes = 0.0

        col_cam_prod = _col(bg_df, "Camions Production sans CCC")
        col_cam_term = _col(bg_df, "Camions Terminaux sans CCC")
        if col_zone and col_cam_prod and col_cam_term:
            camions_zone = (
                bg_df[[col_zone, col_cam_prod, col_cam_term]]
                .dropna(subset=[col_zone])
                .copy()
                .rename(columns={col_zone: "Étage - Zone", col_cam_prod: "Camions Production sans CCC", col_cam_term: "Camions Terminaux sans CCC"})
            )
            camions_zone["Camions Production sans CCC"] = camions_zone["Camions Production sans CCC"].fillna(0)
            camions_zone["Camions Terminaux sans CCC"] = camions_zone["Camions Terminaux sans CCC"].fillna(0)
            camions_zone["Camions totaux"] = camions_zone["Camions Production sans CCC"] + camions_zone["Camions Terminaux sans CCC"]
        else:
            camions_zone = pd.DataFrame(columns=["Étage - Zone", "Camions totaux"])

        col_cam_total = _col(bg_df, "Nombre de Camions")
        if col_cam_total:
            total_camions = float(bg_df[col_cam_total].fillna(0).sum())
        else:
            total_camions = float(camions_zone["Camions totaux"].sum()) if not camions_zone.empty else 0.0

        col_rempl = _col(bg_df, "Remplissage camions sans CCC")
        if col_zone and col_rempl:
            rempl_zone = (
                bg_df[[col_zone, col_rempl]]
                .dropna(subset=[col_zone])
                .copy()
                .rename(columns={col_zone: "Étage - Zone", col_rempl: "Remplissage camions sans CCC"})
            )
            rempl_zone["Remplissage (%)"] = rempl_zone["Remplissage camions sans CCC"].fillna(0)
            if not camions_zone.empty and "Camions totaux" in camions_zone.columns:
                merged = rempl_zone.merge(
                    camions_zone[["Étage - Zone", "Camions totaux"]],
                    on="Étage - Zone",
                    how="left",
                )
                denom = merged["Camions totaux"].sum()
                rempl_moyen = float(
                    (merged["Remplissage (%)"] * merged["Camions totaux"]).sum() / denom
                ) if denom else 0.0
            else:
                rempl_brut = bg_df[col_rempl].dropna()
                rempl_moyen = float(rempl_brut.mean()) if not rempl_brut.empty else 0.0
        else:
            rempl_zone = pd.DataFrame(columns=["Étage - Zone", "Remplissage (%)"])
            rempl_moyen = 0.0

        if col_cam_total and col_mois:
            flux_camions = (
                bg_df[[col_mois, col_cam_total]]
                .dropna(subset=[col_mois])
                .groupby(col_mois, as_index=False)
                .sum()
                .rename(columns={col_mois: "Mois", col_cam_total: "Nombre de Camions"})
            )
            if not flux_camions.empty:
                idx_pic_cam = flux_camions["Nombre de Camions"].idxmax()
                mois_pic_camions = flux_camions.loc[idx_pic_cam, "Mois"]
                pic_camions = float(flux_camions.loc[idx_pic_cam, "Nombre de Camions"])
            else:
                mois_pic_camions = ""
                pic_camions = 0.0
        else:
            flux_camions = pd.DataFrame(columns=["Mois", "Nombre de Camions"])
            mois_pic_camions = ""
            pic_camions = 0.0

        # Camions par type (meme logique que dashboard.py)
        cols = list(bg_df.columns)
        col_etage_type = None
        for c in cols:
            if str(c).startswith("Étage") and c != "Étage - Zone":
                col_etage_type = c
                break
        if col_etage_type:
            idx_etage_type = cols.index(col_etage_type)
            col_type_camion = cols[idx_etage_type + 2] if idx_etage_type + 2 < len(cols) else None
            col_nb_camions_type = cols[idx_etage_type + 3] if idx_etage_type + 3 < len(cols) else None
            if col_type_camion and col_nb_camions_type:
                camions_type = (
                    bg_df[[col_type_camion, col_nb_camions_type]]
                    .dropna(subset=[col_type_camion])
                    .copy()
                )
                camions_type = (
                    camions_type
                    .groupby(col_type_camion, as_index=False)[col_nb_camions_type]
                    .sum()
                )
                camions_type.rename(
                    columns={
                        col_type_camion: "Type de Camion",
                        col_nb_camions_type: "Nombre de Camions",
                    },
                    inplace=True,
                )
            else:
                camions_type = pd.DataFrame(columns=["Type de Camion", "Nombre de Camions"])
        else:
            camions_type = pd.DataFrame(columns=["Type de Camion", "Nombre de Camions"])

        return {
            "palettes_zone": palettes_zone,
            "total_palettes": total_palettes,
            "flux_palettes": flux_palettes,
            "mois_pic_palettes": mois_pic_palettes,
            "pic_palettes": pic_palettes,
            "camions_zone": camions_zone,
            "total_camions": total_camions,
            "flux_camions": flux_camions,
            "mois_pic_camions": mois_pic_camions,
            "pic_camions": pic_camions,
            "rempl_zone": rempl_zone,
            "rempl_moyen": rempl_moyen,
            "camions_type": camions_type,
        }

    def pipeline_avec_ccc(bg_df: pd.DataFrame) -> dict:
        base = pipeline_sans_ccc(bg_df)

        col_mois = _col(bg_df, "Mois")
        col_cam_ccc = _col(bg_df, "Nombre de Camions CCC")
        if col_mois and col_cam_ccc:
            flux_camions_ccc = (
                bg_df[[col_mois, col_cam_ccc]]
                .dropna(subset=[col_mois])
                .groupby(col_mois, as_index=False)
                .sum()
                .rename(columns={col_mois: "Mois", col_cam_ccc: "Nombre de Camions CCC"})
            )
            if not flux_camions_ccc.empty:
                idx_pic_cam_v1 = flux_camions_ccc["Nombre de Camions CCC"].idxmax()
                mois_pic_camions_v1 = flux_camions_ccc.loc[idx_pic_cam_v1, "Mois"]
                pic_camions_v1 = float(flux_camions_ccc.loc[idx_pic_cam_v1, "Nombre de Camions CCC"])
            else:
                mois_pic_camions_v1 = ""
                pic_camions_v1 = 0.0
        else:
            flux_camions_ccc = pd.DataFrame(columns=["Mois", "Nombre de Camions CCC"])
            mois_pic_camions_v1 = ""
            pic_camions_v1 = 0.0

        if col_cam_ccc:
            total_camions_ccc = float(bg_df[col_cam_ccc].fillna(0).sum())
        else:
            total_camions_ccc = 0.0

        col_rempl_ccc = _col(bg_df, "Remplissage camions avec CCC")
        col_zone = _col(bg_df, "Étage - Zone")
        if col_zone and col_rempl_ccc:
            rempl_brut_ccc = bg_df[col_rempl_ccc].dropna()
            rempl_moyen_ccc = float(rempl_brut_ccc.mean()) if not rempl_brut_ccc.empty else 0.0
            rempl_zone_ccc = (
                bg_df[[col_zone, col_rempl_ccc]]
                .dropna(subset=[col_zone])
                .copy()
                .rename(columns={col_zone: "Étage - Zone", col_rempl_ccc: "Remplissage camions avec CCC"})
            )
            rempl_zone_ccc["Remplissage (%)"] = rempl_zone_ccc["Remplissage camions avec CCC"].fillna(0)
        else:
            rempl_moyen_ccc = 0.0
            rempl_zone_ccc = pd.DataFrame(columns=["Étage - Zone", "Remplissage (%)"])

        col_cam_prod_ccc = _col(bg_df, "Camions Production avec CCC")
        col_cam_term_ccc = _col(bg_df, "Camions Terminaux avec CCC")
        if col_zone and col_cam_prod_ccc and col_cam_term_ccc:
            camions_zone_ccc = (
                bg_df[[col_zone, col_cam_prod_ccc, col_cam_term_ccc]]
                .dropna(subset=[col_zone])
                .copy()
                .rename(columns={
                    col_zone: "Étage - Zone",
                    col_cam_prod_ccc: "Camions Production avec CCC",
                    col_cam_term_ccc: "Camions Terminaux avec CCC",
                })
            )
            camions_zone_ccc["Total CCC"] = (
                camions_zone_ccc["Camions Production avec CCC"].fillna(0)
                + camions_zone_ccc["Camions Terminaux avec CCC"].fillna(0)
            )
        else:
            camions_zone_ccc = pd.DataFrame(columns=["Étage - Zone", "Total CCC"])

        if not camions_zone_ccc.empty and "Total CCC" in camions_zone_ccc.columns and not rempl_zone_ccc.empty:
            merged_ccc = rempl_zone_ccc.merge(
                camions_zone_ccc[["Étage - Zone", "Total CCC"]],
                on="Étage - Zone",
                how="left",
            )
            denom_ccc = merged_ccc["Total CCC"].sum()
            rempl_moyen_ccc = float(
                (merged_ccc["Remplissage (%)"] * merged_ccc["Total CCC"]).sum() / denom_ccc
            ) if denom_ccc else 0.0

        return {
            "palettes_zone": base["palettes_zone"],
            "total_palettes": base["total_palettes"],
            "flux_palettes": base["flux_palettes"],
            "mois_pic_palettes": base["mois_pic_palettes"],
            "pic_palettes": base["pic_palettes"],
            "camions_zone": camions_zone_ccc.rename(columns={"Total CCC": "Camions"}) if not camions_zone_ccc.empty else pd.DataFrame(columns=["Étage - Zone", "Camions"]),
            "total_camions": total_camions_ccc,
            "flux_camions": flux_camions_ccc.rename(columns={"Nombre de Camions CCC": "Camions"}) if not flux_camions_ccc.empty else pd.DataFrame(columns=["Mois", "Camions"]),
            "mois_pic_camions": mois_pic_camions_v1,
            "pic_camions": pic_camions_v1,
            "rempl_zone": rempl_zone_ccc,
            "rempl_moyen": rempl_moyen_ccc,
            "camions_type": base["camions_type"],
        }

    def _clean_ccc_familles(df: pd.DataFrame, col: str = "Famille") -> pd.DataFrame:
        exclude = {"stock ccc production", "stock ccc terminaux"}
        if col not in df.columns:
            return df
        mask = (
            df[col]
            .astype(str)
            .str.strip()
            .str.lower()
            .isin(exclude)
        )
        return df.loc[~mask].copy()

    def _df_height(n_rows: int, row_h: int = 32, min_h: int = 180, max_h: int = 2000) -> int:
        return max(min_h, min(max_h, (n_rows + 1) * row_h))

    # 3 bis) Préparation des données de base via pipelines

    metrics_v0 = pipeline_sans_ccc(bg)
    palettes_zone = metrics_v0["palettes_zone"]
    total_palettes = metrics_v0["total_palettes"]
    flux_palettes = metrics_v0["flux_palettes"]
    mois_pic_palettes = metrics_v0["mois_pic_palettes"]
    pic_palettes = metrics_v0["pic_palettes"]
    camions_zone = metrics_v0["camions_zone"]
    total_camions = metrics_v0["total_camions"]
    flux_camions = metrics_v0["flux_camions"]
    mois_pic_camions = metrics_v0["mois_pic_camions"]
    pic_camions = metrics_v0["pic_camions"]
    rempl_zone = metrics_v0["rempl_zone"]
    rempl_moyen = metrics_v0["rempl_moyen"]
    camions_type = metrics_v0["camions_type"]

    metrics_v1 = pipeline_avec_ccc(bg)
    total_camions_ccc = metrics_v1["total_camions"]
    flux_camions_ccc = metrics_v1["flux_camions"].rename(columns={"Camions": "Nombre de Camions CCC"}) if not metrics_v1["flux_camions"].empty else pd.DataFrame(columns=["Mois", "Nombre de Camions CCC"])
    mois_pic_camions_v1 = metrics_v1["mois_pic_camions"]
    pic_camions_v1 = metrics_v1["pic_camions"]
    rempl_zone_ccc = metrics_v1["rempl_zone"]
    rempl_moyen_ccc = metrics_v1["rempl_moyen"]
    camions_zone_ccc = metrics_v1["camions_zone"].rename(columns={"Camions": "Total CCC"}) if not metrics_v1["camions_zone"].empty else pd.DataFrame(columns=["Étage - Zone", "Total CCC"])

    # Hypothèses générales
    lot_col = param.columns[1]  # même logique que ta macro
    try:
        nb_etages = param.loc[param["Lot"] == "Nombre étage :", lot_col].iloc[0]
    except Exception:
        nb_etages = ""
    familles_identifiees = (
        materiel["Nom"].dropna().astype(str).sort_values().unique().tolist()
        if "Nom" in materiel.columns
        else []
    )

    # DPGF + indice depuis Donn?es
    dpgf_date = st.session_state.get("dpgf_date", None)
    dpgf_indice = st.session_state.get("dpgf_indice", "")
    dpgf_date_str = _format_date_fr(dpgf_date)
    planning_indice = st.session_state.get("parametrage", {}).get("planning_indice", "")

    # 4) Gestion des variantes (initialisation)
    if "variants" not in st.session_state:
        st.session_state["variants"] = {}      # {"V2": {"with_ccc": bool, "bytes": ...}}
    if "variant_counter" not in st.session_state:
        st.session_state["variant_counter"] = 2

    # 4) Onglets V0 / V1 / Variantes / Comparatif
    tabs = []

    if use_ccc:
        tabs.append("V1")
    else:
        tabs.append("V0")

    tabs.extend(["Variantes", "Comparatif"])

    tab_objects = st.tabs(tabs)
    main_tab = tab_objects[0]
    tab_var = tab_objects[1]
    tab_comp = tab_objects[2]


    with main_tab:
        if use_ccc:

            try:
                src_v1 = src.copy()
            except Exception:
                src_v1 = pd.DataFrame()

            # 3 onglets internes : Hypothèses / Palettes / Camions
            ong_hyp_v1, ong_pal_v1, ong_cam_v1 = st.tabs(
                ["📘 Hypothèses", "📦 Palettes", "🚚 Camions"]
            )

            # 📘 ONGLET HYPOTHÈSES V1
            with ong_hyp_v1:
                st.markdown("### 📘 Hypothèses")

                h1, h2 = st.columns(2)
                with h1:
                    st.markdown("#### 📄 Document de source")

                    if dpgf_date_str and dpgf_indice:
                        default_dpgf_v1 = f"DPGF indice {dpgf_indice} du {dpgf_date_str}"
                    elif dpgf_date_str:
                        default_dpgf_v1 = f"DPGF du {dpgf_date_str}"
                    elif dpgf_indice:
                        default_dpgf_v1 = f"DPGF indice {dpgf_indice}"
                    else:
                        default_dpgf_v1 = ""

                    dpgf_txt_v1 = st.text_area(
                        "DPGF + Indice :",
                        value=default_dpgf_v1,
                        key="dpgf_v1",
                        placeholder="DPGF indice ? du ?",
                    )

                    pic_file_v1 = st.file_uploader(
                        "Veuillez joindre le fichier PIC", key="pic_v1"
                    )

                    st.markdown(
                        f"- DPGF indice : **{dpgf_indice or '...'}** du **{dpgf_date_str or '...'}**"
                    )

                with h2:
                    st.markdown("#### 🕒 Hypothèse planning")
                    st.markdown(f"- Planning indice : **{planning_indice or '…'}**")

                    planning_hyp_v1 = st.text_area(
                        "Hypothèses prises pour le planning :",
                        key="planning_hyp_v1",
                        placeholder="Décrire ici les hypothèses utilisées pour le planning…",
                    )

                h3, h4 = st.columns(2)
                with h3:
                    st.markdown("#### 📄 Hypothèses de l'étude")
                    st.markdown("- regroupement du matériel en grandes catégories")
                    st.markdown("- conversion des conditionnements en équivalent palette")
                    st.markdown("- 2 phases de travaux par étage")

                with h4:
                    st.markdown("#### ⚙️ Paramètres CCC")

                    def _val_param_ccc(*labels):
                        for libel in labels:
                            try:
                                mask = param["Lot"].apply(lambda x: _norm(x) == _norm(libel))
                                val = param.loc[mask, lot_col].iloc[0]
                                if val != "":
                                    return val
                            except Exception:
                                continue
                        return ""

                    duree_ccc = _val_param_ccc(
                        "Durée de stockage CCC (en mois)",
                        "Duree de stockage CCC (en mois)",
                    )
                    tarif_mois = _val_param_ccc(
                        "Tarif mois de stockage (en €)",
                        "Tarif mois de stockage (en EUR)",
                    )
                    frais_sup = _val_param_ccc(
                        "Frais supplémentaires/palette (en €)",
                        "Frais supplementaires/palette (en EUR)",
                    )
                    frais_liv = _val_param_ccc("Frais de livraison par camion")

                    st.markdown(f"- Durée stockage : **{duree_ccc} mois**")
                    st.markdown(f"- Tarif de stockage : **{tarif_mois} €/mois**")
                    st.markdown(f"- Frais supplémentaires/palette : **{frais_sup} €**")
                    st.markdown(f"- Frais de livraison : **{frais_liv} €**")

                    with st.container():
                        st.markdown("#### Hypothèse de base déportée par famille")

                        col_el = (
                            _find_col(src_v1.columns, "Nom de l'element")
                            or _find_col(src_v1.columns, "Nom de l'élément")
                            or _find_col(src_v1.columns, "Nom de l'élement")
                        )
                        col_use = (
                            _find_col(src_v1.columns, "Utilisation d'une CCC")
                            or _find_col(src_v1.columns, "Utilisation d'un CCC")
                        )
                        col_mat = (
                            _find_col(bg.columns, "Matériel CCC")
                            or _find_col(bg.columns, "Matériel CCC")
                        )
                        col_nb = (
                            _find_col(bg.columns, "Nombre de matériels CCC")
                            or _find_col(bg.columns, "Nombre de matériels CCC")
                        )
    
                        if col_el and col_use and col_mat and col_nb:
                            df_src = src_v1[[col_el, col_use]].dropna()
                            df_src["use_ccc"] = df_src[col_use].astype(str).str.lower().isin(["oui", "yes", "y", "1"])
    
                            df_yes = (
                                df_src.groupby(col_el, as_index=False)["use_ccc"]
                                .any()
                                .rename(columns={col_el: "Famille"})
                            )
    
                            df_qty = (
                                bg[[col_mat, col_nb]]
                                .dropna(subset=[col_mat])
                                .groupby(col_mat, as_index=False)[col_nb]
                                .sum()
                                .rename(columns={col_mat: "Famille", col_nb: "Quantité"})
                            )
    
                            df_merge = pd.merge(df_yes, df_qty, on="Famille", how="left")
                            df_merge["Quantité"] = pd.to_numeric(df_merge["Quantité"], errors="coerce").fillna(0)
                            df_merge["Stocké en CCC ?"] = df_merge["use_ccc"].apply(lambda x: "✅" if x else "❌")
    
                            df_merge = _coerce_quantite_cols(df_merge)
                            df_merge = _clean_ccc_familles(df_merge)
                            display_df = df_merge[["Famille", "Stocké en CCC ?", "Quantité"]]
                            st.dataframe(
                                display_df,
                                use_container_width=True,
                                height=_df_height(len(display_df)),
                            )
                        else:
                            st.info("Colonnes nécessaires introuvables dans Tableau Source / Bilan Graphique")
            # 📦 ONGLET PALETTES (V1)
            with ong_pal_v1:

                    st.markdown("### 📦 Palettes ")

                    colA, colB = st.columns(2)
                    with colA:
                        st.metric(
                            "Palettes équivalentes totales (identiques V0/V1)",
                            f"{total_palettes:,.0f}".replace(",", " "),
                        )
                    with colB:
                        st.metric(
                            "Surface totale (m²)",
                            f"{(total_palettes * 0.96):,.0f}".replace(",", " "),
                        )

                    if _to_internal_lot(params.get("lot", "")).upper() == "GLOBAL" and "Lot" in src_v1.columns:
                        p1, p2 = st.columns(2)
                        with p1:
                            df_pal_lot_v1 = _plot_palettes_par_lot_pie(src_v1, key="pie_palettes_lot_v1")
                        with p2:
                            st.markdown("#### Détail palettes par lot")
                            if df_pal_lot_v1 is not None and not df_pal_lot_v1.empty:
                                dfl = df_pal_lot_v1.copy()
                                dfl["Pourcentage"] = dfl["Pourcentage"].map(lambda v: f"{v:.1f}%")
                                st.dataframe(dfl, use_container_width=True, hide_index=True)
                            else:
                                st.info("Aucune donnée.")

                    c1, c2 = st.columns(2)

                    # Palettes par famille (avant CCC)
                    with c1:
                        st.markdown("#### Palettes par famille")
                        col_fam = (
                            _find_col(src_v1.columns, "Nom de l'element")
                            or _find_col(src_v1.columns, "Nom de l'element")
                            or _find_col(src_v1.columns, "Nom de l'element")
                        )
                        col_pal_eq = _find_col(src_v1.columns, "Nombre palettes equivalent total")
                        if col_fam and col_pal_eq:
                            df_fam_ref = pd.DataFrame({
                                "Famille": src_v1[col_fam].dropna().astype(str).str.strip().unique()
                            })
                            df_fam_ref = df_fam_ref[df_fam_ref["Famille"] != ""]
                            df_fam_ref = df_fam_ref[
                                ~df_fam_ref["Famille"].str.lower().str.startswith("stock ccc")
                            ]

                            df_fam_pal = src_v1[[col_fam, col_pal_eq]].copy()
                            df_fam_pal[col_fam] = df_fam_pal[col_fam].astype(str).str.strip()
                            df_fam_pal[col_pal_eq] = pd.to_numeric(df_fam_pal[col_pal_eq], errors="coerce").fillna(0)
                            df_fam_pal = (
                                df_fam_pal.groupby(col_fam, as_index=False)[col_pal_eq]
                                .sum()
                                .rename(columns={col_fam: "Famille", col_pal_eq: "Palettes"})
                            )

                            df_fam_pal = (
                                df_fam_ref.merge(df_fam_pal, on="Famille", how="left")
                                .fillna({"Palettes": 0})
                                .sort_values("Palettes", ascending=False)
                            )

                            fig_fam_pal = px.bar(
                                df_fam_pal,
                                x="Palettes",
                                y="Famille",
                                orientation="h",
                                color="Palettes",
                                color_continuous_scale=["#EAF4F4", "#2A9D8F", "#1D3557"],
                                text="Palettes",
                            )
                            fig_fam_pal.update_traces(texttemplate="%{x:.0f}", textposition="outside")
                            fig_fam_pal.update_layout(
                                yaxis={"categoryorder": "total ascending"},
                                margin=dict(l=10, r=10, t=20, b=10),
                                coloraxis_showscale=False,
                                height=max(320, min(1400, 34 * len(df_fam_pal))),
                            )
                            st.plotly_chart(fig_fam_pal, key="palettes_famille_v1", use_container_width=True)
                        else:
                            st.info("Colonnes famille/palettes introuvables dans Tableau Source.")

                    # Flux palettes (identique V0)
                    with c2:
                        st.markdown("#### Flux mensuel de palettes")

                        flux_palettes_plot = flux_palettes.copy()
                        if not flux_palettes_plot.empty:
                            flux_palettes_plot["Mois"] = flux_palettes_plot["Mois"].astype(str)

                        fig_flux_pal_v1 = px.area(
                            flux_palettes_plot,
                            x="Mois",
                            y="Volume (nombre de palettes équivalentes)",
                        )

                        # PIC palettes V1
                        if pic_palettes > 0:
                            mois_pic_palettes_str = str(mois_pic_palettes)

                            fig_flux_pal_v1.add_vline(
                                x=mois_pic_palettes_str,
                                line_dash="dot",
                                line_color="red",
                            )
                            fig_flux_pal_v1.add_hline(
                                y=pic_palettes,
                                line_dash="dot",
                                line_color="red",
                            )
                            fig_flux_pal_v1.add_scatter(
                                x=[mois_pic_palettes_str],
                                y=[pic_palettes],
                                mode="markers",
                                marker=dict(color="red", size=10),
                                name="Pic",
                            )
                            fig_flux_pal_v1.add_annotation(
                                x=mois_pic_palettes_str,
                                y=pic_palettes,
                                text=f"Pic : {pic_palettes:.0f} palettes ({mois_pic_palettes_str})",
                                showarrow=True,
                                arrowhead=2,
                                ax=0,
                                ay=-40,
                                font=dict(color="red"),
                            )

                        fig_flux_pal_v1.update_layout(margin=dict(l=10, r=10, t=30, b=40))

                        st.plotly_chart(
                            fig_flux_pal_v1,
                            use_container_width=True,
                            key="flux_palettes_v1",
                        )

                    c3, c4 = st.columns(2)

                    # Matériaux stockés en CCC
                    with c3:
                        st.markdown("#### Matériaux stockés en CCC")

                        mat_col = None
                        qty_col = None
                        for c in bg.columns:
                            cname = str(c).strip().lower()
                            if cname.startswith("matériel ccc"):
                                mat_col = c
                            if cname.startswith("nombre de matériels ccc"):
                                qty_col = c

                        if mat_col and qty_col:
                            if _to_internal_lot(params.get("lot", "")).upper() == "GLOBAL" and "Lot" in src_v1.columns:
                                df_ccc_lot_v1 = _plot_ccc_par_lot_pie(src_v1, key="pie_ccc_lot_v1")
                                st.markdown("#### Détail CCC par lot")
                                if df_ccc_lot_v1 is not None and not df_ccc_lot_v1.empty:
                                    dfc = df_ccc_lot_v1.copy()
                                    dfc["Palettes"] = pd.to_numeric(dfc["Palettes"], errors="coerce").fillna(0).astype(int)
                                    dfc["Pourcentage"] = dfc["Pourcentage"].map(lambda v: f"{v:.1f}%")
                                    st.dataframe(dfc, use_container_width=True, hide_index=True)
                                else:
                                    st.info("Aucune donnée CCC.")

                            df_v1_mat = (
                                bg[[mat_col, qty_col]]
                                .dropna(subset=[mat_col])
                                .groupby(mat_col, as_index=False)[qty_col]
                                .sum()
                            )

                            st.markdown("#### Répartition des matériaux stockés en CCC")
                            fig_mat_v1 = px.bar(
                                df_v1_mat,
                                x=qty_col,
                                y=mat_col,
                                orientation="h",
                            )
                            st.plotly_chart(fig_mat_v1, key="bar_mat_v1", use_container_width=True)
                        else:
                            st.info("Colonnes Matériel CCC absentes")

                    #  Palettes par étage 
                    with c4:
                        st.markdown("#### Répartition des palettes par étage / zone")
                        x_zone = _find_col_contains(palettes_zone.columns, "etage", "zone") or palettes_zone.columns[0]
                        fig_pal_v1 = px.bar(palettes_zone, x=x_zone, y="Palettes", color="Palettes")
                        st.plotly_chart(fig_pal_v1, key="palettes_zone_v1", use_container_width=True)

            # 🚚 ONGLET CAMIONS (V1)
            with ong_cam_v1:

                    st.markdown("### 🚚 Camions")

                    colA, colB = st.columns(2)
                    colA.metric("Nombre total de camions (CCC)", f"{total_camions_ccc:,.0f}")
                    colB.metric("Remplissage moyen (CCC)", f"{rempl_moyen_ccc:.1f} %")

                    c1, c2 = st.columns(2)

                    # Camions par zone
                    with c1:
                        st.markdown("#### Camions par étage")
                        if not camions_zone_ccc.empty:
                            x_zone_ccc = _find_col_contains(camions_zone_ccc.columns, "etage", "zone") or camions_zone_ccc.columns[0]
                            fig_zone_ccc = px.bar(camions_zone_ccc, x=x_zone_ccc, y="Total CCC", color="Total CCC")
                            st.plotly_chart(fig_zone_ccc, key="camions_zone_v1", use_container_width=True)
                        else:
                            st.info("Colonnes camions CCC manquantes")

                        # Flux camions CCC
                    with c2:
                        st.markdown("#### Flux mensuel de camions")

                        if not flux_camions_ccc.empty:

                            flux_camions_plot_v1 = flux_camions_ccc.copy()
                            flux_camions_plot_v1["Nombre de Camions CCC"] = pd.to_numeric(
                                flux_camions_plot_v1["Nombre de Camions CCC"], errors="coerce"
                            ).fillna(0)

                            fig_flux_ccc = px.area(
                                flux_camions_plot_v1,
                                x="Mois",
                                y="Nombre de Camions CCC",
                            )

                            # PIC camions V1
                            if mois_pic_camions_v1:
                                fig_flux_ccc.add_vline(
                                    x=mois_pic_camions_v1,
                                    line_dash="dot",
                                    line_color="red",
                                )
                                fig_flux_ccc.add_hline(
                                    y=pic_camions_v1,
                                    line_dash="dot",
                                    line_color="red",
                                )
                                fig_flux_ccc.add_scatter(
                                    x=[mois_pic_camions_v1],
                                    y=[pic_camions_v1],
                                    mode="markers",
                                    marker=dict(color="red", size=10),
                                    name="Pic de livraison",
                                )
                                fig_flux_ccc.add_annotation(
                                    x=mois_pic_camions_v1,
                                    y=pic_camions_v1,
                                    text=f"Pic : {pic_camions_v1:.0f} camions ({mois_pic_camions_v1})",
                                    showarrow=True,
                                    arrowhead=2,
                                    ax=0,
                                    ay=-40,
                                    font=dict(color="red"),
                                )

                            fig_flux_ccc.update_layout(margin=dict(l=10, r=10, t=30, b=40))

                            st.plotly_chart(
                                fig_flux_ccc,
                                use_container_width=True,
                                key="flux_camions_ccc_v1",
                            )

                        else:
                            st.info("Aucun flux de camions CCC")

                    
                    
                    
                    
                    
                    
                    
                    

                    c3, c4 = st.columns(2)
                    # Remplissage CCC
                    with c3:
                        st.markdown("#### Remplissage par étage")
                        if not rempl_zone_ccc.empty:
                            x_zone_r = _find_col_contains(rempl_zone_ccc.columns, "etage", "zone") or rempl_zone_ccc.columns[0]
                            fig_r_ccc = px.bar(
                                rempl_zone_ccc,
                                x=x_zone_r,
                                y="Remplissage (%)",
                                color="Remplissage (%)",
                            )
                            st.plotly_chart(fig_r_ccc, key="remplissage_ccc_v1", use_container_width=True)
                        else:
                            st.info("Aucune donnée de remplissage CCC disponible")

                    # Typologie des camions (V1 – CCC, colonnes X et Y)
                    with c4:
                        st.markdown("## 🚚 Typologie des camions")

                        try:
                            df_camions_ccc = pd.DataFrame()

                            # Tentative 1: ancien Excel (colonnes X et Y par position)
                            if len(bg.columns) > 24:
                                col_type = bg.columns[23]   # colonne X
                                col_nb   = bg.columns[24]   # colonne Y
                                df_camions_ccc = (
                                    bg[[col_type, col_nb]]
                                    .dropna(subset=[col_type])
                                    .groupby(col_type, as_index=False)[col_nb]
                                    .sum()
                                    .rename(columns={
                                        col_type: "Type de Camion",
                                        col_nb: "Nombre de Camions"
                                    })
                                )

                            # Tentative 2: fallback sans Excel (calcul interne)
                            if df_camions_ccc.empty and isinstance(camions_type, pd.DataFrame) and not camions_type.empty:
                                df_camions_ccc = camions_type.copy()

                            if df_camions_ccc.empty:
                                st.info("Aucun camion (CCC) trouvé dans les colonnes X et Y.")
                            else:
                                for _, row in df_camions_ccc.iterrows():
                                    nom_camion = str(row["Type de Camion"]).strip()
                                    quantite = int(row["Nombre de Camions"])

                                    # Filtrer uniquement les camions connus
                                    if nom_camion not in daba.liste_camions:
                                        continue

                                    img_path = f"images/image_camions/{nom_camion}.png"
                                    c_img, c_nom, c_nb = st.columns([1, 3, 1])

                                    with c_img:
                                        try:
                                            st.image(img_path, width=70)
                                        except:
                                            st.write("🚚")

                                    with c_nom:
                                        st.write(f"**{nom_camion}**")

                                    with c_nb:
                                        st.write(f"**{quantite}**")

                        except Exception as e:
                            st.error(f"Erreur lecture typologie V1 (colonnes X et Y): {e}")           


        else: 

            ong_hyp, ong_pal, ong_cam = st.tabs(
                    ["📘 Hypothèses", "📦 Palettes", "🚚 Camions"]
            )

            # 📘 ONGLET HYPOTHÈSES (V0)
            with ong_hyp:
                st.markdown("### 📘 Hypothèses")

                h1, h2 = st.columns(2)
                with h1:
                    # Document de source
                    st.markdown("### 📄 Document de source")

                    # Pré-remplissage DPGF + Indice à partir de dpgf_date et planning_indice
                    if dpgf_date_str and dpgf_indice:
                        default_dpgf = f"DPGF indice {dpgf_indice} du {dpgf_date_str}"
                    elif dpgf_date_str:
                        default_dpgf = f"DPGF du {dpgf_date_str}"
                    elif dpgf_indice:
                        default_dpgf = f"DPGF indice {dpgf_indice}"
                    else:
                        default_dpgf = ""

                    dpgf_txt = st.text_area(
                        "DPGF + Indice :",
                        value=default_dpgf,
                        key="dpgf_v0",
                        placeholder="DPGF indice ? du ?"
                    )

                    st.file_uploader("Veuillez joindre le fichier PIC", key="pic_v0")

                with h2:
                    # Hypothèse planning
                    st.markdown("### 🕒 Hypothèse planning")
                    st.markdown(f"- Planning indice : **{planning_indice or '…'}**")

                    planning_hyp = st.text_area(
                        "Hypothèses prises pour planning :",
                        key="planning_hyp_v0",
                        placeholder="Décrire ici les hypothèses utilisées pour le planning…"
                    )

                    if planning_hyp.strip() == "":
                        st.markdown(
                            "<p style='color:red;'>Hypothèses planning non complétées</p>",
                            unsafe_allow_html=True,
                        )
                    else:
                        st.markdown(
                            "<p style='color:green;'>Hypothèses planning complétées</p>",
                            unsafe_allow_html=True,
                        )

                h3, h4 = st.columns(2)
                with h3:
                    # Hypothèses de l'étude
                    st.markdown("### 📄 Hypothèses de l'étude")
                    st.markdown("- regroupement du matériel en grandes catégories")
                    st.markdown(
                        "- conversion des conditionnements en équivalent palette "
                        "(palette européenne 1,2 × 0,8)"
                    )
                    st.markdown("- 2 phases de travaux par étage : Production et Terminaux")

                with h4:
                    # Hypothèse de base par famille
                    st.markdown("###  Hypothèse de base déportée par famille")

                    # Familles depuis Tableau Source ou Matériel
                    if "Nom de l'élément" in src.columns:
                        familles_src = (
                            src["Nom de l'élément"]
                            .dropna()
                            .astype(str)
                            .sort_values()
                            .unique()
                        )
                    elif "Nom" in materiel.columns:
                        familles_src = (
                            materiel["Nom"]
                            .dropna()
                            .astype(str)
                            .sort_values()
                            .unique()
                        )
                    else:
                        familles_src = []

                    if len(familles_src) == 0:
                        st.info("Aucune famille trouvée.")
                    else:
                        df_fam = pd.DataFrame({
                            "Famille": familles_src,
                            "Stocké en CCC ?": ["❌"] * len(familles_src),
                            "Quantité": [0] * len(familles_src),
                        })

                        st.dataframe(df_fam, use_container_width=True)

            # 📦 ONGLET PALETTES (V0)
            with ong_pal:
                st.markdown("### 📦 Palettes")

                # Métriques : palettes + surface
                colA, colB = st.columns(2)

                with colA:
                    st.metric(
                        "Palettes équivalentes totales (V0)",
                        f"{total_palettes:,.0f}".replace(",", " "),
                    )

                surface_totale_v0 = total_palettes * 0.96
                with colB:
                    st.metric(
                        "Surface totale occupée (m²)",
                        f"{surface_totale_v0:,.0f}".replace(",", " "),
                    )

                if _to_internal_lot(params.get("lot", "")).upper() == "GLOBAL" and "Lot" in src.columns:
                    p1_v0, p2_v0 = st.columns(2)
                    with p1_v0:
                        df_pal_lot_v0 = _plot_palettes_par_lot_pie(src, key="pie_palettes_lot_v0")
                    with p2_v0:
                        st.markdown("#### Détail palettes par lot")
                        if df_pal_lot_v0 is not None and not df_pal_lot_v0.empty:
                            dfl0 = df_pal_lot_v0.copy()
                            dfl0["Pourcentage"] = dfl0["Pourcentage"].map(lambda v: f"{v:.1f}%")
                            st.dataframe(dfl0, use_container_width=True, hide_index=True)
                        else:
                            st.info("Aucune donnée.")

                # Deux graphiques côte à côte
                c1, c2 = st.columns(2)

                # Palettes par famille (Tableau Source)
                with c1:
                    st.markdown("#### Palettes par famille")
                    col_fam = (
                        _find_col(src.columns, "Nom de l'element")
                        or _find_col(src.columns, "Nom de l'element")
                        or _find_col(src.columns, "Nom de l'element")
                    )
                    col_pal_eq = _find_col(src.columns, "Nombre palettes equivalent total")
                    if col_fam and col_pal_eq:
                        df_fam_ref = pd.DataFrame({
                            "Famille": src[col_fam].dropna().astype(str).str.strip().unique()
                        })
                        df_fam_ref = df_fam_ref[df_fam_ref["Famille"] != ""]
                        df_fam_ref = df_fam_ref[
                            ~df_fam_ref["Famille"].str.lower().str.startswith("stock ccc")
                        ]

                        df_fam_pal = src[[col_fam, col_pal_eq]].copy()
                        df_fam_pal[col_fam] = df_fam_pal[col_fam].astype(str).str.strip()
                        df_fam_pal[col_pal_eq] = pd.to_numeric(df_fam_pal[col_pal_eq], errors="coerce").fillna(0)
                        df_fam_pal = (
                            df_fam_pal.groupby(col_fam, as_index=False)[col_pal_eq]
                            .sum()
                            .rename(columns={col_fam: "Famille", col_pal_eq: "Palettes"})
                        )

                        df_fam_pal = (
                            df_fam_ref.merge(df_fam_pal, on="Famille", how="left")
                            .fillna({"Palettes": 0})
                            .sort_values("Palettes", ascending=False)
                        )

                        fig_fam_pal = px.bar(
                            df_fam_pal,
                            x="Palettes",
                            y="Famille",
                            orientation="h",
                            color="Palettes",
                            color_continuous_scale=["#EAF4F4", "#2A9D8F", "#1D3557"],
                            text="Palettes",
                        )
                        fig_fam_pal.update_traces(texttemplate="%{x:.0f}", textposition="outside")
                        fig_fam_pal.update_layout(
                            yaxis={"categoryorder": "total ascending"},
                            margin=dict(l=10, r=10, t=20, b=10),
                            coloraxis_showscale=False,
                            height=max(320, min(1400, 34 * len(df_fam_pal))),
                        )
                        st.plotly_chart(fig_fam_pal, key="palettes_famille_v0", use_container_width=True)
                    else:
                        st.info("Colonnes famille/palettes introuvables dans Tableau Source.")

                # Flux mensuel de palettes + PIC (V0)
                with c2:
                    st.markdown("#### Flux mensuel de palettes")

                    flux_palettes_plot_v0 = flux_palettes.copy()
                    if not flux_palettes_plot_v0.empty:
                        flux_palettes_plot_v0["Mois"] = flux_palettes_plot_v0["Mois"].astype(str)
                    mois_pic_palettes_str_v0 = str(mois_pic_palettes)

                    fig_flux_pal_v0 = px.area(
                        flux_palettes_plot_v0,
                        x="Mois",
                        y="Volume (nombre de palettes équivalentes)",
                    )

                    if pic_palettes > 0:
                        fig_flux_pal_v0.add_vline(
                            x=mois_pic_palettes_str_v0, line_dash="dot", line_color="red"
                        )
                        fig_flux_pal_v0.add_hline(
                            y=pic_palettes, line_dash="dot", line_color="red"
                        )
                        fig_flux_pal_v0.add_scatter(
                            x=[mois_pic_palettes_str_v0],
                            y=[pic_palettes],
                            mode="markers",
                            marker=dict(color="red", size=10),
                            name="Pic de livraison",
                        )
                        fig_flux_pal_v0.add_annotation(
                            x=mois_pic_palettes_str_v0,
                            y=pic_palettes,
                            text=(
                                f"Pic : {pic_palettes:.0f} palettes "
                                f"({mois_pic_palettes_str_v0})"
                            ),
                            showarrow=True,
                            arrowhead=2,
                            ax=0,
                            ay=-40,
                            font=dict(color="red"),
                        )

                    fig_flux_pal_v0.update_layout(
                        margin=dict(l=10, r=10, t=30, b=40)
                    )
                    st.plotly_chart(
                        fig_flux_pal_v0,
                        use_container_width=True,
                        height=320,
                        key="flux_palettes_v0",
                    )

                # Répartition par étage / zone
                c3, c4 = st.columns(2)
                with c3:
                    st.markdown("#### Répartition des palettes par étage / zone")
                    fig_pal_zone_v0 = px.bar(
                        palettes_zone,
                        x="Étage - Zone",
                        y="Palettes",
                        color="Palettes",
                    )
                    fig_pal_zone_v0.update_layout(margin=dict(l=10, r=10, t=30, b=40))
                    st.plotly_chart(
                        fig_pal_zone_v0,
                        use_container_width=True,
                        height=320,
                        key="palettes_zone_v0",
                    )
                with c4:
                    st.empty()

            # 🚚 ONGLET CAMIONS (V0)
            with ong_cam:
                st.markdown("### 🚚 Camions ")

                colA, colB = st.columns(2)
                with colA:
                    st.metric(
                        "Nombre total de camions ",
                        f"{total_camions:,.0f}".replace(",", " "),
                    )
                with colB:
                    st.metric(
                        "Remplissage moyen camions ",
                        f"{rempl_moyen:.1f} %",
                    )

                c2_, c3_ = st.columns(2)

                # CAMIONS PAR ÉTAGE / ZONE
                with c2_:
                    st.markdown("#### Camions par étage ")
                    fig_cam_zone = px.bar(
                        camions_zone,
                        x="Étage - Zone",
                        y="Camions totaux",
                        color="Camions totaux",
                    )
                    fig_cam_zone.update_layout(
                        margin=dict(l=10, r=10, t=30, b=40)
                    )
                    st.plotly_chart(
                        fig_cam_zone,
                        use_container_width=True,
                        key="cam_v0_zone",
                    )

                # FLUX MENSUEL CAMIONS (CORRIGÉ AVEC LIGNE HORIZONTALE)
                with c3_:
                    st.markdown("#### Flux mensuel de camions ")

                    # Sécurisation des données (important)
                    flux_camions_plot = flux_camions.copy()
                    flux_camions_plot["Nombre de Camions"] = pd.to_numeric(
                        flux_camions_plot["Nombre de Camions"], errors="coerce"
                    ).fillna(0)

                    fig_flux_cam = px.area(
                        flux_camions_plot,
                        x="Mois",
                        y="Nombre de Camions",
                    )

# PIC – Affichage même si pic_camions 0
                    if not flux_camions_plot.empty and mois_pic_camions:

                        # Ligne verticale
                        fig_flux_cam.add_vline(
                            x=mois_pic_camions,
                            line_dash="dot",
                            line_color="red"
                        )

                        # Ligne horizontale (toujours affichée)
                        fig_flux_cam.add_hline(
                            y=pic_camions,
                            line_dash="dot",
                            line_color="red"
                        )

                        # Point rouge
                        fig_flux_cam.add_scatter(
                            x=[mois_pic_camions],
                            y=[pic_camions],
                            mode="markers",
                            marker=dict(color="red", size=10),
                            name="Pic de livraison",
                        )

                        # Annotation
                        fig_flux_cam.add_annotation(
                            x=mois_pic_camions,
                            y=pic_camions,
                            text=f"Pic : {pic_camions:.0f} camions ({mois_pic_camions})",
                            showarrow=True,
                            arrowhead=2,
                            ax=0,
                            ay=-40,
                            font=dict(color="red"),
                        )

                    fig_flux_cam.update_layout(
                        margin=dict(l=10, r=10, t=30, b=40)
                    )

                    st.plotly_chart(
                        fig_flux_cam,
                        use_container_width=True,
                        key="cam_v0_flux_pic",
                    )

                # REMPLISSAGE PAR ÉTAGE / ZONE
                c4, c5 = st.columns(2)
                with c4:
                    st.markdown("#### Remplissage des camions par étage ")
                    if not rempl_zone.empty:
                        fig_rempl = px.bar(
                            rempl_zone,
                            x="Étage - Zone",
                            y="Remplissage (%)",
                            color="Remplissage (%)",
                            color_continuous_scale="Purples",
                        )
                        fig_rempl.update_layout(
                            margin=dict(l=10, r=10, t=30, b=40)
                        )
                        st.plotly_chart(
                            fig_rempl,
                            use_container_width=True,
                            key="cam_v0_rempl",
                        )
                    else:
                        st.info("Aucune donnée de remplissage disponible.")
                with c5:
                    # Typologie des camions (V0 – colonnes T et U)
                    st.markdown("## 🚚 Typologie des camions")

                    try:
                        # Identification stricte des colonnes T et U
                        col_type = bg.columns[19]   # colonne T
                        col_nb   = bg.columns[20]   # colonne U

                        df_camions_v0 = (
                            bg[[col_type, col_nb]]
                            .dropna(subset=[col_type])
                            .groupby(col_type, as_index=False)[col_nb]
                            .sum()
                            .rename(columns={
                                col_type: "Type de Camion",
                                col_nb:   "Nombre de Camions"
                            })
                        )

                        if df_camions_v0.empty:
                            st.info("Aucun camion trouvé dans les colonnes T et U.")
                        else:
                            for _, row in df_camions_v0.iterrows():
                                nom_camion = str(row["Type de Camion"]).strip()
                                quantite = int(row["Nombre de Camions"])

                                # Filtrer uniquement les camions connus
                                if nom_camion not in daba.liste_camions:
                                    continue

                                img_path = f"images/image_camions/{nom_camion}.png"
                                c_img, c_nom, c_nb = st.columns([1, 3, 1])

                                with c_img:
                                    try:
                                        st.image(img_path, width=70)
                                    except:
                                        st.write("🚚")

                                with c_nom:
                                    st.write(f"**{nom_camion}**")

                                with c_nb:
                                    st.write(f"**{quantite}**")

                    except Exception as e:
                        st.error(f"Erreur lecture typologie V0 (colonnes T et U): {e}")



    with tab_var:
        st.subheader("Variantes personnalisées")

        # Choix type de la nouvelle variante (même logique V0/V1)
        choix_type = st.radio(
            "Type de la nouvelle variante :",
            ["Sans CCC", "Avec CCC"],
            horizontal=True,
            key="type_variante_crea"
        )

        # CRÉATION D'UNE NOUVELLE VARIANTE
        if st.button("Créer une variante"):
            vid = f"V{st.session_state['variant_counter']}"

            # Détection du format XLSX / XLSM à partir du fichier d'origine
            import zipfile
            try:
                zip_test = zipfile.ZipFile(io.BytesIO(file_bytes))
                if any("vbaProject.bin" in f.filename for f in zip_test.filelist):
                    ext = ".xlsm"
                else:
                    ext = ".xlsx"
            except:
                st.error("Le fichier Excel d'origine est invalide.")
                st.stop()

            file_name = f"{vid}{ext}"

            # Écriture du fichier sur disque
            try:
                with open(file_name, "wb") as f:
                    f.write(file_bytes)
            except Exception as e:
                st.error(f"Erreur lors de la création de la variante : {e}")
                st.stop()

            # Ajout dans la session
            with open(file_name, "rb") as f:
                st.session_state["variants"][vid] = {
                    "with_ccc": (choix_type == "Avec CCC"),
                    "bytes": f.read(),
                    "ext": ext,
                }

            # Bouton de téléchargement immédiat
            mime = (
                "application/vnd.ms-excel" if ext == ".xlsm"
                else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            with open(file_name, "rb") as f:
                st.download_button(
                    f"Télécharger {file_name}",
                    data=f.read(),
                    file_name=file_name,
                    mime=mime,
                    key=f"download_{vid}_creation",
                )

            st.session_state["variant_counter"] += 1
            st.success(f"Variante {vid} créée.")

        # Liste des variantes existantes
        variants = st.session_state.get("variants", {})

        if not variants:
            st.info("Aucune variante pour le moment.")
        else:
            for vid, meta in variants.items():
                mois_pic_palettes_var = None

                st.markdown(
                    f"### {vid} – {'avec CCC' if meta['with_ccc'] else 'sans CCC'}"
                )

                # Bouton téléchargement du fichier de la variante
                if meta.get("bytes"):
                    st.download_button(
                        f"Télécharger {vid}{meta.get('ext', '.xlsx')}",
                        data=meta["bytes"],
                        file_name=f"{vid}{meta.get('ext', '.xlsx')}",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_{vid}_existing",
                    )

                # Upload d'un fichier Excel modifié pour cette variante
                uploaded_var = st.file_uploader(
                    f"Uploader le fichier modifié pour {vid}",
                    type=["xlsx", "xlsm"],
                    key=f"upload_{vid}",
                )

                if uploaded_var is not None:
                    meta["bytes"] = uploaded_var.read()
                    st.success(f"Fichier de {vid} mis à jour.")

                if meta.get("bytes") is None:
                    st.info("Aucun fichier chargé pour cette variante.")
                    continue

                # Lecture des feuilles du fichier VARIANTE
                try:
                    excel_io_var = io.BytesIO(meta["bytes"])
                    xls_var = pd.ExcelFile(excel_io_var)
                    bg_var = xls_var.parse("Bilan Graphique")
                    param_var = xls_var.parse("Paramétrage")
                    materiel_var = xls_var.parse("Matériel")
                    src_var = xls_var.parse("Tableau Source")
                except Exception as e:
                    st.error(f"Erreur lecture Excel pour {vid}: {e}")
                    continue

                # Application du pipeline adapté
                if meta["with_ccc"]:
                    metrics_var = pipeline_avec_ccc(bg_var)
                else:
                    metrics_var = pipeline_sans_ccc(bg_var)

                palettes_zone_var = metrics_var["palettes_zone"]
                total_palettes_var = metrics_var["total_palettes"]
                flux_palettes_var = metrics_var["flux_palettes"]
                mois_pic_palettes_var = metrics_var["mois_pic_palettes"]
                pic_palettes_var = metrics_var["pic_palettes"]
                camions_zone_var = metrics_var["camions_zone"]
                total_camions_var = metrics_var["total_camions"]
                flux_camions_var = metrics_var["flux_camions"]
                mois_pic_camions_var = metrics_var["mois_pic_camions"]
                pic_camions_var = metrics_var["pic_camions"]
                rempl_zone_var = metrics_var["rempl_zone"]
                rempl_moyen_var = metrics_var["rempl_moyen"]
                camions_type_var = metrics_var["camions_type"]

                with st.expander(f"Afficher le dashboard de {vid}", expanded=False):

                    # EXACTEMENT la même structure : Hypothèses / Palettes / Camions
                    ong_hyp_v, ong_pal_v, ong_cam_v = st.tabs(
                        ["📘 Hypothèses", "📦 Palettes", "🚚 Camions"]
                    )

                    #  VARIANTE SANS CCC  → miroir du DASHBOARD V0
                    if not meta["with_ccc"]:

                        # HYPO V0
                        with ong_hyp_v:
                            st.markdown("### 📘 Hypothèses")
                            h1, h2 = st.columns(2)
                            with h1:
                                st.markdown("#### 📄 Document de source")
                                if dpgf_date_str and dpgf_indice:
                                    default_dpgf = f"DPGF indice {dpgf_indice} du {dpgf_date_str}"
                                elif dpgf_date_str:
                                    default_dpgf = f"DPGF du {dpgf_date_str}"
                                elif dpgf_indice:
                                    default_dpgf = f"DPGF indice {dpgf_indice}"
                                else:
                                    default_dpgf = ""

                                st.text_area(
                                    "DPGF + Indice :",
                                    value=default_dpgf,
                                    key=f"dpgf_var_{vid}",
                                    placeholder="DPGF indice ? du ?",
                                )

                                st.file_uploader(
                                    "Veuillez joindre le fichier PIC",
                                    key=f"pic_var_{vid}",
                                )

                            with h2:
                                st.markdown("#### 🕒 Hypothèse planning")
                                st.markdown(f"- Planning indice : **{planning_indice or '…'}**")

                                planning_hyp_var = st.text_area(
                                    "Hypothèses prises pour planning :",
                                    key=f"planning_hyp_var_{vid}",
                                    placeholder="Décrire ici les hypothèses utilisées pour le planning…",
                                )

                                if planning_hyp_var.strip() == "":
                                    st.markdown(
                                        "<p style='color:red;'>Hypothèses planning non complétées</p>",
                                        unsafe_allow_html=True,
                                    )
                                else:
                                    st.markdown(
                                        "<p style='color:green;'>Hypothèses planning complétées</p>",
                                        unsafe_allow_html=True,
                                    )

                            h3, h4 = st.columns(2)
                            with h3:
                                # Hypothèses de l'étude (idem V0)
                                st.markdown("#### 📄 Hypothèses de l'étude")
                                st.markdown("- regroupement du matériel en grandes catégories")
                                st.markdown(
                                    "- conversion des conditionnements en équivalent palette "
                                    "(palette européenne 1,2 × 0,8)"
                                )
                                st.markdown(
                                    "- 2 phases de travaux par étage : Production et Terminaux"
                                )

                            with h4:
                                st.markdown("#### Hypothèse de base déportée par famille")

                                if "Nom de l'élément" in src_var.columns:
                                    familles_src_v = (
                                        src_var["Nom de l'élément"]
                                        .dropna()
                                        .astype(str)
                                        .sort_values()
                                        .unique()
                                    )
                                elif "Nom" in materiel_var.columns:
                                    familles_src_v = (
                                        materiel_var["Nom"]
                                        .dropna()
                                        .astype(str)
                                        .sort_values()
                                        .unique()
                                    )
                                else:
                                    familles_src_v = []
    
                                if len(familles_src_v) == 0:
                                    st.info(
                                        "Aucune famille trouvée dans Tableau Source / Matériel de la variante."
                                    )
                                else:
                                    df_fam_v = pd.DataFrame(
                                        {
                                            "Famille": familles_src_v,
                                            "Stocké en CCC ?": ["❌" for _ in familles_src_v],
                                        }
                                    )
                                    st.dataframe(df_fam_v, use_container_width=True)
    
                            # PALETTES V0 (variante)
                        with ong_pal_v:
                            st.markdown("### 📦 Palettes")

                            cA, cB = st.columns(2)
                            with cA:
                                st.metric(
                                    "Palettes équivalentes totales (Variante)",
                                    f"{total_palettes_var:,.0f}".replace(",", " "),
                                )
                            with cB:
                                st.metric(
                                    "Surface totale occupée (m²)",
                                    f"{(total_palettes_var * 0.96):,.0f}".replace(",", " "),
                                )

                            c1, c2 = st.columns(2)

                            # Répartition matériaux (on refait comme en V0 mais sur bg_var)
                            with c1:
                                st.markdown("#### Palettes par famille")
                                col_fam_v = (
                                    _find_col(src_var.columns, "Nom de l'element")
                                    or _find_col(src_var.columns, "Nom de l'element")
                                    or _find_col(src_var.columns, "Nom de l'element")
                                )
                                col_pal_eq_v = _find_col(src_var.columns, "Nombre palettes equivalent total")
                                if col_fam_v and col_pal_eq_v:
                                    df_fam_ref_v = pd.DataFrame({
                                        "Famille": src_var[col_fam_v].dropna().astype(str).str.strip().unique()
                                    })
                                    df_fam_ref_v = df_fam_ref_v[df_fam_ref_v["Famille"] != ""]
                                    df_fam_ref_v = df_fam_ref_v[
                                        ~df_fam_ref_v["Famille"].str.lower().str.startswith("stock ccc")
                                    ]

                                    df_fam_pal_v = src_var[[col_fam_v, col_pal_eq_v]].copy()
                                    df_fam_pal_v[col_fam_v] = df_fam_pal_v[col_fam_v].astype(str).str.strip()
                                    df_fam_pal_v[col_pal_eq_v] = pd.to_numeric(df_fam_pal_v[col_pal_eq_v], errors="coerce").fillna(0)
                                    df_fam_pal_v = (
                                        df_fam_pal_v.groupby(col_fam_v, as_index=False)[col_pal_eq_v]
                                        .sum()
                                        .rename(columns={col_fam_v: "Famille", col_pal_eq_v: "Palettes"})
                                    )

                                    df_fam_pal_v = (
                                        df_fam_ref_v.merge(df_fam_pal_v, on="Famille", how="left")
                                        .fillna({"Palettes": 0})
                                        .sort_values("Palettes", ascending=False)
                                    )

                                    fig_fam_pal_v = px.bar(
                                        df_fam_pal_v,
                                        x="Palettes",
                                        y="Famille",
                                        orientation="h",
                                        color="Palettes",
                                        color_continuous_scale=["#EAF4F4", "#2A9D8F", "#1D3557"],
                                        text="Palettes",
                                    )
                                    fig_fam_pal_v.update_traces(texttemplate="%{x:.0f}", textposition="outside")
                                    fig_fam_pal_v.update_layout(
                                        yaxis={"categoryorder": "total ascending"},
                                        margin=dict(l=10, r=10, t=20, b=10),
                                        coloraxis_showscale=False,
                                        height=max(320, min(1400, 34 * len(df_fam_pal_v))),
                                    )
                                    st.plotly_chart(
                                        fig_fam_pal_v,
                                        key=f"pal_fam_v0_{vid}",
                                        use_container_width=True,
                                    )
                                else:
                                    st.info("Colonnes famille/palettes introuvables dans Tableau Source.")

                            # Flux mensuel palettes (variante)
                            with c2:
                                st.markdown("#### Flux mensuel de palettes ")

                                flux_palettes_plot_v = flux_palettes_var.copy()
                                if not flux_palettes_plot_v.empty:
                                    flux_palettes_plot_v["Mois"] = flux_palettes_plot_v["Mois"].astype(str)

# 2e colonne valeur
                                if not flux_palettes_plot_v.empty:
                                    y_col_pal = [
                                        c for c in flux_palettes_plot_v.columns if c != "Mois"
                                    ][0]
                                else:
                                    y_col_pal = "Volume (nombre de palettes équivalentes)"

                                fig_flux_pal_v = px.area(
                                    flux_palettes_plot_v,
                                    x="Mois",
                                    y=y_col_pal,
                                )

                                if pic_palettes_var > 0 and mois_pic_palettes_var:
                                    mois_pic_palettes_str_v = str(mois_pic_palettes_var)
                                    fig_flux_pal_v.add_vline(
                                        x=mois_pic_palettes_str_v,
                                        line_dash="dot",
                                        line_color="red",
                                    )
                                    fig_flux_pal_v.add_hline(
                                        y=pic_palettes_var,
                                        line_dash="dot",
                                        line_color="red",
                                    )
                                    fig_flux_pal_v.add_scatter(
                                        x=[mois_pic_palettes_str_v],
                                        y=[pic_palettes_var],
                                        mode="markers",
                                        marker=dict(color="red", size=10),
                                        name="Pic de livraison",
                                    )

                                fig_flux_pal_v.update_layout(
                                    margin=dict(l=10, r=10, t=30, b=40)
                                )
                                st.plotly_chart(
                                    fig_flux_pal_v,
                                    use_container_width=True,
                                )

                            # Palettes par étage / zone
                            c3, c4 = st.columns(2)
                            with c3:
                                st.markdown("#### Répartition des palettes par étage / zone ")
                                fig_pal_zone_v = px.bar(
                                    palettes_zone_var,
                                    x="Étage - Zone",
                                    y="Palettes",
                                    color="Palettes",
                                )
                                fig_pal_zone_v.update_layout(
                                    margin=dict(l=10, r=10, t=30, b=40)
                                )
                                st.plotly_chart(
                                    fig_pal_zone_v,
                                    use_container_width=True,
                                )
                            with c4:
                                st.empty()

                        # CAMIONS V0 (variante)
                        with ong_cam_v:
                            st.markdown("### 🚚 Camions")

                            colA, colB = st.columns(2)
                            with colA:
                                st.metric(
                                    "Nombre total de camions",
                                    f"{total_camions_var:,.0f}".replace(",", " "),
                                )
                            with colB:
                                st.metric(
                                    "Remplissage moyen camions ",
                                    f"{rempl_moyen_var:.1f} %",
                                )

                            c2_, c3_ = st.columns(2)

                            # Camions par étage / zone
                            with c2_:
                                st.markdown("#### Camions par étage ")
                                fig_cam_zone_v = px.bar(
                                    camions_zone_var,
                                    x="Étage - Zone",
                                    y=camions_zone_var.columns[-1],
                                    color=camions_zone_var.columns[-1],
                                )
                                fig_cam_zone_v.update_layout(
                                    margin=dict(l=10, r=10, t=30, b=40)
                                )
                                st.plotly_chart(
                                    fig_cam_zone_v,
                                    use_container_width=True,
                                )

                            # Flux mensuel camions (variante)
                            with c3_:
                                st.markdown("#### Flux mensuel de camions ")

                                fc_v = flux_camions_var.copy()
                                if not fc_v.empty:
                                    fc_v["Mois"] = fc_v["Mois"].astype(str)
                                    # Nom de la colonne Y (peut être 'Nombre de Camions' ou 'Camions')
                                    y_candidates = [c for c in fc_v.columns if c != "Mois"]
                                    y_col_cam = y_candidates[0] if y_candidates else None
                                    if y_col_cam:
                                        fc_v[y_col_cam] = pd.to_numeric(
                                            fc_v[y_col_cam],
                                            errors="coerce",
                                        ).fillna(0)
                                else:
                                    y_col_cam = None

                                if y_col_cam:
                                    fig_flux_cam_v = px.area(
                                        fc_v,
                                        x="Mois",
                                        y=y_col_cam,
                                    )
                                else:
                                    fig_flux_cam_v = None

                                if mois_pic_camions_var:
                                    fig_flux_cam_v.add_vline(
                                        x=mois_pic_camions_var,
                                        line_dash="dot",
                                        line_color="red",
                                    )
                                    fig_flux_cam_v.add_hline(
                                        y=pic_camions_var,
                                        line_dash="dot",
                                        line_color="red",
                                    )

                                if fig_flux_cam_v is not None:
                                    fig_flux_cam_v.update_layout(
                                        margin=dict(l=10, r=10, t=30, b=40)
                                    )
                                    st.plotly_chart(
                                        fig_flux_cam_v,
                                        use_container_width=True,
                                    )
                                else:
                                    st.info("Aucun flux de camions disponible pour cette variante.")

                            # Remplissage par étage / zone
                            c4, c5 = st.columns(2)
                            with c4:
                                st.markdown("#### Remplissage des camions par étage")
                                if not rempl_zone_var.empty:
                                    fig_rempl_v = px.bar(
                                        rempl_zone_var,
                                        x="Étage - Zone",
                                        y="Remplissage (%)",
                                        color="Remplissage (%)",
                                    )
                                    fig_rempl_v.update_layout(
                                        margin=dict(l=10, r=10, t=30, b=40)
                                    )
                                    st.plotly_chart(
                                        fig_rempl_v,
                                        use_container_width=True,
                                    )
                                else:
                                    st.info("Aucune donnée de remplissage disponible pour cette variante.")
                            with c5:
                                # Typologie camions (variante) – même info que V0
                                # 🚚 Typologie des camions – Variante Sans CCC (structure V0)

                                st.markdown("## 🚚 Typologie des camions ")

                                try:
                                    # Identification stricte des colonnes T et U (comme V0)
                                    col_type = bg_var.columns[19]   # colonne T
                                    col_nb   = bg_var.columns[20]   # colonne U

                                    df_camions_var = (
                                        bg_var[[col_type, col_nb]]
                                        .dropna(subset=[col_type])
                                        .groupby(col_type, as_index=False)[col_nb]
                                        .sum()
                                        .rename(columns={
                                            col_type: "Type de Camion",
                                            col_nb:   "Nombre de Camions"
                                        })
                                    )

                                    if df_camions_var.empty:
                                        st.info("Aucun camion trouvé dans les colonnes T et U pour cette variante.")
                                    else:
                                        for _, row in df_camions_var.iterrows():
                                            nom_camion = str(row["Type de Camion"]).strip()
                                            quantite = int(row["Nombre de Camions"])

                                            # Filtrer uniquement les camions connus
                                            if nom_camion not in daba.liste_camions:
                                                continue

                                            img_path = f"images/image_camions/{nom_camion}.png"
                                            c_img, c_nom, c_nb = st.columns([1, 3, 1])

                                            with c_img:
                                                try:
                                                    st.image(img_path, width=70)
                                                except:
                                                    st.write("🚚")

                                            with c_nom:
                                                st.write(f"**{nom_camion}**")

                                            with c_nb:
                                                st.write(f"**{quantite}**")

                                except Exception as e:
                                    st.error(f"Erreur lecture typologie variante Sans CCC : {e}")

                    #  VARIANTE AVEC CCC  → miroir du DASHBOARD V1
                    else:
                        # HYPO V1 (variante)
                        with ong_hyp_v:
                            st.markdown("### 📘 Hypothèses")
                            h1, h2 = st.columns(2)
                            with h1:
                                st.markdown("#### 📄 Document de source")
                                if dpgf_date_str and dpgf_indice:
                                    default_dpgf = f"DPGF indice {dpgf_indice} du {dpgf_date_str}"
                                elif dpgf_date_str:
                                    default_dpgf = f"DPGF du {dpgf_date_str}"
                                elif dpgf_indice:
                                    default_dpgf = f"DPGF indice {dpgf_indice}"
                                else:
                                    default_dpgf = ""

                                st.text_area(
                                    "DPGF + Indice :",
                                    value=default_dpgf,
                                    key=f"dpgf_var_{vid}",
                                    placeholder="DPGF indice ? du ?",
                                )

                                st.file_uploader(
                                    "Veuillez joindre le fichier PIC",
                                    key=f"pic_var_{vid}",
                                )

                            with h2:
                                st.markdown("#### 🕒 Hypothèse planning")
                                st.markdown(f"- Planning indice : **{planning_indice or '…'}**")

                                planning_hyp_var = st.text_area(
                                    "Hypothèses prises pour planning :",
                                    key=f"planning_hyp_var_{vid}",
                                    placeholder="Décrire ici les hypothèses utilisées pour le planning…",
                                )

                                if planning_hyp_var.strip() == "":
                                    st.markdown(
                                        "<p style='color:red;'>Hypothèses planning non complétées</p>",
                                        unsafe_allow_html=True,
                                    )
                                else:
                                    st.markdown(
                                        "<p style='color:green;'>Hypothèses planning complétées</p>",
                                        unsafe_allow_html=True,
                                    )

                            h3, h4 = st.columns(2)
                            with h3:
                                st.markdown("#### 📄 Hypothèses de l'étude")
                                st.markdown("- regroupement du matériel en grandes catégories")
                                st.markdown("- conversion des conditionnements en équivalent palette")
                                st.markdown("- 2 phases de travaux par étage")

                            with h4:
                                # Paramètres CCC depuis la feuille Paramétrage de la variante
                                st.markdown("#### ⚙️ Paramètres CCC")
                                def _val_param_ccc_var(libel):
                                    try:
                                        lot_col_var = param_var.columns[1]
                                        return param_var.loc[
                                            param_var["Lot"] == libel,
                                            lot_col_var,
                                        ].iloc[0]
                                    except Exception:
                                        return ""

                                duree_ccc_v = _val_param_ccc_var("Durée de stockage CCC (en mois)")
                                tarif_mois_v = _val_param_ccc_var("Tarif mois de stockage (en €)")
                                frais_sup_v = _val_param_ccc_var("Frais supplémentaires/palette (en €)")
                                frais_liv_v = _val_param_ccc_var("Frais de livraison par camion")

                                st.markdown(f"- Durée stockage : **{duree_ccc_v} mois**")
                                st.markdown(f"- Tarif de stockage : **{tarif_mois_v} €/mois**")
                                st.markdown(f"- Frais supplémentaires/palette : **{frais_sup_v} €**")
                                st.markdown(f"- Frais de livraison : **{frais_liv_v} €**")

                            st.markdown("#### Hypothèse de base déportée par famille")

                            if (
                                {"Nom de l'élément", "Utilisation d'une CCC"}.issubset(src_var.columns)
                                and {"Matériel CCC", "Nombre de matériels CCC"}.issubset(bg_var.columns)
                            ):
                                df_src_loc = src_var[
                                    ["Nom de l'élément", "Utilisation d'une CCC"]
                                ].dropna()
                                df_src_loc["use_ccc"] = (
                                    df_src_loc["Utilisation d'une CCC"]
                                    .astype(str)
                                    .str.lower()
                                    .isin(["oui", "yes", "y", "1"])
                                )
    
                                df_yes_v = (
                                    df_src_loc.groupby("Nom de l'élément", as_index=False)["use_ccc"]
                                    .any()
                                    .rename(columns={"Nom de l'élément": "Famille"})
                                )
    
                                df_qty_v = (
                                    bg_var[["Matériel CCC", "Nombre de matériels CCC"]]
                                    .dropna(subset=["Matériel CCC"])
                                    .groupby("Matériel CCC", as_index=False)["Nombre de matériels CCC"]
                                    .sum()
                                    .rename(
                                        columns={
                                            "Matériel CCC": "Famille",
                                            "Nombre de matériels CCC": "Quantité",
                                        }
                                    )
                                )
    
                                df_merge_v = pd.merge(df_yes_v, df_qty_v, on="Famille", how="left")
                                df_merge_v["Quantité"] = pd.to_numeric(
                                    df_merge_v["Quantité"], errors="coerce"
                                ).fillna(0)
                                df_merge_v = _coerce_quantite_cols(df_merge_v)
                                df_merge_v["Stocké en CCC ?"] = df_merge_v["use_ccc"].apply(
                                    lambda x: "✔️" if x else "❌"
                                )
    
                                df_merge_v = _clean_ccc_familles(df_merge_v)
                                display_df_v = df_merge_v[["Famille", "Stocké en CCC ?", "Quantité"]]
                                st.dataframe(
                                    display_df_v,
                                    use_container_width=True,
                                    height=_df_height(len(display_df_v)),
                                )
                            else:
                                st.info(
                                    "Colonnes nécessaires introuvables dans Tableau Source / Bilan Graphique de la variante."
                                )
    
                            # PALETTES V1 (variante)
                        with ong_pal_v:
                            st.markdown("### 📦 Palettes ")

                            colA, colB = st.columns(2)
                            with colA:
                                st.metric(
                                    "Palettes équivalentes totales (Variante)",
                                    f"{total_palettes_var:,.0f}".replace(",", " "),
                                )
                            with colB:
                                st.metric(
                                    "Surface totale (m²)",
                                    f"{(total_palettes_var * 0.96):,.0f}".replace(",", " "),
                                )

                            c1, c2 = st.columns(2)

                            # Palettes par famille
                            with c1:
                                st.markdown("#### Palettes par famille")
                                col_fam_v = (
                                    _find_col(src_var.columns, "Nom de l'element")
                                    or _find_col(src_var.columns, "Nom de l'element")
                                    or _find_col(src_var.columns, "Nom de l'element")
                                )
                                col_pal_eq_v = _find_col(src_var.columns, "Nombre palettes equivalent total")
                                if col_fam_v and col_pal_eq_v:
                                    df_fam_ref_v = pd.DataFrame({
                                        "Famille": src_var[col_fam_v].dropna().astype(str).str.strip().unique()
                                    })
                                    df_fam_ref_v = df_fam_ref_v[df_fam_ref_v["Famille"] != ""]
                                    df_fam_ref_v = df_fam_ref_v[
                                        ~df_fam_ref_v["Famille"].str.lower().str.startswith("stock ccc")
                                    ]

                                    df_fam_pal_v = src_var[[col_fam_v, col_pal_eq_v]].copy()
                                    df_fam_pal_v[col_fam_v] = df_fam_pal_v[col_fam_v].astype(str).str.strip()
                                    df_fam_pal_v[col_pal_eq_v] = pd.to_numeric(df_fam_pal_v[col_pal_eq_v], errors="coerce").fillna(0)
                                    df_fam_pal_v = (
                                        df_fam_pal_v.groupby(col_fam_v, as_index=False)[col_pal_eq_v]
                                        .sum()
                                        .rename(columns={col_fam_v: "Famille", col_pal_eq_v: "Palettes"})
                                    )

                                    df_fam_pal_v = (
                                        df_fam_ref_v.merge(df_fam_pal_v, on="Famille", how="left")
                                        .fillna({"Palettes": 0})
                                        .sort_values("Palettes", ascending=False)
                                    )

                                    fig_fam_pal_v = px.bar(
                                        df_fam_pal_v,
                                        x="Palettes",
                                        y="Famille",
                                        orientation="h",
                                        color="Palettes",
                                        color_continuous_scale=["#EAF4F4", "#2A9D8F", "#1D3557"],
                                        text="Palettes",
                                    )
                                    fig_fam_pal_v.update_traces(texttemplate="%{x:.0f}", textposition="outside")
                                    fig_fam_pal_v.update_layout(
                                        yaxis={"categoryorder": "total ascending"},
                                        margin=dict(l=10, r=10, t=20, b=10),
                                        coloraxis_showscale=False,
                                        height=max(320, min(1400, 34 * len(df_fam_pal_v))),
                                    )
                                    st.plotly_chart(
                                        fig_fam_pal_v,
                                        key=f"pal_fam_v1_{vid}",
                                        use_container_width=True,
                                    )
                                else:
                                    st.info("Colonnes famille/palettes introuvables dans Tableau Source.")

                            # Flux palettes (identique logique V1)
                            with c2:
                                st.markdown("#### Flux mensuel de palettes")

                                flux_palettes_plot_v1 = flux_palettes_var.copy()
                                if not flux_palettes_plot_v1.empty:
                                    flux_palettes_plot_v1["Mois"] = flux_palettes_plot_v1[
                                        "Mois"
                                    ].astype(str)
                                    y_col_pal_v1 = [
                                        c for c in flux_palettes_plot_v1.columns if c != "Mois"
                                    ][0]
                                else:
                                    y_col_pal_v1 = "Volume (nombre de palettes équivalentes)"

                                fig_flux_pal_v1_var = px.area(
                                    flux_palettes_plot_v1,
                                    x="Mois",
                                    y=y_col_pal_v1,
                                )

                                if pic_palettes_var > 0 and mois_pic_palettes_var:
                                    mois_pic_palettes_str_v1 = str(mois_pic_palettes_var)
                                    fig_flux_pal_v1_var.add_vline(
                                        x=mois_pic_palettes_str_v1,
                                        line_dash="dot",
                                        line_color="red",
                                    )
                                    fig_flux_pal_v1_var.add_hline(
                                        y=pic_palettes_var,
                                        line_dash="dot",
                                        line_color="red",
                                    )

                                fig_flux_pal_v1_var.update_layout(
                                    margin=dict(l=10, r=10, t=30, b=40)
                                )
                                st.plotly_chart(
                                    fig_flux_pal_v1_var,
                                    use_container_width=True,
                                )

                            # Répartition palettes par étage / zone
                            c3, c4 = st.columns(2)
                            with c3:
                                st.markdown("#### Matériaux stockés en CCC")

                                mat_col_v = None
                                qty_col_v = None
                                for c in bg_var.columns:
                                    cname = str(c).lower()
                                    if cname.startswith("matériel ccc"):
                                        mat_col_v = c
                                    if cname.startswith("nombre de matériels ccc"):
                                        qty_col_v = c

                                if mat_col_v and qty_col_v:
                                    df_v1_mat_var = (
                                        bg_var[[mat_col_v, qty_col_v]]
                                        .dropna(subset=[mat_col_v])
                                        .groupby(mat_col_v, as_index=False)[qty_col_v]
                                        .sum()
                                    )

                                    st.markdown("#### Répartition des matériaux stockés en CCC")
                                    fig_mat_v1_var = px.bar(
                                        df_v1_mat_var,
                                        x=qty_col_v,
                                        y=mat_col_v,
                                        orientation="h",
                                    )
                                    st.plotly_chart(
                                        fig_mat_v1_var,
                                        use_container_width=True,
                                    )
                                else:
                                    st.info("Colonnes Matériel CCC absentes dans le BG de la variante.")
                            with c4:
                                st.markdown("#### Répartition des palettes par étage / zone")
                                fig_pal_v1_var = px.bar(
                                    palettes_zone_var,
                                    x="Étage - Zone",
                                    y="Palettes",
                                    color="Palettes",
                                )
                                st.plotly_chart(
                                    fig_pal_v1_var,
                                    use_container_width=True,
                                )

                        # CAMIONS V1 (variante)
                        with ong_cam_v:
                            st.markdown("### 🚚 Camions")

                            colA, colB = st.columns(2)
                            with colA:
                                st.metric(
                                    "Nombre total de camions (CCC – Variante)",
                                    f"{total_camions_var:,.0f}".replace(",", " "),
                                )
                            with colB:
                                st.metric(
                                    "Remplissage moyen (CCC – Variante)",
                                    f"{rempl_moyen_var:.1f} %",
                                )

                            c1, c2 = st.columns(2)

                            # Camions par étage / zone CCC
                            with c1:
                                st.markdown("#### Camions par étage")
                                if not camions_zone_var.empty:
# colonne Y 'Camions' (pipeline_avec_ccc)
                                    y_col_zone = [
                                        c for c in camions_zone_var.columns if c != "Étage - Zone"
                                    ][0]
                                    fig_zone_ccc_var = px.bar(
                                        camions_zone_var,
                                        x="Étage - Zone",
                                        y=y_col_zone,
                                        color=y_col_zone,
                                    )
                                    st.plotly_chart(
                                        fig_zone_ccc_var,
                                        use_container_width=True,
                                    )
                                else:
                                    st.info("Aucune donnée de camions CCC par zone pour cette variante.")

                            # Flux camions CCC
                            with c2:
                                st.markdown("#### Flux mensuel de camions")

                                if not flux_camions_var.empty:
                                    fc_v1 = flux_camions_var.copy()
                                    fc_v1["Mois"] = fc_v1["Mois"].astype(str)
                                    y_col_cam_v1 = [
                                        c for c in fc_v1.columns if c != "Mois"
                                    ][0]
                                    fc_v1[y_col_cam_v1] = pd.to_numeric(
                                        fc_v1[y_col_cam_v1],
                                        errors="coerce",
                                    ).fillna(0)

                                    fig_flux_ccc_v = px.area(
                                        fc_v1,
                                        x="Mois",
                                        y=y_col_cam_v1,
                                    )

                                    if mois_pic_camions_var:
                                        fig_flux_ccc_v.add_vline(
                                            x=mois_pic_camions_var,
                                            line_dash="dot",
                                            line_color="red",
                                        )
                                        fig_flux_ccc_v.add_hline(
                                            y=pic_camions_var,
                                            line_dash="dot",
                                            line_color="red",
                                        )

                                    fig_flux_ccc_v.update_layout(
                                        margin=dict(l=10, r=10, t=30, b=40)
                                    )
                                    st.plotly_chart(
                                        fig_flux_ccc_v,
                                        use_container_width=True,
                                    )
                                else:
                                    st.info("Aucun flux de camions CCC pour cette variante.")

                            c3, c4 = st.columns(2)
                            with c3:
                                # Remplissage CCC
                                st.markdown("#### Remplissage par étage")
                                if not rempl_zone_var.empty:
                                    fig_r_ccc_v = px.bar(
                                        rempl_zone_var,
                                        x="Étage - Zone",
                                        y="Remplissage (%)",
                                        color="Remplissage (%)",
                                    )
                                    st.plotly_chart(
                                        fig_r_ccc_v,
                                        use_container_width=True,
                                    )
                                else:
                                    st.info("Aucune donnée de remplissage CCC pour cette variante.")

                            with c4:
                                # Typologie camions CCC (variante)
                                # 🚚 Typologie des camions – Variante Avec CCC (structure V1)

                                st.markdown("## 🚚 Typologie des camions")

                                try:
                                    # Colonnes X et Y index 23 et 24 (comme V1)
                                    col_type = bg_var.columns[23]   # colonne X
                                    col_nb   = bg_var.columns[24]   # colonne Y

                                    df_camions_ccc_var = (
                                        bg_var[[col_type, col_nb]]
                                        .dropna(subset=[col_type])
                                        .groupby(col_type, as_index=False)[col_nb]
                                        .sum()
                                        .rename(columns={
                                            col_type: "Type de Camion",
                                            col_nb:   "Nombre de Camions"
                                        })
                                    )

                                    if df_camions_ccc_var.empty:
                                        st.info("Aucun camion (CCC) trouvé dans les colonnes X et Y pour cette variante.")
                                    else:
                                        for _, row in df_camions_ccc_var.iterrows():
                                            nom_camion = str(row["Type de Camion"]).strip()
                                            quantite = int(row["Nombre de Camions"])

                                            # Filtrer uniquement les camions connus
                                            if nom_camion not in daba.liste_camions:
                                                continue

                                            img_path = f"images/image_camions/{nom_camion}.png"
                                            c_img, c_nom, c_nb = st.columns([1, 3, 1])

                                            with c_img:
                                                try:
                                                    st.image(img_path, width=70)
                                                except:
                                                    st.write("🚚")

                                            with c_nom:
                                                st.write(f"**{nom_camion}**")

                                            with c_nb:
                                                st.write(f"**{quantite}**")

                                except Exception as e:
                                    st.error(f"Erreur lecture typologie variante Avec CCC : {e}")



        with tab_comp:
            st.subheader("Comparatif multi-versions")

            # 1) Construire la liste de toutes les versions
            all_versions = {
                "V0": {"with_ccc": False, "source": "base"},
                "V1": {"with_ccc": True,  "source": "base"},
            }

            for vid, meta in st.session_state.get("variants", {}).items():
                all_versions[vid] = {
                    "with_ccc": meta["with_ccc"],
                    "source": "variant",
                    "bytes": meta["bytes"],
                }

            version_names = list(all_versions.keys())

            if len(version_names) < 1:
                st.info("Aucune version disponible.")
                st.stop()

            # 2) Choix des versions à comparer (multi-sélection)
            selected_versions = st.multiselect(
                "Choisir les versions à comparer",
                version_names,
                default=[v for v in ["V0", "V1"] if v in version_names],
                key="comp_versions_multiselect",
            )

            if len(selected_versions) == 0:
                st.info("Sélectionner au moins une version.")
                st.stop()


            def compute_metrics_for_version(vname, info):
                # Charger les bonnes feuilles
                if info["source"] == "base":
                    bg_loc = _fix_df_columns(bg)
                    param_loc = param
                    src_loc = src
                else:
                    try:
                        excel_io_loc = io.BytesIO(info["bytes"])
                        xls_loc = pd.ExcelFile(excel_io_loc)
                        bg_loc = _fix_df_columns(xls_loc.parse("Bilan Graphique"))
                        param_loc = xls_loc.parse("Paramétrage")
                        src_loc = xls_loc.parse("Tableau Source")
                    except Exception:
                        return {"ok": False}

                col_zone = _find_col(bg_loc.columns, "Étage - Zone") or _find_col_contains(bg_loc.columns, "etage", "zone")

                # Palettes
                try:
                    if not col_zone:
                        raise KeyError("Étage - Zone")
                    pz = (
                        bg_loc[[col_zone, "Production", "Terminaux"]]
                        .dropna(subset=[col_zone])
                        .copy()
                        .rename(columns={col_zone: "Étage - Zone"})
                    )
                    pz["Production"] = pz["Production"].fillna(0)
                    pz["Terminaux"] = pz["Terminaux"].fillna(0)
                    pz["Palettes"] = pz["Production"] + pz["Terminaux"]
                    total_pal = float(pz["Palettes"].sum())
                except Exception:
                    pz = pd.DataFrame()
                    total_pal = 0.0

                # Flux palettes
                if {"Mois", "Volume (nombre de palettes équivalentes)"}.issubset(bg_loc.columns):
                    flux_pal = (
                        bg_loc[["Mois", "Volume (nombre de palettes équivalentes)"]]
                        .dropna(subset=["Mois"])
                        .groupby("Mois", as_index=False)
                        .sum()
                    )
                else:
                    flux_pal = pd.DataFrame(columns=["Mois", "Volume (nombre de palettes équivalentes)"])

                # Camions
                if info["with_ccc"]:
                    col_cam_total = "Nombre de Camions CCC"
                    col_cam_prod = "Camions Production avec CCC"
                    col_cam_term = "Camions Terminaux avec CCC"
                    col_rempl = "Remplissage camions avec CCC"
                else:
                    col_cam_total = "Nombre de Camions"
                    col_cam_prod = "Camions Production sans CCC"
                    col_cam_term = "Camions Terminaux sans CCC"
                    col_rempl = "Remplissage camions sans CCC"

                # Total camions
                if col_cam_total in bg_loc.columns:
                    total_cam = float(bg_loc[col_cam_total].fillna(0).sum())
                    flux_cam = (
                        bg_loc[["Mois", col_cam_total]]
                        .dropna(subset=["Mois"])
                        .groupby("Mois", as_index=False)
                        .sum()
                        .rename(columns={col_cam_total: "Camions"})
                    )
                else:
                    total_cam = 0.0
                    flux_cam = pd.DataFrame(columns=["Mois", "Camions"])

                # Camions zone
                if col_zone and {col_cam_prod, col_cam_term}.issubset(bg_loc.columns):
                    cz = (
                        bg_loc[[col_zone, col_cam_prod, col_cam_term]]
                        .dropna(subset=[col_zone])
                        .copy()
                        .rename(columns={col_zone: "Étage - Zone"})
                    )
                    cz["Camions"] = cz[col_cam_prod].fillna(0) + cz[col_cam_term].fillna(0)
                else:
                    cz = pd.DataFrame(columns=["Étage - Zone", "Camions"])

                # Remplissage
                if col_zone and col_rempl in bg_loc.columns:
                    rz = (
                        bg_loc[[col_zone, col_rempl]]
                        .dropna(subset=[col_zone])
                        .copy()
                        .rename(columns={col_zone: "Étage - Zone"})
                    )
                    rz["Remplissage (%)"] = rz[col_rempl].fillna(0)
                    rbrut = bg_loc[col_rempl].dropna()
                    rmoy = float(rbrut.mean()) if not rbrut.empty else 0.0
                else:
                    rz = pd.DataFrame(columns=["Étage - Zone", "Remplissage (%)"])
                    rmoy = 0.0

                # Coût total logistique
                if "Coût total logistique" in bg_loc.columns:
                    cout_total = float(bg_loc["Coût total logistique"].fillna(0).sum())
                else:
                    cout_total = None

                return {
                    "ok": True,
                    "with_ccc": info["with_ccc"],
                    "bg": bg_loc,
                    "param": param_loc,
                    "src": src_loc,
                    "palettes_zone": pz,
                    "flux_palettes": flux_pal,
                    "total_palettes": total_pal,
                    "camions_zone": cz,
                    "total_camions": total_cam,
                    "flux_camions": flux_cam,
                    "rempl_zone": rz,
                    "rempl_moyen": rmoy,
                    "cout_total": cout_total,
                }

            # 4) Calcul des données pour toutes les versions sélectionnées

            data_versions = {}
            for v in selected_versions:
                data_versions[v] = compute_metrics_for_version(v, all_versions[v])

            if any(not dv["ok"] for dv in data_versions.values()):
                st.error("Impossible de lire les données pour au moins une version.")
                st.stop()

            st.markdown("### 🧩 Performance & Coûts CCC")  # titre plus petit

            def read_ccc_from_bytes(excel_bytes: bytes) -> dict:
                import zipfile
                if not isinstance(excel_bytes, (bytes, bytearray)) or not excel_bytes:
                    return {}
                bio = io.BytesIO(excel_bytes)
                if not zipfile.is_zipfile(bio):
                    return {}
                bio.seek(0)
                try:
                    wb = load_workbook(filename=bio, data_only=True, keep_vba=True)
                except Exception:
                    return {}
                if "Bilan Graphique" not in wb.sheetnames:
                    return {}
                ws = wb["Bilan Graphique"]

                labels = [ws.cell(row=1, column=c).value for c in range(31, 37)]
                values = [ws.cell(row=2, column=c).value for c in range(31, 37)]

                out = {}
                for k, v in zip(labels, values):
                    if k is None:
                        continue
                    out[str(k).strip()] = v
                return out

            def read_ccc_from_bg(bg_df: pd.DataFrame) -> dict:
                if bg_df is None or bg_df.empty:
                    return {}
                labels = [
                    "% Stock CCC",
                    "% réduction Camions",
                    "% remplissage moyen des camions",
                    "Coût CCC stockage",
                    "Coût CCC livraison",
                    "Coût CCC Total",
                ]
                out = {}
                for lab in labels:
                    col = _find_col(bg_df.columns, lab)
                    if not col:
                        continue
                    try:
                        out[lab] = bg_df.loc[0, col]
                    except Exception:
                        continue
                return out

            def fmt_percent(x):
                try:
                    return f"{x * 100:.0f} %"
                except Exception:
                    return "—"

            def fmt_signed_percent(x):
                try:
                    sign = "+" if x >= 0 else "-"
                    return f"{sign}{abs(x) * 100:.0f} %"
                except Exception:
                    return "—"

            def fmt_arrow_percent(x, positive_is_good: bool) -> str:
                try:
                    val = float(x) * 100
                except Exception:
                    return "<span style='color:#9AA0A6'>—</span>"
                is_pos = val >= 0
                arrow = "↑" if is_pos else "↓"
                sign = "+" if is_pos else "-"
                good = is_pos if positive_is_good else not is_pos
                color = "#0F9D58" if good else "#DB4437"
                return f"<span style='color:{color}; font-weight:600;'>{arrow} {sign}{abs(val):.0f} %</span>"

            def fmt_euro(x):
                try:
                    return f"{int(round(x)):,}".replace(",", " ") + " €"
                except Exception:
                    return "—"


            for v in selected_versions:
                info = all_versions[v]

                if not info["with_ccc"]:
                    continue

                excel_bytes_v = file_bytes if info["source"] == "base" else info.get("bytes")
                ccc = read_ccc_from_bytes(excel_bytes_v)
                if not ccc:
                    ccc = read_ccc_from_bg(data_versions[v]["bg"])

                if not ccc:
                    st.info(f"{v} : pas de données CCC disponibles.")
                    st.markdown("---")
                    continue

                st.markdown(f"#### {v}")

                col1, col2, col3 = st.columns(3)

                with col1:
                    with st.container(border=True):
                        st.markdown("<span style='color:gray'>% Stock CCC</span>", unsafe_allow_html=True)
                        st.markdown(f"<h3>{fmt_percent(ccc.get('% Stock CCC'))}</h3>", unsafe_allow_html=True)

                with col2:
                    with st.container(border=True):
                        st.markdown("**KPI camions**")
                        st.markdown("<span style='color:gray'>Réduction camions</span>", unsafe_allow_html=True)
                        st.markdown(
                            fmt_arrow_percent(ccc.get("% réduction Camions"), positive_is_good=False),
                            unsafe_allow_html=True,
                        )
                        st.markdown("<span style='color:gray'>Remplissage moyen des camions</span>", unsafe_allow_html=True)
                        st.markdown(
                            fmt_arrow_percent(ccc.get("% remplissage moyen des camions"), positive_is_good=True),
                            unsafe_allow_html=True,
                        )

                with col3:
                    with st.container(border=True):
                        st.markdown("**KPI Coûts**")

                        c_cost1, c_cost2 = st.columns(2)
                        with c_cost1:
                            st.markdown("<span style='color:gray'>Stockage</span>", unsafe_allow_html=True)
                            st.markdown(
                                f"<h3>{fmt_euro(ccc.get('Coût CCC stockage'))}</h3>",
                                unsafe_allow_html=True
                            )
                        with c_cost2:
                            st.markdown("<span style='color:gray'>Livraison</span>", unsafe_allow_html=True)
                            st.markdown(
                                f"<h3>{fmt_euro(ccc.get('Coût CCC livraison'))}</h3>",
                                unsafe_allow_html=True
                            )

                        st.markdown("<span style='color:gray'>Total</span>", unsafe_allow_html=True)
                        st.markdown(
                            f"<h2 style='color:#0F9D58'>{fmt_euro(ccc.get('Coût CCC Total'))}</h2>",
                            unsafe_allow_html=True
                        )

            # ensuite viennent les onglets
            ong_hyp_comp, ong_pal_comp, ong_cam_comp = st.tabs(["📘 Hypothèses", "📦 Palettes", "🚚 Camions"])

            with ong_hyp_comp:
                h1, h2 = st.columns(2)
                with h1:
                    st.markdown("### 📘 Hypothèses de l'étude")
                    st.markdown("- regroupement du matériel en grandes catégories")
                    st.markdown("- conversion en équivalent palette (1,2 × 0,8 m)")
                    st.markdown("- 2 phases par étage : Production & Terminaux")

                with h2:
                    # Paramètres CCC par version
                    st.markdown("### ⚙️ Paramètres CCC par version")
                    for v, dv in data_versions.items():
                        if not dv["with_ccc"]:
                            continue

                        st.markdown(f"#### {v} – Avec CCC")

                        def _get_param_ccc(dv_local, *labels):
                            for libel in labels:
                                try:
                                    col = dv_local["param"].columns[1]
                                    mask = dv_local["param"]["Lot"].apply(lambda x: _norm(x) == _norm(libel))
                                    val = dv_local["param"].loc[mask, col].iloc[0]
                                    if val != "":
                                        return val
                                except Exception:
                                    continue
                            return ""

                        st.markdown(
                            f"- Durée de stockage CCC : **{_get_param_ccc(dv, 'Durée de stockage CCC (en mois)', 'Duree de stockage CCC (en mois)')} mois**"
                        )
                        st.markdown(
                            f"- Tarif mois de stockage : **{_get_param_ccc(dv, 'Tarif mois de stockage (en €)', 'Tarif mois de stockage (en EUR)')} €**"
                        )
                        st.markdown(
                            f"- Frais supplémentaires/palette : **{_get_param_ccc(dv, 'Frais supplémentaires/palette (en €)', 'Frais supplementaires/palette (en EUR)')} €**"
                        )
                        st.markdown(
                            f"- Frais de livraison par camion : **{_get_param_ccc(dv, 'Frais de livraison par camion')} €**"
                        )

                st.markdown("---")

                st.markdown("### 📦 Hypothèse de base déportée par famille ")

                # df_final existe déjà avec la colonne "Famille"
                # On le reconstruit proprement à partir de ce qui est déjà calculé

                familles = sorted(
                    set(
                        bg["Matériel CCC"].dropna().astype(str)
                        if "Matériel CCC" in bg.columns
                        else []
                    )
                    | set(
                        src["Nom de l'élément"].dropna().astype(str)
                        if "Nom de l'élément" in src.columns
                        else []
                    )
                )

                df_final = pd.DataFrame({"Famille": familles})
                df_final = _clean_ccc_familles(df_final)

                for v, dv in data_versions.items():

                    qty_col = f"Quantité_{v}"
                    flag_col = f"Stocké en CCC ? ({v})"

                    if v == "V0" or not dv["with_ccc"]:
                        if {"Désignation", "Production", "Terminaux"}.issubset(dv["bg"].columns):
                            df_qty = (
                                dv["bg"][["Désignation", "Production", "Terminaux"]]
                                .dropna(subset=["Désignation"])
                                .assign(
                                    Quantite=lambda x: x["Production"].fillna(0)
                                    + x["Terminaux"].fillna(0)
                                )
                                .groupby("Désignation", as_index=False)["Quantite"]
                                .sum()
                                .rename(columns={"Désignation": "Famille", "Quantite": "Quantité"})
                            )
                        else:
                            df_qty = pd.DataFrame(columns=["Famille", "Quantité"])
                    else:
                        if {"Matériel CCC", "Nombre de matériels CCC"}.issubset(dv["bg"].columns):
                            df_qty = (
                                dv["bg"][["Matériel CCC", "Nombre de matériels CCC"]]
                                .dropna(subset=["Matériel CCC"])
                                .groupby("Matériel CCC", as_index=False)["Nombre de matériels CCC"]
                                .sum()
                                .rename(columns={
                                    "Matériel CCC": "Famille",
                                    "Nombre de matériels CCC": "Quantité"
                                })
                            )
                        else:
                            df_qty = pd.DataFrame(columns=["Famille", "Quantité"])

                    df_final = df_final.merge(df_qty, on="Famille", how="left")
                    df_final.rename(columns={"Quantité": qty_col}, inplace=True)
                    df_final[qty_col] = (
                        pd.to_numeric(df_final[qty_col], errors="coerce")
                        .fillna(0)
                        .astype(int)
                    )
                    df_final[flag_col] = df_final[qty_col].apply(lambda x: "✔️" if x > 0 else "❌")

                # 🔥 CETTE LIGNE EST OBLIGATOIRE
                df_final = _coerce_quantite_cols(df_final)
                st.dataframe(
                    df_final,
                    use_container_width=True,
                    height=_df_height(len(df_final)),
                )

                                                


            with ong_pal_comp:
                def _norm_local(s: str) -> str:
                    import unicodedata

                    v = "" if s is None else str(s)
                    v = unicodedata.normalize("NFKD", v)
                    v = "".join(c for c in v if not unicodedata.combining(c))
                    return v.lower().strip()

                def _find_col_contains_local(columns, *tokens: str):
                    tokens_norm = [_norm_local(t) for t in tokens]
                    for col in columns:
                        col_norm = _norm_local(col)
                        if all(t in col_norm for t in tokens_norm):
                            return col
                    return None

                st.markdown("### 📦 Comparaison des palettes")

                # Totaux par version
                cols_tot = st.columns(len(selected_versions))
                for i, v in enumerate(selected_versions):
                    dv = data_versions[v]
                    with cols_tot[i]:
                        st.metric(
                            f"Palettes totales – {v}",
                            f"{dv['total_palettes']:,.0f}".replace(",", " "),
                        )

                pal_left, pal_right = st.columns(2)

                # Palettes par étage / zone
                with pal_left:
                    pal_frames = [
                        dv["palettes_zone"].assign(Version=v)
                        for v, dv in data_versions.items()
                        if "palettes_zone" in dv and not dv["palettes_zone"].empty
                    ]
                    df_pal = pd.concat(pal_frames, ignore_index=True) if pal_frames else pd.DataFrame()

                    if not df_pal.empty:
                        x_zone = (
                            _find_col_contains_local(df_pal.columns, "etage", "zone")
                            or _find_col_contains_local(df_pal.columns, "etage")
                            or _find_col_contains_local(df_pal.columns, "zone")
                        )
                        y_pal = _find_col_contains_local(df_pal.columns, "palette") or "Palettes"
                        if x_zone is None or y_pal not in df_pal.columns:
                            st.info("Colonnes nécessaires introuvables pour l'histogramme palettes.")
                        else:
                            fig_pal = px.bar(
                                df_pal,
                                x=x_zone,
                                y=y_pal,
                                color="Version",
                                barmode="group",
                                title="Palettes par étage / zone",
                            )
                            st.plotly_chart(fig_pal, use_container_width=True)
                    else:
                        st.info("Aucune donnée palettes pour ces versions.")

                # Flux palettes
                with pal_right:
                    flux_frames = [
                        dv["flux_palettes"].assign(Version=v)
                        for v, dv in data_versions.items()
                        if "flux_palettes" in dv and not dv["flux_palettes"].empty
                    ]
                    df_flux_pal = pd.concat(flux_frames, ignore_index=True) if flux_frames else pd.DataFrame()

                    if not df_flux_pal.empty:
                        x_mois = _find_col_contains_local(df_flux_pal.columns, "mois") or "Mois"
                        y_vol = (
                            _find_col_contains_local(df_flux_pal.columns, "volume", "palette")
                            or _find_col_contains_local(df_flux_pal.columns, "palette")
                        )
                        if x_mois not in df_flux_pal.columns or y_vol is None or y_vol not in df_flux_pal.columns:
                            st.info("Colonnes nécessaires introuvables pour le flux palettes.")
                        else:
                            fig_flux_pal = px.line(
                                df_flux_pal,
                                x=x_mois,
                                y=y_vol,
                                color="Version",
                                title="Flux mensuel de palettes",
                            )
                            st.plotly_chart(fig_flux_pal, use_container_width=True)
                    else:
                        st.info("Aucun flux palettes pour ces versions.")

            with ong_cam_comp:
                st.markdown("### 🚚 Comparaison des camions")

                # Totaux + remplissage
                cols_cam = st.columns(len(selected_versions))
                for i, v in enumerate(selected_versions):
                    dv = data_versions[v]
                    with cols_cam[i]:
                        st.metric(
                            f"Camions totaux – {v}",
                            f"{dv['total_camions']:,.0f}".replace(",", " "),
                        )
                        st.metric(
                            f"Remplissage moyen – {v}",
                            f"{dv['rempl_moyen']:.1f} %",
                        )

                st.markdown("---")
                cam_left, cam_right = st.columns(2)

                with cam_left:
                    st.markdown("### 🚚 Camions par étage / zone")

                    cam_frames = [
                        dv["camions_zone"].assign(Version=v)
                        for v, dv in data_versions.items()
                        if not dv["camions_zone"].empty
                    ]
                    df_cam = pd.concat(cam_frames, ignore_index=True) if cam_frames else pd.DataFrame()

                    if not df_cam.empty:
                        fig_cam = px.bar(
                            df_cam,
                            x="Étage - Zone",
                            y="Camions",
                            color="Version",
                            barmode="group",
                            title="Camions par étage / zone",
                        )
                        st.plotly_chart(fig_cam, use_container_width=True)
                    else:
                        st.info("Aucune donnée camions par étage pour ces versions.")

                with cam_right:
                    st.markdown("### 📈 Flux mensuel de camions")

                    df_flux_cam = pd.concat(
                        [
                            dv["flux_camions"].assign(Version=v)
                            for v, dv in data_versions.items()
                            if not dv["flux_camions"].empty
                        ],
                        ignore_index=True,
                    )

                    if not df_flux_cam.empty:
                        fig_flux_cam = px.line(
                            df_flux_cam,
                            x="Mois",
                            y="Camions",
                            color="Version",
                            title="Flux mensuel de camions",
                        )
                        st.plotly_chart(fig_flux_cam, use_container_width=True)
                    else:
                        st.info("Aucun flux camions pour ces versions.")

                st.markdown("---")
                rem_left, rem_right = st.columns(2)
                with rem_left:
                    st.markdown("### 📦 Remplissage des camions par étage / zone")

                    df_rempl = pd.concat(
                        [
                            dv["rempl_zone"].assign(Version=v)
                            for v, dv in data_versions.items()
                            if not dv["rempl_zone"].empty
                        ],
                        ignore_index=True,
                    )

                    if not df_rempl.empty:
                        fig_rempl = px.bar(
                            df_rempl,
                            x="Étage - Zone",
                            y="Remplissage (%)",
                            color="Version",
                            barmode="group",
                            title="Remplissage des camions par étage / zone",
                        )
                        st.plotly_chart(fig_rempl, use_container_width=True)
                    else:
                        st.info("Aucune donnée de remplissage disponible pour ces versions.")
                with rem_right:
                    st.empty()












# Onglet 4 : Entraînement modèles :
elif menu == "Entraînement modèles":
    st.header("Entraînement des modèles")
    st.markdown("""
    ## 
    
    Cette section permet d'entraîner les modèles utilisés pour classer les éléments dans les bordereaux.

    **Étape 1** : Préparer le fichier d'entraînement  
    - Téléchargez le fichier `template_entraînement.xlsx` ci-dessous.  
    - Remplissez les colonnes :
        - **Désignation** : Liste des matériels utilisés dans les bordereaux.  
        - **Catégorie Prédite** : Catégorie correspondante à chaque matériel.  

    **Étape 2** : Choisir le lot d'entraînement  
    **Étape 3** : Charger et lancer l'entraînement  
    """)

    # Liste des modèles disponibles
    models = [m.replace(".pkl", "") for m in os.listdir("models") if m.endswith(".pkl")]
    models = ["TCE" if m == "GLOBAL" else m for m in models]
    if "TCE" not in models:
        models.insert(0, "TCE")

    model_name = st.selectbox("Étape 2 : Choisir le lot d'entraînement", models)

    train_file = st.file_uploader("Déposez un fichier d'entraînement (xlsx)")

    if st.button("Lancer entraînement") and train_file:
        msg = entmod.train_model(train_file, _to_internal_lot(model_name))
        st.success(msg)

# Onglet 5 : Base de données : 

elif menu == "Base de données":
    st.header("🗄️ Base de données SQLite")

    # Choix table + lot
    table_choice = st.selectbox("Choisir une table", ["Matériel", "Conditionnement", "Camion"])
    models = [m.replace(".pkl", "") for m in os.listdir("models") if m.endswith(".pkl")]
    models = ["TCE" if m == "GLOBAL" else m for m in models]
    if "TCE" not in models:
        models.insert(0, "TCE")
    lot_choice = st.selectbox("Modèle", models)
    lot_choice_internal = _to_internal_lot(lot_choice)

    # Charger les données à la demande (comme Gradio)
    if "db_table_choice" not in st.session_state:
        st.session_state["db_table_choice"] = None
    if "db_lot_choice" not in st.session_state:
        st.session_state["db_lot_choice"] = None
    if "db_df" not in st.session_state:
        st.session_state["db_df"] = None

    if st.button("Afficher les données"):
        st.session_state["db_df"] = daba.afficher_donnees(table_choice, lot_choice_internal)
        st.session_state["db_table_choice"] = table_choice
        st.session_state["db_lot_choice"] = lot_choice_internal

    st.subheader("📝 Modifier la table")

    if st.session_state["db_df"] is None:
        st.info("Cliquez sur 'Afficher les données' pour charger la table.")
        st.stop()

    if (st.session_state["db_table_choice"] != table_choice
            or st.session_state["db_lot_choice"] != lot_choice_internal):
        st.warning("La table affichée ne correspond pas aux sélections actuelles. Rechargez les données.")

    df_modifie = st.data_editor(
        st.session_state["db_df"],
        use_container_width=True,
        num_rows="dynamic",
        key="crud_editor"
    )

    if st.button("💾 Enregistrer les modifications"):
        msg = daba.enregistrer_modifications(table_choice, df_modifie, lot_choice_internal)
        st.success(msg)
        st.session_state["db_df"] = daba.afficher_donnees(table_choice, lot_choice_internal)
        st.rerun()


